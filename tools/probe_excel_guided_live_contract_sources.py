#!/usr/bin/env python
"""C13.2-P1 Excel-guided live ETABS source verification gate.

Verification-only utility. It uses an ETABS Excel export/table inventory as a
probe target list, then optionally checks those targets against a live ETABS
model. Excel is inventory only: it can never create VERIFIED_LIVE status.

This tool intentionally does not edit catalogs/schemas, resolve features, run
CheckEngine, emit CheckResult, or implement engineering checks.
"""
from __future__ import annotations

import argparse
import json
import re
import sys
from dataclasses import dataclass
from pathlib import Path
from typing import Any, Mapping, Sequence

ROOT = Path(__file__).resolve().parents[1]
if str(ROOT) not in sys.path:
    sys.path.insert(0, str(ROOT))

READINESS_STATUSES = {
    "VERIFIED_LIVE",
    "NEEDS_LIVE_PROBE",
    "EXCEL_INVENTORY_ONLY",
    "SEMANTIC_REVIEW",
    "PLANNED",
}
MATCH_QUALITIES = {
    "EXACT_TABLE_AND_HEADERS",
    "EXACT_TABLE_HEADER_PARTIAL",
    "KEYWORD_TABLE_HEADER_MATCH",
    "LIVE_TABLE_NOT_FOUND",
    "EXCEL_ONLY",
    "SEMANTIC_REVIEW_REQUIRED",
}
WEAK_ONE_WORD_KEYWORDS = {
    "summary",
    "material",
    "area",
    "wall",
    "drift",
    "forces",
    "properties",
    "assignment",
    "assignments",
    "design",
    "table",
    "data",
}

HEADER_ALIASES: dict[str, tuple[str, ...]] = {
    "UniqueName": ("UniqueName", "Unique Name", "GUID"),
    "Label": ("Label", "Frame", "Beam", "Column", "Object Label", "Object"),
    "Story": ("Story", "Story Name"),
    "Type": ("Type", "Design Type", "Object Type", "Frame Type"),
    "DesignSect": ("DesignSect", "Design Section", "DesignSec", "Design Sect"),
    "AnalysisSect": ("AnalysisSect", "Analysis Section", "AnalysisSec", "Analysis Sect"),
    "Length": ("Length", "Len"),
    "Name": ("Name", "SectionName", "Section Name", "Property", "PropName"),
    "t2": ("t2", "T2", "Width", "B", "b", "bw"),
    "t3": ("t3", "T3", "Depth", "H", "h", "d"),
    "Material": ("Material", "Mat", "Material Name"),
    "Case": ("Case", "OutputCase", "Output Case", "Load Case", "LoadCase"),
    "OutputCase": ("OutputCase", "Output Case", "Case", "Load Case", "LoadCase"),
    "Mode": ("Mode", "Mode Number"),
    "Period": ("Period", "Period sec", "Period (sec)", "T"),
    "UX": ("UX", "U1"),
    "UY": ("UY", "U2"),
    "UZ": ("UZ", "U3"),
    "SumUX": ("SumUX", "Sum UX", "Cumulative UX", "Cumul UX"),
    "SumUY": ("SumUY", "Sum UY", "Cumulative UY", "Cumul UY"),
    "SumUZ": ("SumUZ", "Sum UZ", "Cumulative UZ", "Cumul UZ"),
    "Height": ("Height", "Story Height", "H"),
    "Elevation": ("Elevation", "Elev"),
    "Direction": ("Direction", "Dir"),
    "Drift": ("Drift", "Drift Ratio", "Max Drift"),
    "MaxDrift": ("MaxDrift", "Max Drift", "Maximum Drift"),
    "AvgDrift": ("AvgDrift", "Avg Drift", "Average Drift"),
    "Ratio": ("Ratio", "Max/Avg", "Max Over Avg"),
    "FX": ("FX", "F1", "X Force"),
    "FY": ("FY", "F2", "Y Force"),
    "FZ": ("FZ", "F3", "Z Force"),
    "Frame": ("Frame", "Beam", "Column", "Object", "Label"),
    "Station": ("Station", "Sta", "Location"),
    "AsTop": ("AsTop", "As Top", "Top As", "Top Rebar"),
    "AsBottom": ("AsBottom", "AsBot", "As Bottom", "Bottom As", "Bottom Rebar"),
    "Asw": ("Asw", "VRebar", "Shear Rebar"),
    "PMMRatio": ("PMMRatio", "PMM Ratio", "PMM", "Ratio"),
    "Pier": ("Pier", "Pier Label", "Wall", "Wall Label"),
    "Thickness": ("Thickness", "Thick", "t"),
    "Width": ("Width", "t2", "T2", "B", "b"),
    "Depth": ("Depth", "t3", "T3", "H", "h"),
    "E1": ("E1", "Elastic Modulus", "Modulus of Elasticity", "E"),
    "G12": ("G12", "Shear Modulus", "G"),
    "U12": ("U12", "Poisson", "Poisson Ratio"),
    "UnitWeight": ("UnitWeight", "Unit Weight", "Weight"),
    "UnitMass": ("UnitMass", "Unit Mass", "Mass"),
    "DensityType": ("DensityType", "Density Type", "Density"),
    "SectProp": ("SectProp", "Section Property", "Section", "Property", "PropName"),
    "Shape": ("Shape", "Section Shape"),
    "Fc": ("Fc", "fck", "Concrete Strength"),
    "Fy": ("Fy", "Fye", "Yield Strength"),
}



@dataclass(frozen=True, slots=True)
class FamilyRule:
    family_id: str
    group: str
    exact_aliases: tuple[str, ...]
    keywords: tuple[str, ...]
    required_columns: tuple[str, ...] = ()
    optional_columns: tuple[str, ...] = ()
    semantic_review: bool = False
    planned_only: bool = False
    source_role: str = "observed_source"
    check_unlock_allowed: bool = False
    profile_allowed: tuple[str, ...] = ("verification_gate",)


FAMILY_RULES: tuple[FamilyRule, ...] = (
    FamilyRule(
        "frame_assignments_summary",
        "current_product",
        ("Frame Assignments - Summary", "Frame Assigns - Summary"),
        ("Frame Assignments", "Frame Assigns", "Frame", "Assignment", "Summary"),
        ("UniqueName", "Label", "Story", "Type", "DesignSect"),
        ("AnalysisSect", "Length"),
        profile_allowed=("current_product", "column_geometry", "verification_gate"),
    ),
    FamilyRule(
        "concrete_rectangular_frame_sections",
        "current_product",
        ("Frame Section Property Definitions - Concrete Rectangular", "Frame Sec Def - Conc Rect"),
        ("Frame Section", "Concrete Rectangular", "Conc Rect"),
        ("Name", "t2", "t3"),
        ("Area", "Material"),
        profile_allowed=("current_product", "column_geometry", "verification_gate"),
    ),
    FamilyRule(
        "modal_participating_mass",
        "current_product",
        ("Modal Participating Mass Ratios",),
        ("Modal Participating Mass", "Modal", "Mass"),
        ("Mode", "Period", "UX", "UY", "SumUX", "SumUY"),
        ("Case", "UZ", "SumUZ"),
        profile_allowed=("current_product", "verification_gate"),
    ),
    FamilyRule(
        "story_definitions",
        "story_global",
        ("Story Definitions", "Story Data", "Stories - Summary", "Story Information"),
        ("Story Definitions", "Story Data", "Story", "Definitions"),
        ("Story", "Height", "Elevation"),
        profile_allowed=("story_global", "verification_gate"),
    ),
    FamilyRule(
        "story_drifts",
        "story_global",
        ("Story Drifts", "Diaphragm Drifts", "Story Drift Data"),
        ("Story Drifts", "Drift", "Story"),
        ("Story", "OutputCase", "Direction", "Drift"),
        ("Label",),
        profile_allowed=("story_global", "verification_gate"),
    ),
    FamilyRule(
        "story_max_over_avg_drifts",
        "story_global",
        ("Story Max Over Avg Drifts", "Story Max/Avg Drifts"),
        ("Story Max Over Avg", "Max Over Avg", "Drift"),
        ("Story", "OutputCase", "Direction", "Ratio"),
        ("MaxDrift", "AvgDrift"),
        profile_allowed=("story_global", "verification_gate"),
    ),
    FamilyRule(
        "base_reactions",
        "story_global",
        ("Base Reactions", "Base Reactions Summary"),
        ("Base Reactions", "Base Reaction"),
        ("OutputCase", "FX", "FY", "FZ"),
        profile_allowed=("story_global", "verification_gate"),
    ),
    FamilyRule(
        "material_properties",
        "material_context",
        ("Material Properties", "Material Properties - Summary", "Mat Prop - Basic Mech Props", "Material Properties - Basic Mechanical Properties"),
        ("Material Properties", "Basic Mechanical", "Mat Prop"),
        ("Material", "E1", "G12", "U12"),
        ("UnitWeight", "UnitMass", "DensityType"),
        source_role="basic_mechanical_material_properties",
        profile_allowed=("material_context", "verification_gate"),
    ),
    FamilyRule(
        "material_list_by_story",
        "material_context",
        ("Material List by Story",),
        ("Material List", "Story"),
        ("Story", "Material"),
        source_role="quantity_or_inventory_context_only",
        check_unlock_allowed=False,
        profile_allowed=("material_context", "verification_gate"),
    ),
    FamilyRule(
        "concrete_material_properties",
        "material_context",
        ("Mat Prop - Concrete Data", "Material Properties - Concrete Data"),
        ("Concrete", "Material", "Properties"),
        ("Material", "Fc"),
        source_role="concrete_material_strength_properties",
        profile_allowed=("material_context", "verification_gate"),
    ),
    FamilyRule(
        "rebar_material_properties",
        "material_context",
        ("Mat Prop - Rebar Data", "Material Properties - Rebar Data"),
        ("Rebar", "Material", "Properties"),
        ("Material", "Fy"),
        source_role="rebar_material_strength_properties",
        profile_allowed=("material_context", "verification_gate"),
    ),
    FamilyRule(
        "frame_section_material_assignments",
        "material_context",
        ("Frame Prop - Summary", "Frame Section Property Definitions - Summary"),
        ("Frame", "Section", "Material"),
        ("Name", "Material"),
        ("Shape",),
        source_role="section_property_material_mapping",
        profile_allowed=("material_context", "verification_gate"),
    ),
    FamilyRule(
        "frame_section_assignments",
        "material_context",
        ("Frame Assignments - Section Properties", "Frame Assigns - Sect Prop", "Frame Section Assignments"),
        ("Frame Assignments", "Section Properties"),
        ("Story", "Label", "UniqueName", "SectProp"),
        ("Shape",),
        source_role="section_assignment_context_only",
        check_unlock_allowed=False,
        profile_allowed=("material_context", "verification_gate"),
    ),
    FamilyRule(
        "concrete_beam_design_summary",
        "beam_design_outputs",
        ("Concrete Beam Design Summary - TS 500-2000(R2018)", "Concrete Beam Design Summary"),
        ("Concrete Beam Design", "Beam Design", "Summary"),
        ("Frame", "Station", "AsTop", "AsBottom"),
        ("Asw", "PMMRatio"),
        semantic_review=True,
        profile_allowed=("beam_design_outputs", "verification_gate"),
    ),
    FamilyRule(
        "concrete_beam_flexure_envelope",
        "beam_design_outputs",
        ("Concrete Beam Flexure Envelope - TS 500-2000(R2018)", "Concrete Beam Flexure Envelope"),
        ("Beam", "Flexure", "Envelope"),
        semantic_review=True,
        profile_allowed=("beam_design_outputs", "verification_gate"),
    ),
    FamilyRule(
        "concrete_beam_shear_envelope",
        "beam_design_outputs",
        ("Concrete Beam Shear Envelope - TS 500-2000(R2018)", "Concrete Beam Shear Envelope"),
        ("Beam", "Shear", "Envelope"),
        semantic_review=True,
        profile_allowed=("beam_design_outputs", "verification_gate"),
    ),
    FamilyRule(
        "frame_forces",
        "beam_design_outputs",
        ("Frame Forces", "Element Forces - Frames", "Frame Element Forces"),
        ("Frame", "Forces"),
        semantic_review=True,
        profile_allowed=("beam_design_outputs", "verification_gate"),
    ),
    FamilyRule(
        "concrete_column_design_summary",
        "column_design_outputs",
        ("Concrete Column Design Summary - TS 500-2000(R2018)", "Concrete Column Design Summary"),
        ("Concrete Column Design", "Column Design", "Summary"),
        semantic_review=True,
        profile_allowed=("column_design_outputs", "verification_gate"),
    ),
    FamilyRule(
        "concrete_column_pmm_envelope",
        "column_design_outputs",
        ("Concrete Column PMM Envelope - TS 500-2000(R2018)", "Concrete Column PMM Envelope"),
        ("Column", "PMM", "Envelope"),
        semantic_review=True,
        profile_allowed=("column_design_outputs", "verification_gate"),
    ),
    FamilyRule(
        "column_forces",
        "column_design_outputs",
        ("Column Forces", "Frame Forces"),
        ("Column", "Forces"),
        semantic_review=True,
        profile_allowed=("column_design_outputs", "verification_gate"),
    ),
    FamilyRule(
        "area_assignments_summary",
        "wall_area",
        ("Area Assigns - Summary", "Area Assignments - Summary", "Area Object Assignments"),
        ("Area Assignments", "Area Assigns", "Area", "Summary"),
        profile_allowed=("wall_area", "verification_gate"),
    ),
    FamilyRule(
        "wall_section_properties",
        "wall_area",
        ("Wall Property Definitions - Specified", "Area Section Props - Summary", "Area Section Property Definitions"),
        ("Wall", "Area Section", "Section Property", "Shell"),
        profile_allowed=("wall_area", "verification_gate"),
    ),
    FamilyRule(
        "pier_assignments",
        "wall_area",
        ("Area Assigns - Pier Labels", "Pier Assignments", "Pier Labels"),
        ("Pier", "Assignment", "Wall"),
        profile_allowed=("wall_area", "verification_gate"),
    ),
    FamilyRule(
        "pier_section_properties",
        "wall_area",
        ("Pier Section Properties",),
        ("Pier", "Section", "Properties"),
        ("Pier", "Length", "Thickness"),
        profile_allowed=("wall_area", "verification_gate"),
    ),
    FamilyRule(
        "pier_forces",
        "wall_area",
        ("Pier Forces", "Wall Forces", "Pier Element Forces"),
        ("Pier", "Forces", "Wall"),
        semantic_review=True,
        profile_allowed=("wall_area", "verification_gate"),
    ),
    FamilyRule(
        "shear_wall_design_summary",
        "wall_area",
        ("Shear Wall Design Summary - TS 500-2000(R2018)", "Shear Wall Design Summary"),
        ("Wall", "Design", "Summary"),
        semantic_review=True,
        profile_allowed=("wall_area", "verification_gate"),
    ),
)

PROBE_PROFILES: dict[str, dict[str, Any]] = {
    "current_product": {"families": ("frame_assignments_summary", "concrete_rectangular_frame_sections", "modal_participating_mass"), "default_safe": True},
    "column_geometry": {"families": ("frame_assignments_summary", "concrete_rectangular_frame_sections")},
    "story_global": {"group": "story_global"},
    "material_context": {"group": "material_context"},
    "beam_design_outputs": {"group": "beam_design_outputs"},
    "column_design_outputs": {"group": "column_design_outputs"},
    "wall_area": {"group": "wall_area"},
    "verification_gate": {"group": "all"},
}

ARTIFACT_NAMES = (
    "connection_report.json",
    "excel_inventory_parse_report.json",
    "excel_table_family_classification.json",
    "live_available_tables.json",
    "excel_to_live_table_match_report.json",
    "live_header_comparison_report.json",
    "live_sample_rows_report.json",
    "source_promotion_recommendation.json",
    "semantic_review_sources.json",
    "needs_live_probe_sources.json",
    "c13_2_expansion_decision_report.json",
    "C13_2_P1_EXCEL_GUIDED_LIVE_SOURCE_VERIFICATION.md",
)


def _norm(text: Any) -> str:
    return re.sub(r"[^a-z0-9]+", "", str(text or "").casefold())


def _tokens(text: Any) -> set[str]:
    return {t for t in re.split(r"[^a-z0-9]+", str(text or "").casefold()) if t and t not in WEAK_ONE_WORD_KEYWORDS}


def _json_safe(value: Any) -> Any:
    if isinstance(value, Mapping):
        return {str(k): _json_safe(v) for k, v in value.items()}
    if isinstance(value, (list, tuple, set, frozenset)):
        return [_json_safe(v) for v in value]
    try:
        json.dumps(value)
        return value
    except TypeError:
        return repr(value)


def _write_json(path: Path, payload: Any) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(json.dumps(_json_safe(payload), indent=2, ensure_ascii=False), encoding="utf-8")


def _write_text(path: Path, text: str) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(text, encoding="utf-8")


def aliases_for(column: str) -> tuple[str, ...]:
    return HEADER_ALIASES.get(column, (column,))


def _has_header(headers: Sequence[Any], logical_column: str) -> bool:
    header_norms = {_norm(h) for h in headers if _norm(h)}
    for alias in aliases_for(logical_column):
        alias_norm = _norm(alias)
        if not alias_norm:
            continue
        if alias_norm in header_norms:
            return True
        if len(alias_norm) > 2 and any(alias_norm in h or h in alias_norm for h in header_norms):
            return True
    return False


def validate_expected_headers(headers: Sequence[Any], required: Sequence[str], optional: Sequence[str] = ()) -> dict[str, Any]:
    matched_required = [c for c in required if _has_header(headers, c)]
    missing_required = [c for c in required if c not in matched_required]
    matched_optional = [c for c in optional if _has_header(headers, c)]
    return {
        "validation_applies": bool(required),
        "passed": not missing_required,
        "matched_required": matched_required,
        "missing_required": missing_required,
        "matched_optional": matched_optional,
        "alias_policy_used": True,
        "alias_proof": {c: list(aliases_for(c)) for c in required},
    }


def classify_table_family(table_name: str, headers: Sequence[str] = ()) -> str:
    name_norm = _norm(table_name)
    header_norm = " ".join(_norm(h) for h in headers)
    for rule in FAMILY_RULES:
        if any(_norm(a) == name_norm for a in rule.exact_aliases):
            return rule.family_id
    if "modal" in name_norm and "participating" in name_norm and "mass" in name_norm:
        return "modal_participating_mass"
    if "frame" in name_norm and ("assign" in name_norm or "assignments" in name_norm) and "summary" in name_norm:
        return "frame_assignments_summary"
    if "framesection" in name_norm and ("concreterectangular" in name_norm or "concrect" in name_norm):
        return "concrete_rectangular_frame_sections"
    if "storymax" in name_norm and "avg" in name_norm and "drift" in name_norm:
        return "story_max_over_avg_drifts"
    if "story" in name_norm and "drift" in name_norm:
        return "story_drifts"
    if "story" in name_norm and ("definition" in name_norm or "data" in name_norm or "information" in name_norm):
        return "story_definitions"
    if "basereaction" in name_norm:
        return "base_reactions"
    if "matpropconcrete" in name_norm or ("concrete" in name_norm and "material" in name_norm):
        return "concrete_material_properties"
    if "matproprebar" in name_norm or ("rebar" in name_norm and "material" in name_norm):
        return "rebar_material_properties"
    if "materiallistbystory" in name_norm or ("materiallist" in name_norm and "story" in name_norm):
        return "material_list_by_story"
    if "frameassign" in name_norm and "section" in name_norm:
        return "frame_section_assignments"
    if "frameprop" in name_norm and "summary" in name_norm:
        return "frame_section_material_assignments"
    if "material" in name_norm:
        return "material_properties"
    if "beamdesignsummary" in name_norm or "concretebeamdesignsummary" in name_norm:
        return "concrete_beam_design_summary"
    if "beamflexure" in name_norm and "envelope" in name_norm:
        return "concrete_beam_flexure_envelope"
    if "beamshear" in name_norm and "envelope" in name_norm:
        return "concrete_beam_shear_envelope"
    if "columndesignsummary" in name_norm or "concretecolumndesignsummary" in name_norm:
        return "concrete_column_design_summary"
    if "column" in name_norm and "pmm" in name_norm and "envelope" in name_norm:
        return "concrete_column_pmm_envelope"
    if "areassign" in name_norm or "areaassign" in name_norm:
        return "area_assignments_summary"
    if "areasection" in name_norm or "wallproperty" in name_norm:
        return "wall_section_properties"
    if "pier" in name_norm and "assign" in name_norm:
        return "pier_assignments"
    if "piersection" in name_norm:
        return "pier_section_properties"
    if "pierforces" in name_norm or "wallforces" in name_norm:
        return "pier_forces"
    if "shearwall" in name_norm and "designsummary" in name_norm:
        return "shear_wall_design_summary"
    # Header hints as last resort.
    if all(_norm(c) in header_norm for c in ("typename", "designsect")):
        return "frame_assignments_summary"
    return "unknown_or_ambiguous"


def family_rule(family_id: str) -> FamilyRule | None:
    for rule in FAMILY_RULES:
        if rule.family_id == family_id:
            return rule
    return None


def family_rules_for_profile(profile: str) -> list[FamilyRule]:
    """Return the configured family rules for a profile.

    This function intentionally returns the profile definition only.  Runtime
    probing must further scope the result to families observed in the Excel/JSON
    inventory unless --include-planned-families is explicitly used.
    """
    if profile not in PROBE_PROFILES:
        raise ValueError(f"Unknown probe profile: {profile}")
    spec = PROBE_PROFILES[profile]
    if "families" in spec:
        wanted = set(spec["families"])
        return [r for r in FAMILY_RULES if r.family_id in wanted]
    group = spec.get("group")
    if group == "all":
        return list(FAMILY_RULES)
    return [r for r in FAMILY_RULES if r.group == group]


def observed_family_ids(tables: Sequence[Mapping[str, Any]]) -> set[str]:
    """Families actually observed in the Excel/JSON inventory.

    Unknown/ambiguous rows are not probe targets.  They remain inventory evidence
    only and cannot cause live ETABS fetches.
    """
    ids = {str(t.get("family_id") or "") for t in tables}
    return {family_id for family_id in ids if family_rule(family_id) is not None}


def scoped_family_rules_for_inventory(
    profile: str,
    tables: Sequence[Mapping[str, Any]],
    *,
    include_planned_families: bool = False,
) -> list[FamilyRule]:
    """Profile rules narrowed to inventory-observed families by default.

    C13.2-P1 is a verification gate.  It must not probe every internal family
    rule merely because the code knows about it.  Absent families may be listed
    as PLANNED only when include_planned_families=True, and even then they must
    not be fetched live.
    """
    rules = family_rules_for_profile(profile)
    observed = observed_family_ids(tables)
    if include_planned_families:
        return rules
    return [rule for rule in rules if rule.family_id in observed]


def parse_inventory_json(path: Path) -> list[dict[str, Any]]:
    payload = json.loads(path.read_text(encoding="utf-8"))
    if isinstance(payload, list):
        raw_tables = payload
    elif isinstance(payload, Mapping):
        raw_tables = payload.get("tables") or payload.get("sheets") or payload.get("inventory") or []
    else:
        raw_tables = []
    tables: list[dict[str, Any]] = []
    for idx, item in enumerate(raw_tables):
        if isinstance(item, str):
            tables.append({"excel_table_name": item, "headers": [], "row_count": None, "inventory_index": idx})
        elif isinstance(item, Mapping):
            name = item.get("excel_table_name") or item.get("table_name") or item.get("sheet_name") or item.get("name")
            headers = item.get("headers") or item.get("columns") or item.get("field_keys") or []
            row_count = item.get("row_count") or item.get("rows") or item.get("number_records")
            tables.append({"excel_table_name": str(name or f"table_{idx}"), "headers": [str(h) for h in headers], "row_count": row_count, "inventory_index": idx})
    return tables


def _first_nonempty_row_values(ws: Any) -> tuple[int | None, list[str]]:
    for row_index, row in enumerate(ws.iter_rows(values_only=True), start=1):
        values = ["" if v is None else str(v).strip() for v in row]
        non_empty = [v for v in values if v]
        if len(non_empty) >= 2:
            return row_index, values
    return None, []


def parse_inventory_xlsx(path: Path) -> list[dict[str, Any]]:
    try:
        from openpyxl import load_workbook  # type: ignore
    except Exception as exc:  # pragma: no cover - environment specific
        raise RuntimeError("openpyxl is required to parse .xlsx inventory files") from exc
    wb = load_workbook(path, read_only=True, data_only=True)
    tables: list[dict[str, Any]] = []
    for idx, sheet_name in enumerate(wb.sheetnames):
        ws = wb[sheet_name]
        header_row_index, headers = _first_nonempty_row_values(ws)
        tables.append({
            "excel_table_name": sheet_name,
            "headers": [h for h in headers if h],
            "row_count": max(0, (ws.max_row or 0) - (header_row_index or 1)),
            "inventory_index": idx,
            "header_row_index": header_row_index,
        })
    return tables


def parse_inventory(excel_inventory: Path | None, inventory_json: Path | None) -> tuple[list[dict[str, Any]], dict[str, Any]]:
    if inventory_json:
        tables = parse_inventory_json(inventory_json)
        source_type = "json"
        source_path = str(inventory_json)
    elif excel_inventory:
        tables = parse_inventory_xlsx(excel_inventory)
        source_type = "xlsx"
        source_path = str(excel_inventory)
    else:
        raise ValueError("Either --excel-inventory or --inventory-json is required")
    for table in tables:
        table["family_id"] = classify_table_family(str(table.get("excel_table_name") or ""), table.get("headers") or [])
    family_counts: dict[str, int] = {}
    for table in tables:
        family_counts[str(table["family_id"])] = family_counts.get(str(table["family_id"]), 0) + 1
    report = {
        "inventory_source_type": source_type,
        "inventory_source_path": source_path,
        "table_count": len(tables),
        "classified_family_counts": family_counts,
        "excel_role": "probe_target_inventory_only",
        "excel_can_create_verified_live": False,
    }
    return tables, report


def classify_inventory_tables(tables: Sequence[Mapping[str, Any]]) -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []
    for table in tables:
        family_id = str(table.get("family_id") or "unknown_or_ambiguous")
        rule = family_rule(family_id)
        rows.append({
            "excel_table_name": table.get("excel_table_name"),
            "family_id": family_id,
            "group": rule.group if rule else "unknown_or_ambiguous",
            "semantic_review": bool(rule.semantic_review) if rule else False,
            "headers": list(table.get("headers") or []),
            "row_count": table.get("row_count"),
            "initial_status": "EXCEL_INVENTORY_ONLY" if not rule or not rule.semantic_review else "SEMANTIC_REVIEW",
        })
    return rows


def _score_live_candidate(live_name: str, rule: FamilyRule, excel_names: Sequence[str]) -> int:
    live_norm = _norm(live_name)
    score = 0
    if any(_norm(a) == live_norm for a in rule.exact_aliases):
        score += 10000
    for alias in rule.exact_aliases:
        if _norm(alias) and _norm(alias) in live_norm:
            score += 1000
    live_tokens = _tokens(live_name)
    for excel_name in excel_names:
        overlap = live_tokens & _tokens(excel_name)
        score += len(overlap) * 100
        if _norm(excel_name) == live_norm:
            score += 5000
    for keyword in rule.keywords:
        kw_tokens = _tokens(keyword)
        if kw_tokens and kw_tokens <= live_tokens:
            score += 300
    return score


def match_excel_to_live_tables(
    tables: Sequence[Mapping[str, Any]],
    live_table_names: Sequence[str],
    profile: str,
    max_candidate_tables_per_family: int,
    *,
    include_planned_families: bool = False,
) -> list[dict[str, Any]]:
    rules = scoped_family_rules_for_inventory(profile, tables, include_planned_families=include_planned_families)
    by_family: dict[str, list[Mapping[str, Any]]] = {}
    for table in tables:
        family_id = str(table.get("family_id") or "unknown_or_ambiguous")
        by_family.setdefault(family_id, []).append(table)
    live_lookup = {_norm(name): name for name in live_table_names}
    rows: list[dict[str, Any]] = []
    for rule in rules:
        family_tables = by_family.get(rule.family_id, [])
        planned_absent = not bool(family_tables)
        excel_names = [str(t.get("excel_table_name")) for t in family_tables]
        exact_matches: list[str] = []
        # Planned/absent families are report-only planning rows when explicitly
        # requested.  They must never produce live fetch candidates.
        if planned_absent:
            live_matches = []
            match_basis = "planned_absent"
            candidate_count_before_cap = 0
        else:
            for name in excel_names + list(rule.exact_aliases):
                matched = live_lookup.get(_norm(name))
                if matched and matched not in exact_matches:
                    exact_matches.append(matched)
            if exact_matches:
                live_matches = exact_matches[:max_candidate_tables_per_family]
                match_basis = "exact"
                candidate_count_before_cap = len(exact_matches)
            else:
                scored = [(name, _score_live_candidate(name, rule, excel_names)) for name in live_table_names]
                ranked = [name for name, score in sorted(scored, key=lambda item: (-item[1], item[0])) if score > 0]
                live_matches = ranked[:max_candidate_tables_per_family]
                match_basis = "keyword_or_alias"
                candidate_count_before_cap = len(ranked)
        rows.append({
            "family_id": rule.family_id,
            "group": rule.group,
            "excel_table_names": excel_names,
            "excel_table_name": excel_names[0] if excel_names else None,
            "live_candidate_tables": live_matches,
            "live_table_name": live_matches[0] if live_matches else None,
            "match_basis": match_basis if (live_matches or planned_absent) else "none",
            "planned_absent": planned_absent,
            "planned_live_fetch_allowed": False if planned_absent else True,
            "candidate_count_before_cap": candidate_count_before_cap,
            "candidate_count_after_cap": len(live_matches),
            "candidate_truncation_applied": len(live_matches) < candidate_count_before_cap,
            "semantic_review": rule.semantic_review,
            "required_columns": list(rule.required_columns),
            "optional_columns": list(rule.optional_columns),
        })
    return rows


def _compact_shape(raw: Any) -> list[dict[str, Any]]:
    if not isinstance(raw, (list, tuple)):
        return []
    out: list[dict[str, Any]] = []
    for idx, item in enumerate(raw):
        text = repr(item)
        out.append({"index": idx, "type": type(item).__name__, "len": len(item) if isinstance(item, (list, tuple, str)) else None, "repr": text[:200]})
    return out


def _fetch_live_table(database_tables: Any, table_name: str, max_sample_rows: int) -> tuple[list[str], list[dict[str, Any]], dict[str, Any]]:
    from tbdy_engine.providers.etabs_display_table_fetcher import fetch_display_table  # lazy import
    result = fetch_display_table(database_tables, table_name, max_rows=max_sample_rows)
    parsed = result.parsed
    headers = [str(h) for h in list(getattr(parsed, "field_keys", []) or [])]
    rows = [dict(r) for r in list(getattr(parsed, "rows", []) or [])[:max_sample_rows]]
    raw_shape = {
        "attempted_table_name": table_name,
        "raw_type": type(getattr(result, "raw_response", None)).__name__,
        "raw_slot_summary": _compact_shape(getattr(result, "raw_response", None)),
        "fetch_status": getattr(parsed, "fetch_status", None),
        "row_count_reported": getattr(parsed, "row_count_reported", None),
    }
    return headers, rows, raw_shape


def compare_live_headers(match_rows: Sequence[Mapping[str, Any]], fetched: Mapping[str, Mapping[str, Any]]) -> list[dict[str, Any]]:
    out: list[dict[str, Any]] = []
    for match in match_rows:
        rule = family_rule(str(match.get("family_id")))
        live_name = match.get("live_table_name")
        live_payload = fetched.get(str(live_name)) if live_name else None
        headers = list(live_payload.get("headers") or []) if live_payload else []
        validation = validate_expected_headers(headers, rule.required_columns if rule else (), rule.optional_columns if rule else ())
        excel_headers = []
        for table_name in match.get("excel_table_names") or []:
            # Filled by caller in promotion rows; comparison stores live proof only.
            pass
        out.append({
            "family_id": match.get("family_id"),
            "live_table_name": live_name,
            "live_headers": headers,
            "expected_header_validation": validation,
            "live_sample_row_count": len(live_payload.get("rows") or []) if live_payload else 0,
            "error": live_payload.get("error") if live_payload else None,
        })
    return out


def column_geometry_gate(sample_rows_by_table: Mapping[str, Mapping[str, Any]]) -> dict[str, Any]:
    assignment_rows: list[Mapping[str, Any]] = []
    section_rows: list[Mapping[str, Any]] = []
    for live_name, payload in sample_rows_by_table.items():
        norm = _norm(live_name)
        rows = list(payload.get("rows") or [])
        if "frameassignment" in norm or "frameassign" in norm:
            assignment_rows.extend(rows)
        if "framesection" in norm or "framesecdef" in norm or "concreterectangular" in norm:
            section_rows.extend(rows)
    column_rows = [r for r in assignment_rows if str(r.get("Type") or r.get("Design Type") or r.get("Object Type") or "").strip().casefold() == "column"]
    column_sections = sorted({str(r.get("DesignSect") or r.get("Design Section") or "").strip() for r in column_rows if str(r.get("DesignSect") or r.get("Design Section") or "").strip()})
    section_names = sorted({str(r.get("Name") or r.get("SectionName") or r.get("Section Name") or "").strip() for r in section_rows if str(r.get("Name") or r.get("SectionName") or r.get("Section Name") or "").strip()})
    missing = [s for s in column_sections if s not in section_names]
    passed = bool(column_rows) and bool(column_sections) and not missing
    type_counts: dict[str, int] = {}
    for r in assignment_rows:
        key = str(r.get("Type") or r.get("Design Type") or r.get("Object Type") or "").strip() or "<blank>"
        type_counts[key] = type_counts.get(key, 0) + 1
    return {
        "passed": passed,
        "status": "VERIFIED_LIVE_FOR_COLUMN_GEOMETRY_CONTRACT" if passed else "NEEDS_LIVE_PROBE",
        "assignment_row_count": len(assignment_rows),
        "type_distribution": type_counts,
        "column_row_count": len(column_rows),
        "column_sections": column_sections,
        "concrete_rectangular_section_count": len(section_names),
        "missing_column_sections": missing,
        "matched_column_sections": [s for s in column_sections if s in section_names],
        "sample_column_rows": [dict(r) for r in column_rows[:10]],
    }


def semantic_source_role_validation(family_id: str, live_table_name: Any, live_headers: Sequence[Any]) -> dict[str, Any]:
    """Validate that live table semantics match the requested family.

    A table can exist and headers can partially match while still proving the
    wrong source role.  C13.2-P1 may only recommend VERIFIED_LIVE when the live
    source role is semantically compatible with the family id.
    """
    name_norm = _norm(live_table_name)
    blockers: list[str] = []
    passed = True

    if family_id == "material_properties":
        if "materiallist" in name_norm:
            passed = False
            blockers.append("Material List tables are inventory/quantity context, not basic mechanical material properties")
        for col in ("Material", "E1", "G12", "U12"):
            if not _has_header(live_headers, col):
                passed = False
                blockers.append(f"missing semantic material property header: {col}")

    if family_id == "frame_section_material_assignments":
        proves_section_assignment_only = "frameassignment" in name_norm or "frameassign" in name_norm
        missing_material = not _has_header(live_headers, "Material")
        if proves_section_assignment_only or missing_material:
            passed = False
        if proves_section_assignment_only and missing_material:
            blockers.append("live table proves section assignment, not material assignment; Material header missing")
            blockers.append("missing Material header required for section-material mapping")
        elif proves_section_assignment_only:
            blockers.append("live table proves section assignment, not section-to-material mapping")
        elif missing_material:
            blockers.append("Material header missing; cannot prove section-to-material mapping")
        # Section-material mapping should be proven by a section property summary
        # source, not by an object assignment source.  A keyword table/header match
        # is not enough for VERIFIED_LIVE when the source role is different.

    return {"passed": passed, "blockers": blockers}


def build_promotion_rows(
    classification_rows: Sequence[Mapping[str, Any]],
    match_rows: Sequence[Mapping[str, Any]],
    header_comparison: Sequence[Mapping[str, Any]],
    live_mode: bool,
) -> list[dict[str, Any]]:
    excel_headers_by_family: dict[str, list[str]] = {}
    excel_names_by_family: dict[str, list[str]] = {}
    for row in classification_rows:
        family_id = str(row.get("family_id"))
        excel_names_by_family.setdefault(family_id, []).append(str(row.get("excel_table_name")))
        for header in row.get("headers") or []:
            if header not in excel_headers_by_family.setdefault(family_id, []):
                excel_headers_by_family[family_id].append(str(header))
    header_by_family = {str(row.get("family_id")): row for row in header_comparison}
    rows: list[dict[str, Any]] = []
    for match in match_rows:
        family_id = str(match.get("family_id"))
        rule = family_rule(family_id)
        comp = header_by_family.get(family_id, {})
        validation = comp.get("expected_header_validation") or {"passed": False}
        live_headers = list(comp.get("live_headers") or [])
        live_sample_row_count = int(comp.get("live_sample_row_count") or 0)
        live_table_name = match.get("live_table_name")
        semantic_validation = semantic_source_role_validation(family_id, live_table_name, live_headers)
        planned_absent = bool(match.get("planned_absent"))
        observed_in_excel = (not planned_absent) and bool(excel_names_by_family.get(family_id))
        planned_without_excel_evidence = planned_absent
        live_fetch_allowed = bool(match.get("planned_live_fetch_allowed")) and observed_in_excel and not planned_absent
        blockers: list[str] = []
        if planned_absent:
            recommended_status = "PLANNED"
            match_quality = "EXCEL_ONLY"
            blockers.append("family absent from Excel inventory; planned only and not live-fetched")
        elif rule and rule.semantic_review:
            recommended_status = "SEMANTIC_REVIEW"
            match_quality = "SEMANTIC_REVIEW_REQUIRED"
            blockers.append("design/force output semantics require human review; no check unlock")
        elif not live_mode:
            recommended_status = "EXCEL_INVENTORY_ONLY"
            match_quality = "EXCEL_ONLY"
            blockers.append("parse-only mode; live ETABS proof absent")
        elif not live_table_name:
            recommended_status = "NEEDS_LIVE_PROBE"
            match_quality = "LIVE_TABLE_NOT_FOUND"
            blockers.append("no matching live table found")
        elif not validation.get("passed"):
            recommended_status = "NEEDS_LIVE_PROBE"
            match_quality = "EXACT_TABLE_HEADER_PARTIAL" if match.get("match_basis") == "exact" else "KEYWORD_TABLE_HEADER_MATCH"
            blockers.append("live table headers do not prove expected semantics")
            blockers.extend(str(b) for b in semantic_validation.get("blockers", []))
        elif live_sample_row_count <= 0:
            recommended_status = "NEEDS_LIVE_PROBE"
            match_quality = "EXACT_TABLE_HEADER_PARTIAL"
            blockers.append("live headers found but no sample rows")
        elif not semantic_validation.get("passed"):
            recommended_status = "NEEDS_LIVE_PROBE"
            match_quality = "EXACT_TABLE_HEADER_PARTIAL" if match.get("match_basis") == "exact" else "KEYWORD_TABLE_HEADER_MATCH"
            blockers.extend(str(b) for b in semantic_validation.get("blockers", []))
        else:
            recommended_status = "VERIFIED_LIVE"
            match_quality = "EXACT_TABLE_AND_HEADERS" if match.get("match_basis") == "exact" else "KEYWORD_TABLE_HEADER_MATCH"
        rows.append({
            "family_id": family_id,
            "excel_table_name": (excel_names_by_family.get(family_id) or [None])[0],
            "all_excel_table_names": excel_names_by_family.get(family_id, []),
            "live_table_name": live_table_name,
            "excel_headers": excel_headers_by_family.get(family_id, []),
            "live_headers": live_headers,
            "live_sample_row_count": live_sample_row_count,
            "source_role": rule.source_role if rule else "unknown_or_ambiguous",
            "check_unlock_allowed": False,
            "semantic_source_role_validation": semantic_validation,
            "match_quality": match_quality,
            "recommended_status": recommended_status,
            "planned_absent": planned_absent,
            "observed_in_excel": observed_in_excel,
            "planned_without_excel_evidence": planned_without_excel_evidence,
            "live_fetch_allowed": live_fetch_allowed,
            "can_expand_contract_now": recommended_status == "VERIFIED_LIVE",
            "can_implement_check_now": False,
            "blockers": blockers,
        })
    return rows


def expansion_decision_report(
    inventory_report: Mapping[str, Any],
    promotion_rows: Sequence[Mapping[str, Any]],
    column_gate: Mapping[str, Any],
    live_mode: bool,
) -> dict[str, Any]:
    verified = [r for r in promotion_rows if r.get("recommended_status") == "VERIFIED_LIVE"]
    semantic = [r for r in promotion_rows if r.get("recommended_status") == "SEMANTIC_REVIEW"]
    needs_probe = [r for r in promotion_rows if r.get("recommended_status") == "NEEDS_LIVE_PROBE"]
    excel_only = [r for r in promotion_rows if r.get("recommended_status") == "EXCEL_INVENTORY_ONLY"]
    planned = [r for r in promotion_rows if r.get("recommended_status") == "PLANNED"]
    current_safe_checks = [
        "beam_geometry_min_width",
        "beam_geometry_min_depth",
        "beam_depth_width_ratio",
        "column_geometry_report_contract_ready",
        "modal_mass_UX_UY_report_verdict",
    ]
    full_expansion_now = False
    reason = "C13.2-P1 is a verification gate only; full contract/schema expansion requires human approval after reviewing live proof."
    if not live_mode:
        reason = "Parse-only mode: Excel inventory cannot promote sources to VERIFIED_LIVE."
    elif needs_probe or semantic:
        reason = "Some families still need live proof or semantic review; full contract/schema expansion is not safe yet."
    return {
        "full_c13_2_contract_expansion_now": full_expansion_now,
        "reason": reason,
        "excel_used_as_probe_target_inventory": True,
        "excel_not_used_as_production_input": True,
        "live_mode": live_mode,
        "inventory_table_count": inventory_report.get("table_count"),
        "promotion_counts": {
            "VERIFIED_LIVE": len(verified),
            "SEMANTIC_REVIEW": len(semantic),
            "NEEDS_LIVE_PROBE": len(needs_probe),
            "EXCEL_INVENTORY_ONLY": len(excel_only),
            "PLANNED": len(planned),
        },
        "verified_live_families": [r.get("family_id") for r in verified],
        "semantic_review_families": [r.get("family_id") for r in semantic],
        "needs_live_probe_families": [r.get("family_id") for r in needs_probe],
        "planned_families": [r.get("family_id") for r in planned],
        "human_approval_required_before_contract_expansion": True,
        "column_geometry_gate": column_gate,
        "current_safe_check_capacity": {
            "current_safe_check_count": 5,
            "safe_check_or_report_items": current_safe_checks,
            "conclusion": "Product remains geometry/modal-report limited. Full TBDY is not unlocked.",
        },
        "blocked_check_families": {
            "rebar": "ETABS design output / provided rebar semantics not verified",
            "beam_flexure": "moment envelope / rebar semantics need live proof + semantic review",
            "beam_shear": "shear demand/design output semantics need live proof + semantic review",
            "column_axial_pmm": "force/design summary semantics not verified for checks",
            "wall_checks": "area/pier/wall table semantics not verified for checks",
            "story_drift": "story drift table and output-case/limit policy not verified for check contract",
            "base_shear": "base reaction/seismic source semantics not verified for final check",
        },
        "safe_to_implement_checks_now": False,
    }


def build_markdown_report(decision: Mapping[str, Any], inventory_report: Mapping[str, Any], promotion_rows: Sequence[Mapping[str, Any]]) -> str:
    def section_rows(status: str) -> list[Mapping[str, Any]]:
        return [r for r in promotion_rows if r.get("recommended_status") == status]

    def md_table(rows: Sequence[Mapping[str, Any]]) -> str:
        if not rows:
            return "_(none)_\n"
        headers = ["family_id", "excel_table_name", "live_table_name", "recommended_status", "match_quality", "can_expand_contract_now"]
        lines = ["| " + " | ".join(headers) + " |", "| " + " | ".join("---" for _ in headers) + " |"]
        for row in rows:
            lines.append("| " + " | ".join(str(row.get(h, "")) for h in headers) + " |")
        return "\n".join(lines) + "\n"

    parts = [
        "# C13.2-P1 Excel-Guided Live Source Verification",
        "",
        "## Decision",
        "",
        f"Full C13.2 contract/schema expansion now: **{decision.get('full_c13_2_contract_expansion_now')}**",
        "",
        str(decision.get("reason")),
        "",
        "Excel is used only as probe target inventory. Live ETABS proof is required for VERIFIED_LIVE. This sprint implements no engineering checks.",
        "",
        "## Inventory Summary",
        "",
        f"Inventory table count: {inventory_report.get('table_count')}",
        "",
        "## VERIFIED_LIVE",
        md_table(section_rows("VERIFIED_LIVE")),
        "## SEMANTIC_REVIEW",
        md_table(section_rows("SEMANTIC_REVIEW")),
        "## NEEDS_LIVE_PROBE",
        md_table(section_rows("NEEDS_LIVE_PROBE")),
        "## EXCEL_INVENTORY_ONLY",
        md_table(section_rows("EXCEL_INVENTORY_ONLY")),
        "## PLANNED",
        md_table(section_rows("PLANNED")),
        "## Current Safe Check Capacity",
        "",
        f"Current safe check/report capacity: {decision.get('current_safe_check_capacity', {}).get('current_safe_check_count')} items.",
        "",
        "The product remains geometry/modal-report limited. Full TBDY is not unlocked.",
        "",
        "safe_to_implement_checks_now: false",
    ]
    return "\n".join(parts)


def _empty_live_outputs() -> tuple[list[str], list[dict[str, Any]], dict[str, Mapping[str, Any]], list[dict[str, Any]], list[dict[str, Any]]]:
    return [], [], {}, [], []


def run_probe(
    *,
    out: Path,
    excel_inventory: Path | None,
    inventory_json: Path | None,
    live_etabs: bool,
    probe_profile: str,
    max_candidate_tables_per_family: int,
    max_sample_rows: int,
    preferred_output_case: str | None,
    include_planned_families: bool = False,
) -> int:
    out.mkdir(parents=True, exist_ok=True)
    connection_report = {
        "live_etabs_requested": live_etabs,
        "live_etabs_connected": False,
        "etabs_model_available": False,
        "preferred_output_case": preferred_output_case,
        "probe_profile": probe_profile,
        "display_selection_attempted": False,
        "display_selection_status": "NOT_ATTEMPTED",
        "errors": [],
        "warnings": [],
        "include_planned_families": bool(include_planned_families),
        "architecture_guardrail": {
            "excel_role": "probe_target_inventory_only",
            "excel_is_production_input": False,
            "catalog_schema_expansion_in_this_sprint": False,
            "feature_resolver_modified": False,
            "check_engine_modified": False,
            "safe_to_implement_checks_now": False,
        },
    }
    try:
        rules = family_rules_for_profile(probe_profile)
    except ValueError as exc:
        connection_report["errors"].append(str(exc))
        _write_json(out / "connection_report.json", connection_report)
        return 2
    if max_candidate_tables_per_family < 1 or max_sample_rows < 0:
        connection_report["errors"].append("--max-candidate-tables-per-family must be >=1 and --max-sample-rows must be >=0")
        _write_json(out / "connection_report.json", connection_report)
        return 2

    tables, inventory_report = parse_inventory(excel_inventory, inventory_json)
    classification_rows = classify_inventory_tables(tables)

    live_table_names, match_rows, fetched_by_table, header_comparison, live_sample_rows = _empty_live_outputs()
    raw_shapes: list[dict[str, Any]] = []
    live_mode = False
    if live_etabs:
        live_mode = True
        try:
            from tbdy_engine.etabs.connection import ETABSConnection, get_available_tables  # lazy import
            from tbdy_engine.providers.etabs_display_table_fetcher import select_output_for_display  # lazy import
            conn = ETABSConnection()
            ok, message = conn.connect()
            connection_report["live_etabs_connected"] = bool(ok)
            connection_report["etabs_model_available"] = bool(ok)
            if not ok:
                connection_report["errors"].append(message)
            else:
                sap = conn.get_sap()
                database_tables = sap.DatabaseTables
                if preferred_output_case:
                    connection_report["display_selection_attempted"] = True
                    selection = select_output_for_display(database_tables, preferred_output_case)
                    connection_report["display_selection_status"] = "SUCCESS" if selection.get("display_selection_success") else "FAILED"
                    connection_report["display_selection_diagnostics"] = selection
                live_table_names = list(get_available_tables(sap))
                match_rows = match_excel_to_live_tables(
                    tables,
                    live_table_names,
                    probe_profile,
                    max_candidate_tables_per_family,
                    include_planned_families=include_planned_families,
                )
                unique_live_names = []
                for row in match_rows:
                    for table_name in row.get("live_candidate_tables") or []:
                        if table_name and table_name not in unique_live_names:
                            unique_live_names.append(table_name)
                for live_name in unique_live_names:
                    try:
                        headers, rows, raw_shape = _fetch_live_table(database_tables, live_name, max_sample_rows)
                        fetched_by_table[live_name] = {"headers": headers, "rows": rows, "error": None}
                        live_sample_rows.append({"live_table_name": live_name, "sample_row_count": len(rows), "rows": rows, "row_limit": max_sample_rows})
                        raw_shapes.append(raw_shape)
                    except Exception as exc:  # pragma: no cover - live COM failures
                        fetched_by_table[live_name] = {"headers": [], "rows": [], "error": str(exc)}
                        live_sample_rows.append({"live_table_name": live_name, "sample_row_count": 0, "rows": [], "error": str(exc), "row_limit": max_sample_rows})
                header_comparison = compare_live_headers(match_rows, fetched_by_table)
        except Exception as exc:  # pragma: no cover - live environment failures
            connection_report["errors"].append(str(exc))
            live_mode = False
    else:
        # Parse-only mode is inventory-scoped.  It must not create fake rows for
        # known families absent from the Excel inventory unless explicitly asked
        # to include planned families.
        parse_rules = scoped_family_rules_for_inventory(
            probe_profile, tables, include_planned_families=include_planned_families
        )
        match_rows = []
        for r in parse_rules:
            family_tables = [t for t in tables if t.get("family_id") == r.family_id]
            planned_absent = not bool(family_tables)
            match_rows.append({
                "family_id": r.family_id,
                "group": r.group,
                "excel_table_names": [t["excel_table_name"] for t in family_tables],
                "excel_table_name": family_tables[0]["excel_table_name"] if family_tables else None,
                "live_candidate_tables": [],
                "live_table_name": None,
                "match_basis": "planned_absent" if planned_absent else "parse_only",
                "planned_absent": planned_absent,
                "planned_live_fetch_allowed": False if planned_absent else True,
                "candidate_count_before_cap": 0,
                "candidate_count_after_cap": 0,
                "candidate_truncation_applied": False,
                "semantic_review": r.semantic_review,
                "required_columns": list(r.required_columns),
                "optional_columns": list(r.optional_columns),
            })
        header_comparison = compare_live_headers(match_rows, {})

    promotion_rows = build_promotion_rows(classification_rows, match_rows, header_comparison, live_mode=live_etabs and connection_report.get("live_etabs_connected") is True)
    column_gate = column_geometry_gate(fetched_by_table)
    decision = expansion_decision_report(inventory_report, promotion_rows, column_gate, live_mode=live_etabs and connection_report.get("live_etabs_connected") is True)
    semantic_rows = [r for r in promotion_rows if r.get("recommended_status") == "SEMANTIC_REVIEW"]
    needs_rows = [r for r in promotion_rows if r.get("recommended_status") == "NEEDS_LIVE_PROBE"]
    md = build_markdown_report(decision, inventory_report, promotion_rows)

    _write_json(out / "connection_report.json", connection_report)
    _write_json(out / "excel_inventory_parse_report.json", inventory_report)
    _write_json(out / "excel_table_family_classification.json", classification_rows)
    _write_json(out / "live_available_tables.json", {"available_table_count": len(live_table_names), "tables": [{"table_name": n, "table_key": _norm(n)} for n in live_table_names]})
    _write_json(out / "excel_to_live_table_match_report.json", match_rows)
    _write_json(out / "live_header_comparison_report.json", header_comparison)
    _write_json(out / "live_sample_rows_report.json", live_sample_rows)
    _write_json(out / "source_promotion_recommendation.json", promotion_rows)
    _write_json(out / "semantic_review_sources.json", semantic_rows)
    _write_json(out / "needs_live_probe_sources.json", needs_rows)
    _write_json(out / "c13_2_expansion_decision_report.json", decision)
    _write_text(out / "C13_2_P1_EXCEL_GUIDED_LIVE_SOURCE_VERIFICATION.md", md)
    return 0 if not connection_report["errors"] else (1 if live_etabs else 0)


def build_arg_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(description="Excel-guided live ETABS source verification gate for C13.2-P1")
    source = parser.add_mutually_exclusive_group(required=True)
    source.add_argument("--excel-inventory", type=Path)
    source.add_argument("--inventory-json", type=Path)
    parser.add_argument("--out", required=True, type=Path)
    parser.add_argument("--live-etabs", action="store_true")
    parser.add_argument("--probe-profile", default="verification_gate", choices=list(PROBE_PROFILES.keys()))
    parser.add_argument("--max-candidate-tables-per-family", type=int, default=3)
    parser.add_argument("--max-sample-rows", type=int, default=50)
    parser.add_argument("--preferred-output-case", default=None)
    parser.add_argument("--include-planned-families", action="store_true", help="Report absent known families as PLANNED rows; planned rows are never live-fetched.")
    return parser


def main(argv: Sequence[str] | None = None) -> int:
    args = build_arg_parser().parse_args(argv)
    return run_probe(
        out=args.out,
        excel_inventory=args.excel_inventory,
        inventory_json=args.inventory_json,
        live_etabs=bool(args.live_etabs),
        probe_profile=str(args.probe_profile),
        max_candidate_tables_per_family=int(args.max_candidate_tables_per_family),
        max_sample_rows=int(args.max_sample_rows),
        preferred_output_case=args.preferred_output_case,
        include_planned_families=bool(args.include_planned_families),
    )


if __name__ == "__main__":
    raise SystemExit(main())
