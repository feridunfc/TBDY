#!/usr/bin/env python
"""C13.2-P3 offline ETABS input/export source inventory.

Inventory-only utility for ETABS-exported input/model-definition workbooks.
It never imports or connects to ETABS/COM, never emits engineering verdicts,
never creates CheckResult data, and never promotes stable source contracts.
"""
from __future__ import annotations

import argparse
import json
import re
import sys
from dataclasses import dataclass
from pathlib import Path
from typing import Any, Iterable, Mapping, Sequence

try:
    from openpyxl import load_workbook
except Exception as exc:  # pragma: no cover
    raise SystemExit(f"openpyxl is required to read ETABS input/export workbooks: {exc}")

SPRINT = "C13.2-P3 full input export inventory add-on"
GENERATED_BY = "tools/inventory_c13_2_p3_full_input_export.py"
EVIDENCE_KIND = "EXCEL_EXPORT_EVIDENCE_ONLY"
SAFE_TO_IMPLEMENT_CHECKS_NOW = False
CHECK_UNLOCK_ALLOWED = False

OUTPUT_FILES = (
    "workbook_inventory.json",
    "table_header_inventory.json",
    "table_sample_inventory.json",
    "source_family_fit_report.json",
    "check_engine_source_readiness_inventory.json",
    "wall_pier_story_material_chain_report.json",
    "live_probe_target_recommendations.json",
    "acceptance_policy_recommendations.json",
    "c13_2_p3_full_input_inventory_summary.json",
)

TARGET_DOMAINS = (
    "all",
    "identity",
    "materials",
    "frame",
    "beam",
    "column",
    "wall_pier",
    "slab_area",
    "story_global",
    "modal",
    "drift",
    "loads_combos",
    "design_outputs",
    "foundations_soil",
)

FORBIDDEN_IMPORT_TOKENS = ("comtypes", "win32com", "ETABSConnection", "SapModel")


def _json_default(value: Any) -> Any:
    if isinstance(value, Path):
        return str(value)
    return str(value)


def _write_json(path: Path, payload: Any) -> None:
    path.write_text(json.dumps(payload, indent=2, ensure_ascii=False, default=_json_default), encoding="utf-8")


def _cell_text(value: Any) -> str:
    if value is None:
        return ""
    text = str(value).strip()
    return re.sub(r"\s+", " ", text)


def normalize_name(value: Any) -> str:
    text = _cell_text(value).lower()
    text = text.replace("/", " ").replace("-", " ")
    text = re.sub(r"[^a-z0-9]+", "_", text)
    return re.sub(r"_+", "_", text).strip("_")


def _normalize_columns(columns: Sequence[str]) -> list[str]:
    return [normalize_name(col) for col in columns]


def _has_any(normalized_columns: set[str], names: Iterable[str]) -> bool:
    wanted = {normalize_name(name) for name in names}
    return bool(normalized_columns & wanted)


def _find_any(normalized_columns: set[str], names: Iterable[str]) -> list[str]:
    wanted = {normalize_name(name): name for name in names}
    return [wanted[key] for key in sorted(normalized_columns & set(wanted))]


def _sheet_norm(sheet_name: str) -> str:
    return normalize_name(sheet_name)


def _name_matches(sheet_name: str, aliases: Sequence[str]) -> bool:
    norm = _sheet_norm(sheet_name)
    return any(normalize_name(alias) == norm for alias in aliases)


def _name_contains(sheet_name: str, keywords: Sequence[str]) -> bool:
    norm = _sheet_norm(sheet_name)
    return any(normalize_name(keyword) in norm for keyword in keywords)


def _is_probable_unit_row(values: Sequence[str]) -> bool:
    unit_tokens = {"kn", "m", "mm", "cm", "n", "kg", "tonf", "sec", "rad", "c", "unit", "units", "%", "ratio"}
    non_empty = [normalize_name(v) for v in values if _cell_text(v)]
    if not non_empty:
        return True
    hits = sum(1 for item in non_empty if item in unit_tokens or item.endswith("_m") or item.endswith("_mm"))
    return hits >= max(1, len(non_empty) // 2)


def _header_score(values: Sequence[str]) -> int:
    non_empty = [_cell_text(v) for v in values if _cell_text(v)]
    if not non_empty:
        return -100
    normalized = [normalize_name(v) for v in non_empty]
    known = {
        "story", "name", "height", "elevation", "bselev", "material", "e1", "g12", "u12", "fc", "fy",
        "pier", "width_bottom", "width_top", "thickness_bottom", "thickness_top", "wall", "area", "unique_name",
        "label", "designsect", "analysissect", "sectprop", "section", "propname", "outputcase", "direction",
        "drift", "ratio", "mode", "period", "ux", "uy", "sumux", "sumuy", "fx", "fy", "fz", "load",
        "case", "combo", "pattern", "diaphragm", "joint", "frame", "station", "p", "v2", "v3", "m2", "m3",
    }
    score = len(non_empty) * 2
    score += sum(6 for item in normalized if item in known)
    score += sum(2 for item in normalized if any(token in item for token in ("story", "pier", "material", "section", "wall")))
    if _is_probable_unit_row(non_empty):
        score -= 20
    # Header rows usually have more unique labels than unit rows.
    score += len(set(normalized))
    return score


def detect_header_and_samples(ws: Any, max_sample_rows: int) -> dict[str, Any]:
    rows = []
    max_scan_rows = min(ws.max_row or 0, 30)
    max_scan_cols = min(ws.max_column or 0, 80)
    for row in ws.iter_rows(min_row=1, max_row=max_scan_rows, max_col=max_scan_cols, values_only=True):
        rows.append([_cell_text(v) for v in row])

    best_index = 0
    best_score = -10**9
    for idx, row in enumerate(rows):
        score = _header_score(row)
        if score > best_score:
            best_index = idx
            best_score = score

    columns = [v for v in rows[best_index] if _cell_text(v)] if rows else []
    if not columns and rows:
        columns = [_cell_text(v) for v in rows[0] if _cell_text(v)]
        best_index = 0

    sample_rows: list[dict[str, Any]] = []
    header_len = len(columns)
    start = best_index + 2
    skipped_unit_row = False
    if start - 1 < len(rows) and _is_probable_unit_row(rows[start - 1][:header_len]):
        start += 1
        skipped_unit_row = True

    for row in ws.iter_rows(min_row=start, max_col=max(header_len, 1), values_only=True):
        values = [_cell_text(v) for v in row[:header_len]]
        if not any(values):
            continue
        sample_rows.append({columns[i]: values[i] if i < len(values) else "" for i in range(header_len)})
        if len(sample_rows) >= max_sample_rows:
            break

    # ETABS exports can be very large. Full-sheet non-empty scanning is slow in
    # read-only mode, so use worksheet dimensions as a bounded inventory estimate
    # and clearly label it in notes. Header/sample detection remains bounded.
    non_empty_row_count = int(ws.max_row or 0)

    notes = ["non_empty_row_count uses worksheet max_row as bounded inventory estimate"]
    if skipped_unit_row:
        notes.append("unit row below header skipped for samples")
    if best_score < 0:
        notes.append("weak header detection confidence")

    return {
        "detected_header_row": best_index + 1 if rows else None,
        "columns": columns,
        "normalized_columns": _normalize_columns(columns),
        "non_empty_row_count": non_empty_row_count,
        "sample_rows": sample_rows,
        "sample_row_count": len(sample_rows),
        "notes": notes,
    }


@dataclass(frozen=True, slots=True)
class FamilyRule:
    family_id: str
    domain: str
    aliases: tuple[str, ...]
    keywords: tuple[str, ...]
    required_any: tuple[tuple[str, ...], ...]
    supporting: tuple[str, ...]
    semantic_role: str
    semantic_review: bool = False


FAMILY_RULES: tuple[FamilyRule, ...] = (
    FamilyRule("frame_assignments_summary", "identity", ("Frame Assignments - Summary",), ("frame assignments summary",), (("UniqueName", "Label"), ("Story",),), ("DesignSect", "AnalysisSect", "Type", "Length"), "frame identity/section context"),
    FamilyRule("frame_section_assignments", "frame", ("Frame Assignments - Section Properties",), ("frame assignments section",), (("Story", "Label", "UniqueName"), ("SectProp", "Section", "PropName"),), ("AnalysisSect", "DesignSect"), "frame section assignment context"),
    FamilyRule("frame_section_material_assignments", "frame", ("Frame Section Property Definitions - Summary",), ("frame section property definitions summary",), (("Name",), ("Material",), ("Shape",),), ("Type", "Color"), "section property material mapping"),
    FamilyRule("frame_object_geometry", "frame", ("Frame Object Connectivity",), ("frame object connectivity",), (("UniqueName", "Frame", "Label"),), ("Xi", "Yi", "Zi", "Xj", "Yj", "Zj", "PointI", "PointJ"), "frame object geometry context"),
    FamilyRule("area_assignments_summary", "slab_area", ("Area Assignments - Summary",), ("area assignments summary",), (("UniqueName", "Area", "Label"),), ("Story", "SectProp", "Pier", "Diaphragm"), "area identity/assignment context"),
    FamilyRule("area_object_connectivity", "slab_area", ("Area Object Connectivity",), ("area object connectivity",), (("UniqueName", "Area", "Label"),), ("Point1", "Point2", "Point3", "Point4", "Story"), "area object geometry context"),
    FamilyRule("wall_bays", "wall_pier", ("Wall Bays",), ("wall bays",), (("Wall", "Pier", "Name"),), ("Story", "Bay", "Label"), "wall topology context"),
    FamilyRule("wall_object_connectivity", "wall_pier", ("Wall Object Connectivity",), ("wall object connectivity",), (("Story", "Name", "Wall", "Pier", "Area"),), ("Point1", "Point2", "Point3", "Point4"), "wall object connectivity supporting evidence"),
    FamilyRule("story_definitions", "story_global", ("Story Definitions", "Story Data", "Story Definitions - Summary"), ("story definitions", "story data"), (("Story", "Name"), ("Height",),), ("Elevation", "MasterStory", "SimilarTo", "SpliceAbove", "SpliceHeight"), "story metadata/elevation-height context"),
    FamilyRule("tower_and_base_story_definition", "story_global", ("Tower and Base Story Definition",), ("tower base story definition",), (("BSElev", "Base Elevation"),), ("Tower", "Base", "Story"), "base elevation for derived story elevation policy"),
    FamilyRule("material_properties_basic_mechanical", "materials", ("Material Properties - Basic Mechanical Properties", "Mat Prop - Basic Mech Props"), ("basic mechanical", "material properties"), (("Material", "Name"), ("E1",), ("G12",), ("U12",),), ("Type", "Weight", "UnitWeight"), "raw material mechanical constants"),
    FamilyRule("concrete_material_properties", "materials", ("Material Properties - Concrete Data", "Mat Prop - Concrete Data"), ("concrete data", "material properties concrete"), (("Material", "Name"),), ("Fc", "Fcs", "fck", "E1"), "raw concrete material evidence"),
    FamilyRule("rebar_material_properties", "materials", ("Material Properties - Rebar Data", "Mat Prop - Rebar Data"), ("rebar data", "material properties rebar"), (("Material", "Name"),), ("Fy", "Fu"), "raw rebar material evidence"),
    FamilyRule("material_list_by_story", "materials", ("Material List by Story",), ("material list by story",), (("Story",), ("Material",),), ("Weight", "Volume"), "quantity/inventory context only"),
    FamilyRule("material_list_by_object_type", "materials", ("Material List by Object Type",), ("material list by object type",), (("ObjectType", "Object Type", "Type"), ("Material",),), ("Weight", "Volume"), "quantity/inventory context only"),
    FamilyRule("material_list_by_section_property", "materials", ("Material List by Section Prop", "Material List by Section Property"), ("material list by section",), (("Section", "SectProp", "PropName"), ("Material",),), ("Weight", "Volume"), "quantity/inventory context only"),
    FamilyRule("concrete_rectangular_frame_sections", "frame", ("Frame Section Property Definitions - Concrete Rectangular",), ("concrete rectangular",), (("Name",), ("t2", "Width"), ("t3", "Depth"),), ("Material", "Area"), "rectangular frame section dimensions"),
    FamilyRule("frame_section_property_summary", "frame", ("Frame Section Property Definitions - Summary",), ("frame section property definitions summary",), (("Name",),), ("Material", "Shape"), "frame section property summary"),
    FamilyRule("column_section_properties", "column", ("Column Section Properties",), ("column section",), (("Name", "Section"),), ("Material", "Width", "Depth", "t2", "t3"), "column section property context"),
    FamilyRule("beam_section_properties", "beam", ("Beam Section Properties",), ("beam section",), (("Name", "Section"),), ("Material", "Width", "Depth", "t2", "t3"), "beam section property context"),
    FamilyRule("pier_assignments", "wall_pier", ("Pier Assignments", "Area Assigns - Pier Labels"), ("pier assignments", "pier labels"), (("Story", "Area", "Label", "Pier"),), ("Pier", "UniqueName"), "area-to-pier label mapping"),
    FamilyRule("pier_section_properties", "wall_pier", ("Pier Section Properties",), ("pier section properties",), (("Story",), ("Pier",), ("Width Bottom", "Width Top", "Width"), ("Thickness Bottom", "Thickness Top", "Thickness"),), ("Material", "Section", "PropName", "WallProp"), "direct pier section geometry/material source"),
    FamilyRule("wall_section_properties", "wall_pier", ("Wall Section Properties",), ("wall section properties",), (("Name", "Section", "WallProp"),), ("Material", "Thickness"), "wall section property context"),
    FamilyRule("area_section_properties", "slab_area", ("Area Section Props - Summary", "Area Section Properties"), ("area section props", "area section properties"), (("Name", "Section", "PropName"),), ("Material", "Thickness", "Type"), "area section property context"),
    FamilyRule("area_section_assignments", "slab_area", ("Area Assigns - Sect Prop",), ("area assigns sect prop",), (("Area", "UniqueName", "Label"), ("SectProp", "Section", "PropName"),), ("Story",), "area-to-section assignment mapping"),
    FamilyRule("area_assigns_pier_labels", "wall_pier", ("Area Assigns - Pier Labels",), ("area assigns pier labels",), (("Area", "UniqueName", "Label"), ("Pier",),), ("Story",), "area-to-pier mapping"),
    FamilyRule("area_assigns_sect_prop", "slab_area", ("Area Assigns - Sect Prop",), ("area assigns sect prop",), (("Area", "UniqueName", "Label"), ("SectProp", "Section", "PropName"),), ("Story",), "area-to-section mapping"),
    FamilyRule("wall_property_def_specified", "wall_pier", ("Wall Property Def - Specified",), ("wall property def specified",), (("Name", "WallProp", "Section"),), ("Material", "Thickness"), "wall section material/thickness definition"),
    FamilyRule("modal_participating_mass", "modal", ("Modal Participating Mass Ratios",), ("modal participating mass",), (("Mode",), ("Period",), ("UX",), ("UY",),), ("SumUX", "SumUY", "UZ", "SumUZ"), "modal/global mass participation"),
    FamilyRule("modal_periods", "modal", ("Modal Periods And Frequencies", "Modal Periods"), ("modal periods",), (("Mode",), ("Period",),), ("Frequency",), "modal periods context"),
    FamilyRule("base_reactions", "story_global", ("Base Reactions",), ("base reactions",), (("OutputCase", "Case", "Load Case"), ("FX", "F1"), ("FY", "F2"),), ("FZ", "MX", "MY", "MZ"), "base reactions observed source"),
    FamilyRule("mass_summary", "modal", ("Mass Summary",), ("mass summary",), (("Mass", "Weight"),), ("Story", "Diaphragm"), "mass summary context"),
    FamilyRule("center_of_mass_rigidity", "modal", ("Centers of Mass and Rigidity", "Center of Mass and Rigidity"), ("mass rigidity",), (("Story",),), ("XCCM", "YCCM", "XCR", "YCR"), "center of mass/rigidity context"),
    FamilyRule("story_drifts", "drift", ("Story Drifts",), ("story drifts",), (("Story",), ("OutputCase", "Case"), ("Direction",), ("Drift",),), ("Label",), "story drift observed source"),
    FamilyRule("story_max_over_avg_drifts", "drift", ("Story Max Over Avg Drifts",), ("max over avg",), (("Story",), ("OutputCase", "Case"), ("Direction",), ("Ratio",),), ("MaxDrift", "AvgDrift"), "torsion irregularity candidate observed source"),
    FamilyRule("story_forces", "story_global", ("Story Forces",), ("story forces",), (("Story",),), ("OutputCase", "FX", "FY", "FZ", "MX", "MY", "MZ"), "story force observed source", True),
    FamilyRule("story_shears", "story_global", ("Story Shears",), ("story shears",), (("Story",),), ("VX", "VY", "Shear"), "story shear observed source", True),
    FamilyRule("diaphragm_drifts", "drift", ("Diaphragm Drifts",), ("diaphragm drifts",), (("Story",), ("Drift",),), ("Direction", "OutputCase"), "diaphragm drift observed source"),
    FamilyRule("load_patterns", "loads_combos", ("Load Pattern Definitions", "Load Patterns"), ("load pattern",), (("Load", "Name", "Pattern"),), ("Type", "SelfWtMult"), "load pattern inventory"),
    FamilyRule("load_cases", "loads_combos", ("Load Case Definitions", "Load Cases"), ("load case",), (("Case", "Name", "LoadCase"),), ("Type",), "load case inventory"),
    FamilyRule("load_combinations", "loads_combos", ("Load Combination Definitions", "Load Combinations"), ("load combination",), (("Combo", "Name", "Combination"),), ("Case", "SF", "ScaleFactor"), "load combination inventory"),
    FamilyRule("response_spectrum_cases", "loads_combos", ("Response Spectrum Case Definitions", "Response Spectrum Modal Info"), ("response spectrum",), (("Case", "Name"),), ("Function", "Scale"), "response spectrum case inventory"),
    FamilyRule("diaphragm_assignments", "loads_combos", ("Diaphragm Assignments", "Area Assigns - Diaphragms"), ("diaphragm assignments",), (("Diaphragm",),), ("Story", "Area"), "diaphragm assignment context"),
    FamilyRule("concrete_beam_design_results", "design_outputs", ("Conc Bm Sum - TS 500-2000R2018", "Conc Bm Flx Env - TS 500-R2018", "Conc Bm Shr Env - TS 500-R2018"), ("conc bm", "beam design"), (("Beam", "Frame", "Label"),), ("AsTop", "AsBot", "VRebar", "Combo"), "beam design output semantic-review only", True),
    FamilyRule("concrete_column_design_results", "design_outputs", ("Conc Col Sum - TS 500-2000R2018", "Conc Col PMM Env - TS 500-R2018", "Conc Col Shr Env - TS 500-R2018"), ("conc col", "column design"), (("Column", "Frame", "Label"),), ("PMMRatio", "VRebar", "Combo"), "column design output semantic-review only", True),
    FamilyRule("concrete_wall_design_results", "design_outputs", ("Pier Dgn Sum - TS 500-R2018", "Shear Wall Design Combo Data"), ("wall design", "pier dgn"), (("Pier",),), ("AsVertical", "AsHorizontal", "Combo"), "wall design output semantic-review only", True),
    FamilyRule("pier_forces", "design_outputs", ("Pier Forces",), ("pier forces",), (("Pier",),), ("P", "V2", "V3", "M2", "M3", "OutputCase"), "pier force source semantic-review only", True),
    FamilyRule("spandrel_forces", "design_outputs", ("Spandrel Forces",), ("spandrel forces",), (("Spandrel",),), ("P", "V2", "V3", "M2", "M3"), "spandrel force source semantic-review only", True),
    FamilyRule("shell_stresses", "design_outputs", ("Shell Stresses",), ("shell stresses",), (("Area", "Element"),), ("S11", "S22", "S12"), "shell stress source semantic-review only", True),
    FamilyRule("joint_reactions", "design_outputs", ("Joint Reactions",), ("joint reactions",), (("Joint", "Point"),), ("F1", "F2", "F3", "M1", "M2", "M3"), "joint reaction source semantic-review only", True),
    FamilyRule("frame_forces", "design_outputs", ("Frame Forces",), ("frame forces",), (("Frame", "Label"),), ("Station", "P", "V2", "V3", "M2", "M3"), "frame force source semantic-review only", True),
    FamilyRule("area_forces", "design_outputs", ("Area Forces",), ("area forces",), (("Area", "Element"),), ("F11", "F22", "F12", "M11", "M22"), "area force source semantic-review only", True),
    FamilyRule("point_springs", "foundations_soil", ("Point Spring Assignments", "Point Springs"), ("point spring",), (("Point", "Joint"),), ("K1", "K2", "K3"), "support/foundation spring context"),
    FamilyRule("area_springs", "foundations_soil", ("Area Spring Assignments", "Area Springs"), ("area spring",), (("Area",),), ("K", "Subgrade"), "soil/foundation area spring context"),
    FamilyRule("soil_pressure", "foundations_soil", ("Soil Pressures", "Soil Pressure"), ("soil pressure",), (("Area", "Element"),), ("Pressure", "P"), "soil pressure source semantic-review only", True),
    FamilyRule("foundation_reactions", "foundations_soil", ("Foundation Reactions",), ("foundation reactions",), (("Point", "Area", "Joint"),), ("F1", "F2", "F3"), "foundation reaction source semantic-review only", True),
    FamilyRule("uplift_candidates", "foundations_soil", ("Joint Reactions", "Soil Pressures"), ("uplift", "joint reactions", "soil pressure"), (("F3", "Pressure"),), ("Point", "Area"), "uplift candidate source semantic-review only", True),
)


def _required_found_missing(required_any: Sequence[Sequence[str]], normalized_columns: set[str]) -> tuple[list[str], list[str]]:
    found: list[str] = []
    missing: list[str] = []
    for group in required_any:
        options = tuple(group)
        match = _find_any(normalized_columns, options)
        label = " OR ".join(options)
        if match:
            found.append(label)
        else:
            missing.append(label)
    return found, missing


def _confidence(matching: bool, missing: Sequence[str], supporting_found: Sequence[str], semantic_review: bool) -> str:
    if not matching:
        return "NONE"
    if semantic_review:
        return "MEDIUM" if len(missing) <= 1 else "LOW"
    if not missing:
        return "HIGH"
    if supporting_found or len(missing) <= 1:
        return "MEDIUM"
    return "LOW"


def _family_matches_sheet(rule: FamilyRule, sheet_name: str) -> bool:
    return _name_matches(sheet_name, rule.aliases) or _name_contains(sheet_name, rule.keywords)


def build_family_fit(headers: Sequence[Mapping[str, Any]], target_domain: str) -> dict[str, Any]:
    family_reports: dict[str, Any] = {}
    for rule in FAMILY_RULES:
        if target_domain != "all" and rule.domain != target_domain:
            continue
        matching_sheets = []
        all_found: set[str] = set()
        all_missing: set[str] = set()
        all_supporting: set[str] = set()
        for sheet in headers:
            if not _family_matches_sheet(rule, str(sheet["sheet_name"])):
                continue
            norm_cols = set(sheet.get("normalized_columns", []))
            found, missing = _required_found_missing(rule.required_any, norm_cols)
            supporting = _find_any(norm_cols, rule.supporting)
            matching_sheets.append({
                "sheet_name": sheet["sheet_name"],
                "required_columns_found": found,
                "required_columns_missing": missing,
                "supporting_columns_found": supporting,
                "detected_header_row": sheet.get("detected_header_row"),
            })
            all_found.update(found)
            all_missing.update(missing)
            all_supporting.update(supporting)
        if matching_sheets:
            missing_overall = sorted(set(" OR ".join(group) for group in rule.required_any) - all_found)
        else:
            missing_overall = sorted(" OR ".join(group) for group in rule.required_any)
        confidence = _confidence(bool(matching_sheets), missing_overall, sorted(all_supporting), rule.semantic_review)
        family_reports[rule.family_id] = {
            "matching_sheets": matching_sheets,
            "required_columns_found": sorted(all_found),
            "required_columns_missing": missing_overall,
            "supporting_columns_found": sorted(all_supporting),
            "confidence": confidence,
            "evidence_kind": EVIDENCE_KIND,
            "semantic_role": rule.semantic_role,
            "readiness_hint": "SEMANTIC_REVIEW_REQUIRED" if rule.semantic_review and matching_sheets else "INVENTORY_ONLY_READY_FOR_LIVE_PROBE" if matching_sheets else "NOT_FOUND",
            "check_unlock_allowed": CHECK_UNLOCK_ALLOWED,
            "safe_to_implement_checks_now": SAFE_TO_IMPLEMENT_CHECKS_NOW,
            "notes": [
                "Excel export evidence only; live ETABS exact proof required before promotion.",
                "Design/force/result evidence remains semantic-review only." if rule.semantic_review else "No engineering check is implemented or unlocked.",
            ],
        }
    return family_reports


def _family_has_high(families: Mapping[str, Any], family_id: str) -> bool:
    row = families.get(family_id) or {}
    return row.get("confidence") in {"HIGH", "MEDIUM"} and bool(row.get("matching_sheets"))


def build_chain_report(families: Mapping[str, Any]) -> dict[str, Any]:
    story_defs = _family_has_high(families, "story_definitions")
    tower_base = _family_has_high(families, "tower_and_base_story_definition")
    material_basic = _family_has_high(families, "material_properties_basic_mechanical")
    concrete = _family_has_high(families, "concrete_material_properties")
    rebar = _family_has_high(families, "rebar_material_properties")
    pier_section = _family_has_high(families, "pier_section_properties")
    supporting_wall = any(_family_has_high(families, fid) for fid in (
        "wall_bays", "wall_object_connectivity", "area_assigns_pier_labels", "area_assigns_sect_prop", "wall_property_def_specified", "area_section_properties"
    ))
    def status(parts: Sequence[bool]) -> str:
        if all(parts):
            return "COMPLETE_FOR_OFFLINE_EVIDENCE"
        if any(parts):
            return "PARTIAL_FOR_OFFLINE_EVIDENCE"
        return "INCOMPLETE"
    return {
        "evidence_kind": EVIDENCE_KIND,
        "story_elevation_chain": {
            "story_definitions_present": story_defs,
            "tower_and_base_story_definition_present": tower_base,
            "story_or_name_column_present": bool((families.get("story_definitions") or {}).get("required_columns_found")),
            "height_column_present": "Height" in " ".join((families.get("story_definitions") or {}).get("required_columns_found", [])),
            "bselev_column_present": tower_base,
            "derived_elevation_supported": story_defs and tower_base,
            "elevation_is_direct_column_required": False,
            "chain_status": status([story_defs, tower_base]),
            "check_unlock_allowed": False,
        },
        "material_chain": {
            "basic_mechanical_present": material_basic,
            "concrete_data_present": concrete,
            "rebar_data_present": rebar,
            "mechanical_constants_are_raw_evidence_only": True,
            "chain_status": status([material_basic, concrete or rebar]),
            "check_unlock_allowed": False,
        },
        "wall_pier_chain": {
            "pier_section_properties_present": pier_section,
            "supporting_wall_mapping_present": supporting_wall,
            "direct_pier_geometry_evidence": pier_section,
            "supporting_tables_do_not_unlock_checks_alone": True,
            "chain_status": status([pier_section, supporting_wall]),
            "check_unlock_allowed": False,
        },
        "safe_to_implement_checks_now": False,
        "stable_contract_promotion": False,
    }


CHECK_AREA_DEPENDENCIES: dict[str, tuple[str, ...]] = {
    "beam_geometry": ("frame_assignments_summary", "concrete_rectangular_frame_sections", "frame_object_geometry"),
    "beam_material": ("frame_section_material_assignments", "material_properties_basic_mechanical", "concrete_material_properties"),
    "beam_longitudinal_rebar_sources": ("concrete_beam_design_results",),
    "beam_transverse_rebar_sources": ("concrete_beam_design_results",),
    "beam_flexure_design_outputs": ("concrete_beam_design_results",),
    "beam_shear_design_outputs": ("concrete_beam_design_results",),
    "beam_capacity_design_sources": ("frame_forces", "concrete_beam_design_results"),
    "beam_deflection_sources": ("story_drifts", "frame_forces"),
    "unsupported_beam_section_detection": ("frame_section_property_summary", "concrete_rectangular_frame_sections"),
    "column_geometry": ("frame_assignments_summary", "concrete_rectangular_frame_sections"),
    "column_material": ("frame_section_material_assignments", "material_properties_basic_mechanical", "concrete_material_properties"),
    "column_pmm_design_outputs": ("concrete_column_design_results",),
    "column_shear_design_outputs": ("concrete_column_design_results",),
    "column_longitudinal_rebar_sources": ("concrete_column_design_results",),
    "column_transverse_rebar_sources": ("concrete_column_design_results",),
    "column_axial_load_sources": ("frame_forces", "concrete_column_design_results"),
    "column_capacity_design_sources": ("frame_forces", "concrete_column_design_results"),
    "unsupported_column_section_detection": ("frame_section_property_summary", "concrete_rectangular_frame_sections"),
    "wall_geometry": ("pier_section_properties", "wall_section_properties", "wall_property_def_specified"),
    "wall_thickness": ("pier_section_properties", "wall_property_def_specified", "wall_section_properties"),
    "wall_material": ("pier_section_properties", "wall_property_def_specified", "material_properties_basic_mechanical"),
    "pier_identity": ("pier_assignments", "area_assigns_pier_labels"),
    "pier_section_geometry": ("pier_section_properties",),
    "pier_assignment_mapping": ("area_assigns_pier_labels", "pier_assignments"),
    "wall_area_object_mapping": ("wall_object_connectivity", "area_object_connectivity"),
    "wall_boundary_geometry_context": ("wall_bays", "wall_object_connectivity", "pier_section_properties"),
    "pier_force_sources": ("pier_forces",),
    "wall_design_output_sources": ("concrete_wall_design_results",),
    "slab_area_identity": ("area_assignments_summary", "area_object_connectivity"),
    "slab_section_properties": ("area_section_properties",),
    "slab_thickness": ("area_section_properties", "area_assigns_sect_prop"),
    "area_section_assignment": ("area_assigns_sect_prop", "area_section_assignments"),
    "diaphragm_assignment": ("diaphragm_assignments",),
    "area_force_sources": ("area_forces",),
    "punching_candidate_sources": ("joint_reactions", "area_forces", "slab_area_identity"),
    "story_height": ("story_definitions",),
    "story_elevation_or_derived_elevation": ("story_definitions", "tower_and_base_story_definition"),
    "tower_base_elevation": ("tower_and_base_story_definition",),
    "modal_mass_participation": ("modal_participating_mass",),
    "base_reactions": ("base_reactions",),
    "story_drift": ("story_drifts",),
    "story_max_over_avg_drift": ("story_max_over_avg_drifts",),
    "torsion_irregularity_candidate_sources": ("story_max_over_avg_drifts",),
    "load_patterns": ("load_patterns",),
    "load_cases": ("load_cases",),
    "load_combinations": ("load_combinations",),
    "response_spectrum_cases": ("response_spectrum_cases",),
    "combo_family_candidate_mapping": ("load_combinations", "load_cases", "response_spectrum_cases"),
}


def build_check_readiness(families: Mapping[str, Any]) -> dict[str, Any]:
    out: dict[str, Any] = {
        "evidence_kind": EVIDENCE_KIND,
        "safe_to_implement_checks_now": False,
        "check_unlock_allowed": False,
        "check_areas": {},
    }
    for area, deps in CHECK_AREA_DEPENDENCIES.items():
        apparent = [fid for fid in deps if (families.get(fid) or {}).get("matching_sheets")]
        semantic = any((families.get(fid) or {}).get("readiness_hint") == "SEMANTIC_REVIEW_REQUIRED" for fid in apparent)
        if semantic:
            readiness = "SEMANTIC_REVIEW_REQUIRED"
        elif apparent:
            readiness = "INVENTORY_ONLY_READY_FOR_LIVE_PROBE"
        else:
            readiness = "NOT_FOUND"
        missing = [fid for fid in deps if fid not in apparent]
        out["check_areas"][area] = {
            "apparent_source_tables": [sheet["sheet_name"] for fid in apparent for sheet in (families.get(fid) or {}).get("matching_sheets", [])],
            "source_families_observed": apparent,
            "required_headers_available": {fid: (families.get(fid) or {}).get("required_columns_found", []) for fid in apparent},
            "missing_headers_or_unknowns": missing,
            "readiness": readiness,
            "live_probe_required": True,
            "stable_contract_promotion_required_before_check": True,
            "check_implementation_allowed_now": False,
            "check_unlock_allowed": False,
            "notes": ["Inventory only. Live ETABS proof and stable contract promotion are required before any check implementation."],
        }
    return out


def build_live_recommendations(families: Mapping[str, Any], target_domain: str) -> dict[str, Any]:
    recommendations = []
    for rule in FAMILY_RULES:
        if target_domain != "all" and rule.domain != target_domain:
            continue
        fam = families.get(rule.family_id) or {}
        matching = fam.get("matching_sheets", [])
        if matching:
            table_names = [row["sheet_name"] for row in matching]
            priority = "HIGH" if fam.get("confidence") == "HIGH" else "MEDIUM"
            reason = "Workbook contains matching sheet/header evidence; live exact proof is required before promotion."
        else:
            table_names = list(rule.aliases[:1])
            priority = "LOW"
            reason = "Relevant future target inferred from policy; no matching workbook sheet found."
        for table_name in table_names:
            recommendations.append({
                "table_name": table_name,
                "target_domain": rule.domain,
                "target_family": rule.family_id,
                "reason": reason,
                "expected_headers": [" OR ".join(group) for group in rule.required_any],
                "supporting_headers": list(rule.supporting),
                "priority": priority,
                "check_unlock_allowed": False,
            })
    return {
        "evidence_kind": EVIDENCE_KIND,
        "live_required_before_merge": True,
        "recommendations": recommendations,
        "safe_to_implement_checks_now": False,
    }


def build_acceptance_policy() -> dict[str, Any]:
    return {
        "promote_now": False,
        "live_required_before_merge": True,
        "excel_evidence_is_not_live_proof": True,
        "safe_to_implement_checks_now": False,
        "check_unlock_allowed": False,
        "source_acceptance_classes": [
            "VERIFIED_LIVE_CANDIDATE",
            "EXCEL_EXPORT_EVIDENCE_ONLY",
            "PARTIAL_CONTEXT_ONLY",
            "NEEDS_LIVE_PROBE",
            "SEMANTIC_REVIEW",
            "NOT_FOR_CHECK_UNLOCK",
        ],
        "recommended_policy": {
            "material_properties": [
                "Excel can identify table/header candidates.",
                "Future promotion requires live exact table proof.",
                "Mechanical constants are raw evidence only.",
            ],
            "story_definitions": [
                "Story/Name + Height + Tower and Base Story Definition/BSElev supports derived elevation policy.",
                "Per-story Elevation direct column is not mandatory if derived_elevation_supported is explicit and human-reviewed.",
                "Future promotion requires live exact table proof.",
            ],
            "pier_section_properties": [
                "Pier Section Properties can be accepted as direct geometry/material source if Story + Pier + Width/Thickness columns exist.",
                "Literal Section column is not mandatory for Pier Section Properties if direct geometry exists.",
                "Future promotion requires live exact table proof.",
            ],
            "wall_pier_mapping": [
                "Wall Object Connectivity, Wall Bays, Area Assigns - Pier Labels, Area Assigns - Sect Prop, and Wall Property Def - Specified are supporting chain evidence.",
                "Alone they do not unlock checks.",
                "Future promotion requires live exact table proof and human review.",
            ],
            "design_outputs": [
                "Design/force/result tables require semantic review.",
                "Excel-export presence alone does not prove check semantics.",
                "Live exact proof and source-role classification are required before any check uses them.",
            ],
        },
        "p4_entry_criteria": [
            "C13.2-P3 live ETABS proof rerun after hotfix2",
            "material_properties VERIFIED_LIVE_CANDIDATE",
            "story_definitions VERIFIED_LIVE_CANDIDATE with derived_elevation_supported true or explicit Elevation column",
            "pier_section_properties VERIFIED_LIVE_CANDIDATE with direct_section_geometry_present true",
            "no check unlock",
            "no stable contract promotion unless separately reviewed",
        ],
        "merge_blockers": [
            "no live ETABS proof after hotfix2",
            "source output sets safe_to_implement_checks_now true",
            "any stable contract file changed",
            "any FeatureResolver/CheckEngine/report renderer changed",
            "Excel evidence treated as production input",
            "engineering verdict created from Excel",
        ],
    }


def inventory_workbook(xlsx: Path, out_dir: Path, max_sample_rows: int, target_domain: str) -> dict[str, Any]:
    wb = load_workbook(xlsx, read_only=True, data_only=True)
    out_dir.mkdir(parents=True, exist_ok=True)
    headers = []
    samples = []
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        detected = detect_header_and_samples(ws, max_sample_rows)
        header_record = {
            "sheet_name": sheet_name,
            "detected_header_row": detected["detected_header_row"],
            "columns": detected["columns"],
            "normalized_columns": detected["normalized_columns"],
            "non_empty_row_count": detected["non_empty_row_count"],
            "sample_row_count": detected["sample_row_count"],
            "notes": detected["notes"],
        }
        headers.append(header_record)
        samples.append({
            "sheet_name": sheet_name,
            "sample_rows": detected["sample_rows"],
            "max_sample_rows": max_sample_rows,
        })

    workbook_inventory = {
        "workbook_path": str(xlsx),
        "sheet_count": len(wb.sheetnames),
        "sheet_names": list(wb.sheetnames),
        "generated_by": GENERATED_BY,
        "etabs_required": False,
        "excel_production_input": False,
        "stable_contract_promotion": False,
        "checks_implemented": False,
        "safe_to_implement_checks_now": False,
    }
    families = build_family_fit(headers, target_domain)
    chain = build_chain_report(families)
    readiness = build_check_readiness(families)
    live_recs = build_live_recommendations(families, target_domain)
    policy = build_acceptance_policy()
    detected_families = sorted(fid for fid, row in families.items() if row.get("matching_sheets"))
    summary = {
        "sprint": SPRINT,
        "etabs_required": False,
        "excel_production_input": False,
        "stable_contract_promotion": False,
        "checks_implemented": False,
        "safe_to_implement_checks_now": False,
        "workbook_processed": str(xlsx),
        "sheet_count": len(wb.sheetnames),
        "families_detected": detected_families,
        "check_areas_inventory_count": len(readiness["check_areas"]),
        "recommended_live_probe_target_count": len(live_recs["recommendations"]),
        "acceptance_policy_written": True,
        "status": "INVENTORY_AND_POLICY_ONLY",
    }

    payloads = {
        "workbook_inventory.json": workbook_inventory,
        "table_header_inventory.json": {"tables": headers, "safe_to_implement_checks_now": False},
        "table_sample_inventory.json": {"tables": samples, "safe_to_implement_checks_now": False},
        "source_family_fit_report.json": {"families": families, "safe_to_implement_checks_now": False, "evidence_kind": EVIDENCE_KIND},
        "check_engine_source_readiness_inventory.json": readiness,
        "wall_pier_story_material_chain_report.json": chain,
        "live_probe_target_recommendations.json": live_recs,
        "acceptance_policy_recommendations.json": policy,
        "c13_2_p3_full_input_inventory_summary.json": summary,
    }
    for name, payload in payloads.items():
        _write_json(out_dir / name, payload)
    return summary


def parse_args(argv: Sequence[str] | None = None) -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Offline ETABS input/export workbook inventory for C13.2-P3.")
    parser.add_argument("--xlsx", required=True, help="Path to ETABS input/export workbook (.xlsx)")
    parser.add_argument("--out", required=True, help="Output directory for JSON inventory artifacts")
    parser.add_argument("--max-sample-rows", type=int, default=10)
    parser.add_argument("--target-domain", choices=TARGET_DOMAINS, default="all")
    return parser.parse_args(argv)


def main(argv: Sequence[str] | None = None) -> int:
    args = parse_args(argv)
    xlsx = Path(args.xlsx)
    if not xlsx.exists():
        print(f"Workbook not found: {xlsx}", file=sys.stderr)
        return 2
    if xlsx.suffix.lower() not in {".xlsx", ".xlsm"}:
        print(f"Expected .xlsx/.xlsm workbook, got: {xlsx}", file=sys.stderr)
        return 2
    summary = inventory_workbook(xlsx, Path(args.out), max(0, args.max_sample_rows), args.target_domain)
    print(json.dumps(summary, indent=2, ensure_ascii=False))
    return 0


if __name__ == "__main__":  # pragma: no cover
    raise SystemExit(main())
