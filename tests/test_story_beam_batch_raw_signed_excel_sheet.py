from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook, load_workbook

from tbdy_engine.design.beams.etabs_story_beam_batch_runner import (
    _raw_signed_evidence_rows,
    _write_raw_signed_evidence_sheet,
)


def test_raw_signed_evidence_rows_from_governing() -> None:
    governing = {
        "Ve_left_kN": {
            "combo": "Grav_Ult",
            "station": 0.0,
            "etabs_local_axis_component": "V2",
            "etabs_raw_signed_value": -50.006,
            "design_demand_magnitude": 50.006,
            "sign_convention": "ETABS raw signed local force is preserved; design/check demand uses positive magnitude.",
        }
    }

    rows = _raw_signed_evidence_rows(governing)

    assert rows == [
        {
            "action": "Ve_left_kN",
            "combo": "Grav_Ult",
            "station": 0.0,
            "etabs_local_axis_component": "V2",
            "etabs_raw_signed_value": -50.006,
            "design_demand_magnitude": 50.006,
            "sign_convention": "ETABS raw signed local force is preserved; design/check demand uses positive magnitude.",
        }
    ]


def test_write_raw_signed_evidence_sheet(tmp_path: Path) -> None:
    xlsx_path = tmp_path / "engine_report.xlsx"

    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Checks"
    worksheet.append(["component", "check_type", "status"])
    worksheet.append(["300", "beam_shear_ve_le_vr", "OK"])
    workbook.save(xlsx_path)

    governing = {
        "Vd_left_kN": {
            "combo": "Grav_Ult",
            "station": 0.0,
            "etabs_local_axis_component": "V2",
            "etabs_raw_signed_value": -50.006,
            "design_demand_magnitude": 50.006,
            "sign_convention": "ETABS raw signed local force is preserved; design/check demand uses positive magnitude.",
        },
        "Md_left_neg_kNm": {
            "combo": "Grav_Ult",
            "station": 0.0,
            "etabs_local_axis_component": "M3",
            "etabs_raw_signed_value": -26.194,
            "design_demand_magnitude": 26.194,
            "sign_convention": "ETABS raw signed local force is preserved; design/check demand uses positive magnitude.",
        },
    }

    _write_raw_signed_evidence_sheet(xlsx_path, governing)

    loaded = load_workbook(xlsx_path)
    assert "Checks" in loaded.sheetnames
    assert "ETABS_Raw_Evidence" in loaded.sheetnames

    sheet = loaded["ETABS_Raw_Evidence"]
    headers = [cell.value for cell in sheet[1]]
    assert headers == [
        "action",
        "combo",
        "station",
        "etabs_local_axis_component",
        "etabs_raw_signed_value",
        "design_demand_magnitude",
        "sign_convention",
    ]

    first_row = [cell.value for cell in sheet[2]]
    assert first_row[0] == "Vd_left_kN"
    assert first_row[3] == "V2"
    assert first_row[4] == -50.006
    assert first_row[5] == 50.006
