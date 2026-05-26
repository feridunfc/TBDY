from __future__ import annotations

import json
import os
from pathlib import Path

import pandas as pd
import pytest
from openpyxl import load_workbook

import tbdy_engine.etabs.providers.beam_provider as beam_provider
from tbdy_engine.etabs.providers.beam_provider import (
    BEAM_TABLE_CANDIDATES,
    BeamEtabsProvider,
    BeamProviderError,
)
from tbdy_engine.etabs.table_access import EtabsTableAccessResult, EtabsTableAccessStatus
from tbdy_engine.runner_v2 import TBDYEngineV2


BEAM_CHECK_IDS = {"beam_geometry", "beam_flexure", "beam_shear", "beam_ductility"}


def _beam_design_summary_df() -> pd.DataFrame:
    return pd.DataFrame(
        [
            {
                "Story": "S1",
                "Frame": "B1",
                "DesignSect": "B30x60",
                "Status": "OK",
                "TotTopRebar": 1200.0,
                "TotBotRebar": 1100.0,
                "TotTrnRebar": 250.0,
            }
        ]
    )


def _beam_flexure_df() -> pd.DataFrame:
    return pd.DataFrame(
        [
            {"Story": "S1", "Frame": "B1", "Location": "Middle", "OutputCase": "EQX", "M3": 80.0},
            {"Story": "S1", "Frame": "B1", "Location": "End-I", "OutputCase": "EQY", "M3": -120.0},
        ]
    )


def _beam_shear_df() -> pd.DataFrame:
    return pd.DataFrame(
        [
            {"Story": "S1", "Frame": "B1", "Location": "End-J", "VCombo": "EQX", "Shear": 55.0},
            {"Story": "S1", "Frame": "B1", "Location": "End-I", "VCombo": "EQY", "Shear": -65.0},
        ]
    )


def _ok_result(table_name: str, df: pd.DataFrame) -> EtabsTableAccessResult:
    return EtabsTableAccessResult(
        table_name=table_name,
        status=EtabsTableAccessStatus.OK,
        model_filename="mock.edb",
        row_count=int(df.shape[0]),
        column_count=int(df.shape[1]),
        df=df,
    )


def _fake_read_factory(attempted: list[str]):
    table_frames = {
        "Concrete Beam Design Summary - TS 500-2000(R2018)": _beam_design_summary_df(),
        "Concrete Beam Flexure Envelope - TS 500-2000(R2018)": _beam_flexure_df(),
        "Concrete Beam Shear Envelope - TS 500-2000(R2018)": _beam_shear_df(),
    }

    def fake_read(table_name: str) -> EtabsTableAccessResult:
        attempted.append(table_name)
        df = table_frames.get(table_name)
        if df is None:
            return EtabsTableAccessResult(
                table_name=table_name,
                status=EtabsTableAccessStatus.TABLE_UNAVAILABLE,
                model_filename="mock.edb",
                error=f"unexpected table: {table_name}",
            )
        return _ok_result(table_name, df)

    return fake_read


def _json_check_ids(report_dir: Path) -> set[str]:
    payload = json.loads((report_dir / "engine_report.json").read_text(encoding="utf-8"))
    return {str(row["check_id"]) for row in payload["checks"]}


def _excel_check_ids(report_dir: Path) -> set[str]:
    workbook = load_workbook(report_dir / "engine_report.xlsx")
    rows = list(workbook["Details"].iter_rows(values_only=True))
    header = list(rows[0])
    check_id_index = header.index("check_id")
    return {str(row[check_id_index]) for row in rows[1:] if row[check_id_index]}


def test_beam_provider_builds_context_and_engine_reports(monkeypatch, tmp_path):
    attempted: list[str] = []
    monkeypatch.setattr(beam_provider, "read_etabs_table_on_demand", _fake_read_factory(attempted))

    provider = BeamEtabsProvider()
    ctx = provider.build_context()
    result = TBDYEngineV2(ctx, report_dir=tmp_path).run()

    assert result["status"] in {"OK", "PARTIAL"}
    assert (tmp_path / "engine_report.json").exists()
    assert (tmp_path / "engine_report.xlsx").exists()
    assert BEAM_CHECK_IDS.issubset(_json_check_ids(tmp_path))
    assert BEAM_CHECK_IDS.issubset(_excel_check_ids(tmp_path))

    expected_attempts = [candidates[0] for candidates in BEAM_TABLE_CANDIDATES.values()]
    assert attempted == expected_attempts
    diagnostics = getattr(ctx, "beam_provider_diagnostics")
    assert set(diagnostics["attempted_tables"]) == set(BEAM_TABLE_CANDIDATES)
    assert diagnostics["selected_tables"] == {
        "beam_design_summary": "Concrete Beam Design Summary - TS 500-2000(R2018)",
        "beam_flexure_envelope": "Concrete Beam Flexure Envelope - TS 500-2000(R2018)",
        "beam_shear_envelope": "Concrete Beam Shear Envelope - TS 500-2000(R2018)",
    }
    assert diagnostics["missing_tables"] == {}


def test_beam_runtime_delivery_source_uses_real_runner_pipeline_only():
    source = Path(__file__).read_text(encoding="utf-8")
    forbidden_fake = "fake" + "_" + "evaluator"

    assert "TBDYEngineV2" in source
    assert ".run()" in source
    assert "_build_evaluators" not in source
    assert "CheckAdapter" not in source
    assert "ReportingFacade" not in source
    assert "JSONReporter" not in source
    assert "ExcelReporter" not in source
    assert forbidden_fake not in source


@pytest.mark.etabs_smoke
def test_live_beam_provider_runs_engine_reports_when_enabled(tmp_path):
    if os.environ.get("TBDY_RUN_ETABS_LIVE_SMOKE") != "1":
        pytest.skip("Live ETABS beam provider disabled; set TBDY_RUN_ETABS_LIVE_SMOKE=1")

    provider = BeamEtabsProvider()
    try:
        ctx = provider.build_context()
    except BeamProviderError as exc:
        pytest.skip(f"Required live beam ETABS tables missing: {exc.diagnostics}")

    result = TBDYEngineV2(ctx, report_dir=tmp_path).run()

    assert result["status"] in {"OK", "PARTIAL"}
    assert (tmp_path / "engine_report.json").exists()
    assert (tmp_path / "engine_report.xlsx").exists()
    assert BEAM_CHECK_IDS.issubset(_json_check_ids(tmp_path))
    assert BEAM_CHECK_IDS.issubset(_excel_check_ids(tmp_path))
    diagnostics = getattr(ctx, "beam_provider_diagnostics")
    assert diagnostics["attempted_tables"]
    assert set(diagnostics["selected_tables"]) == set(BEAM_TABLE_CANDIDATES)
    assert diagnostics["missing_tables"] == {}
