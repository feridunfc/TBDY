from __future__ import annotations

import json
from pathlib import Path

from openpyxl import load_workbook

import apps.streamlit_beam_design_app as app


def test_r22d_report_source_contract() -> None:
    source = Path("apps/streamlit_beam_design_app.py").read_text(encoding="utf-8-sig")

    required = [
        "Generate ETABS Design Crosscheck Report",
        "ETABS_Design_Crosscheck",
        "etabs_design_crosscheck.json",
        "etabs_design_crosscheck.md",
        "etabs_design_crosscheck.xlsx",
        "ETABS design output does not validate BeamCore",
        "does not prove TBDY compliance",
    ]

    for text in required:
        assert text in source


def test_r22d_markdown_renderer_contains_crosscheck_rows() -> None:
    payload = {
        "rows": [
            {
                "story": "+14.5",
                "label": "B4",
                "unique_name": "300",
                "section": "B40x70",
                "location": "End-I",
                "region": "left",
                "status": "DIAGNOSTIC",
                "etabs_negative_moment_kNm": -49.7361,
                "beamcore_negative_moment_kNm": 26.1939,
                "negative_moment_delta_kNm": 23.5422,
                "etabs_as_top_cm2": 2.62,
                "beamcore_top_required_cm2": 1.3104,
                "top_as_delta_cm2": 1.3096,
            }
        ]
    }

    markdown = app._r22d_render_etabs_design_crosscheck_markdown(payload)

    assert "# ETABS Design Output Crosscheck" in markdown
    assert "Diagnostic comparison only" in markdown
    assert "B4" in markdown
    assert "End-I" in markdown
    assert "DIAGNOSTIC" in markdown
    assert "2.62" in markdown


def test_r22d_write_xlsx_sheet(tmp_path: Path) -> None:
    xlsx_path = tmp_path / "etabs_design_crosscheck.xlsx"

    rows = [
        {
            "story": "+14.5",
            "label": "B4",
            "unique_name": "300",
            "section": "B40x70",
            "location": "Middle",
            "region": "middle",
            "status": "DIAGNOSTIC",
            "etabs_as_bot_cm2": 2.25,
            "beamcore_bottom_required_cm2": 2.4431,
            "bottom_as_delta_cm2": 0.1931,
            "message": "Diagnostic comparison only",
        }
    ]

    app._r22d_write_etabs_design_crosscheck_xlsx(xlsx_path, rows)

    workbook = load_workbook(xlsx_path)
    assert workbook.sheetnames == ["ETABS_Design_Crosscheck"]

    sheet = workbook["ETABS_Design_Crosscheck"]
    headers = [cell.value for cell in sheet[1]]

    assert "location" in headers
    assert "etabs_as_bot_cm2" in headers
    assert "beamcore_bottom_required_cm2" in headers

    row = [cell.value for cell in sheet[2]]
    assert row[headers.index("label")] == "B4"
    assert row[headers.index("location")] == "Middle"
    assert row[headers.index("status")] == "DIAGNOSTIC"


def test_r22d_bundle_writes_json_markdown_xlsx(tmp_path: Path, monkeypatch) -> None:
    class FakeSessionState(dict):
        pass

    class FakeStreamlit:
        session_state = FakeSessionState(
            {
                "r22c_etabs_design_crosscheck_rows": [
                    {
                        "story": "+14.5",
                        "label": "B4",
                        "unique_name": "300",
                        "section": "B40x70",
                        "location": "End-J",
                        "region": "right",
                        "status": "DIAGNOSTIC",
                        "etabs_negative_moment_kNm": -63.2114,
                        "beamcore_negative_moment_kNm": 13.538,
                        "negative_moment_delta_kNm": 49.6734,
                    }
                ]
            }
        )

    monkeypatch.setattr(app, "st", FakeStreamlit)

    result = app._write_r22d_etabs_design_crosscheck_report_bundle(tmp_path)

    assert result["status"] == "OK"

    json_path = tmp_path / "etabs_design_crosscheck.json"
    md_path = tmp_path / "etabs_design_crosscheck.md"
    xlsx_path = tmp_path / "etabs_design_crosscheck.xlsx"

    assert json_path.exists()
    assert md_path.exists()
    assert xlsx_path.exists()

    payload = json.loads(json_path.read_text(encoding="utf-8"))
    assert payload["status"] == "DIAGNOSTIC"
    assert payload["rows"][0]["location"] == "End-J"
    assert "does not validate BeamCore" in payload["claim_boundary"]

    assert "End-J" in md_path.read_text(encoding="utf-8")
