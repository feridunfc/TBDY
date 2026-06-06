from __future__ import annotations

from dataclasses import dataclass
from pathlib import Path

from openpyxl import load_workbook

from tbdy_engine.design.beams.beam_core_artifacts import _report_payload, _write_xlsx_report


@dataclass
class FakeCheck:
    component: str = "300"
    story: str = "+14.5"
    section: str = "B40x70"
    check_type: str = "beam_flexure_bottom_area_provided_ge_required"
    status: str = "OK"
    demand: float = 2.1956
    capacity: float = 10.0
    ratio: float = 0.21956
    unit: str = "cm2"
    code_ref: str = "TBDY"
    messages: tuple[str, ...] = ("ok",)
    evidence: dict[str, object] | None = None

    def __post_init__(self) -> None:
        if self.evidence is None:
            self.evidence = {
                "story": self.story,
                "section_name": self.section,
                "core_check_evidence_by_id": {
                    "300:+14.5:beam_flexure_bottom_area_provided_ge_required": {
                        "Md_kNm": 43.7621,
                        "As_required_cm2": 2.1956,
                        "provided_area_cm2": 10.0,
                    }
                },
            }


def test_r25a_json_has_central_evidence_map_and_compact_check_evidence() -> None:
    payload = _report_payload((FakeCheck(),))

    assert payload["summary"]["total"] == 1
    assert "evidence_by_id" in payload

    check = payload["checks"][0]
    assert check["evidence_ref"] == "300:+14.5:beam_flexure_bottom_area_provided_ge_required"
    assert check["ratio_type"] == "demand_over_capacity"
    assert check["pass_rule"] == "ratio <= 1.0"

    assert "core_check_evidence_by_id" not in check["evidence"]
    assert check["evidence"]["Md_kNm"] == 43.7621

    evidence_by_id = payload["evidence_by_id"]
    assert check["evidence_ref"] in evidence_by_id
    assert evidence_by_id[check["evidence_ref"]]["As_required_cm2"] == 2.1956


def test_r25a_xlsx_checks_sheet_has_ratio_and_evidence_columns(tmp_path: Path) -> None:
    path = tmp_path / "engine_report.xlsx"

    written = _write_xlsx_report((FakeCheck(),), path)

    assert written == path

    workbook = load_workbook(path, read_only=True)
    sheet = workbook["Checks"]
    headers = [cell.value for cell in next(sheet.iter_rows(min_row=1, max_row=1))]

    for header in ("id", "story", "section", "ratio_type", "pass_rule", "evidence_ref", "messages"):
        assert header in headers
