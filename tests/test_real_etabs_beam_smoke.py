from __future__ import annotations

import json
import os
from pathlib import Path

import pytest

from tbdy_engine.etabs.table_access import read_etabs_table_on_demand
from tbdy_engine.etabs.normalizers.beam_design import build_beam_context_from_tables
from tbdy_engine.runner_v2 import run_engine_v2


BEAM_DESIGN_SUMMARY_TABLE_CANDIDATES = (
    "Concrete Beam Design Summary - TS 500-2000(R2018)",
)
BEAM_FLEXURE_TABLE_CANDIDATES = (
    "Concrete Beam Flexure Envelope - TS 500-2000(R2018)",
    "Concrete Beam Flexure Envelope -  TS 500-2000(R2018)",
)
BEAM_SHEAR_TABLE_CANDIDATES = (
    "Concrete Beam Shear Envelope - TS 500-2000(R2018)",
    "Concrete Beam Shear Envelope -  TS 500-2000(R2018)",
)

EXPECTED_EXCEL_SHEETS = {"Summary", "Kiriş Kesme", "Kiriş Donatı Seçimi", "Beam Checks", "Evidence"}

FLEXURE_REBAR_AREA_KEYS = {
    "required_area",
    "area",
    "as_required",
    "as_top",
    "as_bottom",
    "AsTop",
    "AsBot",
    "total_required_area",
    "i_top_required_area",
    "j_top_required_area",
    "bottom_required_area",
}

FORBIDDEN_JSON_FIELDS = {
    "report_metadata",
    "runtime_bridge",
    "report_contract",
    "evaluation_errors",
    "evaluation_skipped",
    "execution_order",
    "cache_stats",
    "coverage",
    "distributions",
    "json_snapshot",
    "excel_snapshot",
    "action_summary",
}


def _require_real_etabs_enabled() -> None:
    if os.environ.get("RUN_REAL_ETABS_BEAM_SMOKE") != "1":
        pytest.skip("Set RUN_REAL_ETABS_BEAM_SMOKE=1 on a Windows/ETABS machine to execute the live smoke.")


def _read_first_available_table(candidate_names: tuple[str, ...]):
    attempts: list[str] = []
    for table_name in candidate_names:
        result = read_etabs_table_on_demand(table_name)
        if result.ok:
            assert result.has_data, f"{table_name} returned no rows"
            return table_name, result
        attempts.append(f"{table_name}: status={result.status} error={result.error}")
    pytest.fail("No ETABS table candidate could be read. Attempts: " + " | ".join(attempts))


def _column_values(sheet, column: int, start_row: int) -> list[object]:
    return [sheet.cell(row=row, column=column).value for row in range(start_row, sheet.max_row + 1)]


def _has_real_value(values: list[object]) -> bool:
    return any(value not in (None, "", "NO_DATA") for value in values)


def _first_flexure_row_keys(context: dict[str, object]) -> set[str]:
    design_metadata = context.get("design_metadata", {})
    grouped = design_metadata.get("beam_flexure_grouped", {}) if isinstance(design_metadata, dict) else {}
    if not isinstance(grouped, dict):
        return set()
    for group in grouped.values():
        if not isinstance(group, dict):
            continue
        row = group.get("governing_ratio") or group.get("governing_positive") or group.get("governing_negative")
        if isinstance(row, dict):
            return set(row.keys())
    return set()


def _print_first_flexure_row_keys(context: dict[str, object]) -> set[str]:
    keys = _first_flexure_row_keys(context)
    print(f"NORMALIZED_FLEXURE_ROW_KEYS={sorted(keys)}")
    return keys


@pytest.mark.real_etabs
def test_real_etabs_beam_smoke_produces_json_and_excel_reports(tmp_path: Path) -> None:
    _require_real_etabs_enabled()
    pytest.importorskip("openpyxl")

    design_summary_table, design_summary = _read_first_available_table(BEAM_DESIGN_SUMMARY_TABLE_CANDIDATES)
    flexure_table, flexure = _read_first_available_table(BEAM_FLEXURE_TABLE_CANDIDATES)
    shear_table, shear = _read_first_available_table(BEAM_SHEAR_TABLE_CANDIDATES)

    context = build_beam_context_from_tables(
        {
            "beam_design_summary": design_summary.df,
            "beam_design_summary_source_table": design_summary_table,
            "beam_flexure_envelope": flexure.df,
            "beam_flexure_envelope_source_table": flexure_table,
            "beam_shear_envelope": shear.df,
            "beam_shear_envelope_source_table": shear_table,
        }
    )
    flexure_row_keys = _print_first_flexure_row_keys(context)

    design_metadata = context.get("design_metadata", {})
    assert design_metadata.get("beam_design_summary_rows")
    assert "beam_flexure_grouped" in design_metadata
    assert "beam_shear_grouped" in design_metadata

    result = run_engine_v2(context, report_dir=tmp_path)

    assert result["reports"].keys() == {"json", "excel"}
    assert "json_snapshot" not in result["reports"]
    assert "excel_snapshot" not in result["reports"]
    assert "action_summary" not in result["reports"]

    json_path = Path(result["reports"]["json"])
    assert json_path.exists()
    assert json_path.name == "engine_report.json"

    payload = json.loads(json_path.read_text(encoding="utf-8"))
    assert set(payload) == {"summary", "checks"}
    assert FORBIDDEN_JSON_FIELDS.isdisjoint(payload)

    check_types = {row["check_type"] for row in payload["checks"]}
    assert {"beam_geometry", "beam_flexure", "beam_shear"}.issubset(check_types)

    for row in payload["checks"]:
        assert {
            "id",
            "component",
            "check_type",
            "status",
            "demand",
            "capacity",
            "ratio",
            "evidence",
            "messages",
            "story",
            "section",
            "unit",
            "code_ref",
        } == set(row)
        assert FORBIDDEN_JSON_FIELDS.isdisjoint(row)

    excel_path = Path(result["reports"]["excel"])
    assert excel_path.exists()
    assert excel_path.name == "engine_report.xlsx"

    import openpyxl

    workbook = openpyxl.load_workbook(excel_path)
    assert set(workbook.sheetnames) == EXPECTED_EXCEL_SHEETS
    assert "Eval_Skipped" not in workbook.sheetnames
    assert "Eval_Errors" not in workbook.sheetnames
    assert "Report_Contract" not in workbook.sheetnames
    assert "Details" not in workbook.sheetnames
    assert workbook["Kiriş Kesme"]["A1"].value == "KİRİŞ KESME KAPASİTE HESABI"
    assert workbook["Kiriş Donatı Seçimi"]["A1"].value == "KİRİŞ DONATI SEÇİMİ"

    shear_sheet = workbook["Kiriş Kesme"]
    assert _has_real_value(_column_values(shear_sheet, 4, 17) + _column_values(shear_sheet, 20, 17))

    flexure_sheet = workbook["Kiriş Donatı Seçimi"]
    assert flexure_sheet["A1"].value == "KİRİŞ DONATI SEÇİMİ"
    if FLEXURE_REBAR_AREA_KEYS.intersection(flexure_row_keys):
        required_area_columns = [5, 8, 11, 14, 19, 21]
        required_or_ratio_values: list[object] = []
        for column in required_area_columns:
            required_or_ratio_values.extend(_column_values(flexure_sheet, column, 16))
        assert _has_real_value(required_or_ratio_values)
    else:
        print("FLEXURE_REBAR_AREA_SOURCE=NOT_AVAILABLE_IN_NORMALIZED_DATA")
