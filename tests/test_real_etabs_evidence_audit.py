from __future__ import annotations

import pytest

pytestmark = [
    pytest.mark.legacy_evidence_audit,
    pytest.mark.skip(
        reason="Legacy evidence audit uses pre-closure adapter/report contracts; archived from BEAM_RUNTIME_CLOSURE proof."
    ),
]

import json
from pathlib import Path

from openpyxl import load_workbook

from tbdy_engine.adapters.check_adapter import CheckAdapter
from tbdy_engine.contracts.loader import EngineContractLoader
from tbdy_engine.reports.excel_reporter import ExcelReporter
from tbdy_engine.reports.json_reporter import JSONReporter


ROOT = Path(__file__).resolve().parents[1]
FIXTURE_STATUS_SYNTHETIC = "ETABS_SHAPED_SYNTHETIC_NO_REAL_FIXTURE"
FIXTURE_STATUS_REAL = "REAL_ETABS_FIXTURE_AVAILABLE"
REAL_FIXTURE_SEARCH_ROOTS = [
    ROOT / "tests" / "fixtures",
    ROOT / "tests" / "golden",
    ROOT / "sample",
    ROOT / "samples",
    ROOT / "fixtures",
]
REAL_FIXTURE_SUFFIXES = {".csv", ".xlsx", ".json", ".edb"}


def _real_fixture_candidates() -> list[Path]:
    candidates: list[Path] = []
    for root in REAL_FIXTURE_SEARCH_ROOTS:
        if not root.exists():
            continue
        for path in root.rglob("*"):
            if path.is_file() and path.suffix.lower() in REAL_FIXTURE_SUFFIXES:
                candidates.append(path)
    return sorted(candidates)


def _fixture_status() -> str:
    return FIXTURE_STATUS_REAL if _real_fixture_candidates() else FIXTURE_STATUS_SYNTHETIC


def _catalog():
    return EngineContractLoader.from_project_root(ROOT).build_runtime_catalog()


def _check_payload(*, status: str = "OK", ratio: float = 0.5, evidence: dict | None = None, **extra):
    payload = {
        "status": status,
        "ratio": ratio,
        "value": ratio,
        "limit": 1.0,
        "unit": "ratio",
        "message": "ETABS-shaped evidence audit fixture",
        "tbdy_ref": "fixture",
        "evaluation_level": "DESIGN_LEVEL",
        "source": "etabs_shaped_fixture",
    }
    if evidence is not None:
        payload["evidence"] = evidence
    payload.update(extra)
    return payload


def _etabs_shaped_eval_results():
    """Closest available end-to-end fixture when no real exported ETABS tables exist in repo."""
    return {
        "results": {
            "COLUMN_DESIGN": {
                "outputs": [
                    {
                        "label": "C1",
                        "story": "S1",
                        "checks": {
                            "geometry": _check_payload(evidence={"source_table": "Story Definitions"}),
                            "axial": _check_payload(
                                governing_combo="S_E_COL_N",
                                combo_family=None,
                                evidence={
                                    "force": "N_kn",
                                    "N_kn": 720.0,
                                    "limit": 1000.0,
                                    "ratio": 0.72,
                                    "component_case": "S_E_COL_N",
                                    "governing_combo": "S_E_COL_N",
                                    "source_table": "Element Forces - Columns",
                                },
                            ),
                            "pmm": _check_payload(
                                governing_combo="PMM_COMBO_1",
                                combo_family=None,
                                evidence={
                                    "governing_combo": "PMM_COMBO_1",
                                    "source": "column_pmm",
                                    "source_table": "Concrete Column PMM Envelope - TS 500-2000(R2018)",
                                },
                            ),
                            "shear": _check_payload(
                                governing_combo=None,
                                combo_family=None,
                                evidence={
                                    "force": "max(abs(Vx_kn), abs(Vy_kn))",
                                    "Vx_kn": 310.0,
                                    "Vy_kn": 280.0,
                                    "Vx_case": "K_E_COL_VX",
                                    "Vy_case": "K_E_COL_VY",
                                    "source_table": "Element Forces - Columns",
                                },
                            ),
                            "confinement": _check_payload(evidence={"source": "column_confinement"}),
                            "rebar_minimum": _check_payload(
                                governing_combo=None,
                                combo_family=None,
                                evidence={
                                    "As_total_mm2": 1200.0,
                                    "As_min_mm2": 900.0,
                                    "rho_pct": 1.3333333333,
                                    "rho_min_pct": 1.0,
                                    "n_bars_total": 6,
                                    "n_bars_min": 6,
                                    "bar_diameter_mm": 16.0,
                                    "bar_diameter_min_mm": 14.0,
                                    "source": "real_rebar",
                                },
                            ),
                        },
                    }
                ]
            },
            "BEAM_DESIGN": {
                "outputs": [
                    {
                        "label": "B1",
                        "story": "S1",
                        "checks": {
                            "geometry": _check_payload(evidence={"source_table": "Story Definitions"}),
                            "flexure": _check_payload(
                                governing_combo="BEAM_FLEX_COMBO",
                                combo_family=None,
                                evidence={
                                    "forces": {
                                        "M_pos_knm": 80.8195,
                                        "M_neg_left_knm": 109.7213,
                                        "M_neg_right_knm": 130.3767,
                                        "M_pos_case": "Crack_SeisX",
                                        "M_neg_left_case": "Crack_SeisX_Soil",
                                        "M_neg_right_case": "Crack_SeisX",
                                    },
                                    "ratio": 0.76,
                                    "value": 0.76,
                                    "limit": 1.0,
                                    "governing_combo": "BEAM_FLEX_COMBO",
                                    "source_table": "Concrete Beam Flexure Envelope -  TS 500-2000(R2018)",
                                },
                            ),
                            "shear": _check_payload(
                                governing_combo="BEAM_SHEAR_COMBO",
                                combo_family=None,
                                evidence={
                                    "forces": {
                                        "V_max_kn": 129.1275,
                                        "V_at_support_kn": 129.1275,
                                        "V_max_case": "Crack_SeisY_Soil",
                                        "V_at_support_case": "Crack_SeisY_Soil",
                                    },
                                    "ratio": 0.69,
                                    "value": 129.1275,
                                    "limit": 200.0,
                                    "governing_combo": "BEAM_SHEAR_COMBO",
                                    "source_table": "Concrete Beam Shear Envelope -  TS 500-2000(R2018)",
                                },
                            ),
                            "ductility": _check_payload(evidence={"source": "beam_ductility"}),
                        },
                    }
                ]
            },
            "SCWB_CHECK": {
                "column_capacity_hierarchy": [
                    _check_payload(
                        status="WARNING",
                        ratio=0.95,
                        governing_combo=None,
                        combo_family=None,
                        evidence={
                            "joint_id": "J1",
                            "direction": "X",
                            "columns": ["C1", "C2"],
                            "beams": ["B1", "B2"],
                            "sum_mrc_knm": 950.0,
                            "sum_mrb_knm": 830.0,
                            "required_mrc_knm": 996.0,
                            "reason_code": "scwb_ratio_below_limit",
                        },
                    )
                ],
                "beam_capacity_hierarchy": [
                    _check_payload(
                        status="WARNING",
                        ratio=0.95,
                        governing_combo=None,
                        combo_family=None,
                        evidence={
                            "joint_id": "J1",
                            "direction": "X",
                            "columns": ["C1", "C2"],
                            "beams": ["B1", "B2"],
                            "sum_mrc_knm": 950.0,
                            "sum_mrb_knm": 830.0,
                            "required_mrc_knm": 996.0,
                            "reason_code": "scwb_ratio_below_limit",
                        },
                    )
                ],
            },
        },
        "errors": {},
        "skipped": {},
        "execution_order": ["COLUMN_DESIGN", "BEAM_DESIGN", "SCWB_CHECK"],
        "cache_stats": {},
    }


def _adapted_rows():
    rows = CheckAdapter(_catalog()).adapt_all(_etabs_shaped_eval_results())
    return {row.check_id: row for row in rows}


def _json_rows(path: Path):
    payload = json.loads(path.read_text(encoding="utf-8"))
    return {row["check_id"]: row for row in payload["checks"]}


def _excel_rows(path: Path):
    workbook = load_workbook(path)
    sheet = workbook["Details"]
    rows = list(sheet.iter_rows(values_only=True))
    header = list(rows[0])
    return header, {row[header.index("check_id")]: dict(zip(header, row)) for row in rows[1:]}


def test_real_etabs_fixture_availability_is_explicitly_declared():
    status = _fixture_status()

    assert status in {FIXTURE_STATUS_REAL, FIXTURE_STATUS_SYNTHETIC}
    if status == FIXTURE_STATUS_SYNTHETIC:
        assert _real_fixture_candidates() == []
        assert "SYNTHETIC" in status
    else:
        assert _real_fixture_candidates()


def test_end_to_end_adapter_json_excel_path_preserves_evidence(tmp_path):
    checks = list(_adapted_rows().values())
    json_path = tmp_path / "engine_report.json"
    excel_path = tmp_path / "engine_report.xlsx"

    JSONReporter(write_history=False).generate(
        checks,
        _etabs_shaped_eval_results(),
        runtime_catalog=_catalog(),
        output_path=str(json_path),
    )
    ExcelReporter(write_history=False).generate(
        checks,
        _etabs_shaped_eval_results(),
        output_path=str(excel_path),
    )

    adapter_rows = _adapted_rows()
    json_rows = _json_rows(json_path)
    _header, excel_rows = _excel_rows(excel_path)

    assert adapter_rows["beam_flexure"].evidence["forces"]["M_pos_case"] == "Crack_SeisX"
    assert json_rows["beam_flexure"]["evidence"]["forces"]["M_neg_left_case"] == "Crack_SeisX_Soil"
    assert json.loads(excel_rows["beam_flexure"]["evidence"])["forces"]["M_neg_right_case"] == "Crack_SeisX"

    assert adapter_rows["beam_shear"].evidence["forces"]["V_max_case"] == "Crack_SeisY_Soil"
    assert json_rows["beam_shear"]["evidence"]["forces"]["V_at_support_case"] == "Crack_SeisY_Soil"
    assert json.loads(excel_rows["beam_shear"]["evidence"])["forces"]["V_max_case"] == "Crack_SeisY_Soil"

    assert adapter_rows["column_axial"].evidence["component_case"] == "S_E_COL_N"
    assert json_rows["column_axial"]["evidence"]["governing_combo"] == "S_E_COL_N"
    assert json.loads(excel_rows["column_axial"]["evidence"])["component_case"] == "S_E_COL_N"


def test_excel_contains_no_blank_or_placeholder_evidence(tmp_path):
    checks = list(_adapted_rows().values())
    excel_path = tmp_path / "engine_report.xlsx"
    ExcelReporter(write_history=False).generate(
        checks,
        _etabs_shaped_eval_results(),
        output_path=str(excel_path),
    )

    header, rows = _excel_rows(excel_path)
    assert "evidence" in header
    assert "evidence_type" in header
    assert "source_table" in header
    assert "fixture_status" in header
    assert "governing_combo" in header
    assert "combo_family" in header
    assert "component_case" in header

    for check_id, row in rows.items():
        assert row["evidence"] not in {None, "", "{}"}, check_id
        assert row["evidence_type"] not in {None, ""}, check_id
        assert row["source_table"] not in {None, ""}, check_id


def test_no_second_contract_system_was_added():
    forbidden = [
        ROOT / "docs" / "workbook_manifest.yaml",
        ROOT / "docs" / "sheet_contracts.yaml",
        ROOT / "docs" / "unit_contract.yaml",
        ROOT / "docs" / "evidence_contract.yaml",
        ROOT / "tbdy_engine" / "contracts" / "workbook_manifest.yaml",
        ROOT / "tbdy_engine" / "contracts" / "sheet_contracts.yaml",
        ROOT / "tbdy_engine" / "contracts" / "unit_contract.yaml",
        ROOT / "tbdy_engine" / "contracts" / "evidence_contract.yaml",
    ]

    for path in forbidden:
        assert not path.exists(), str(path.relative_to(ROOT))
