from __future__ import annotations

import json
from pathlib import Path

from openpyxl import load_workbook

from tbdy_engine.adapters.check_adapter import CheckResult
from tbdy_engine.reports.excel_reporter import ExcelReporter


def _check(check_id: str, evidence: dict | None) -> CheckResult:
    return CheckResult(
        check_id=check_id,
        check_name=check_id.replace("column_", "").replace("beam_", ""),
        evaluation="FIXTURE",
        status="OK",
        ratio=0.5,
        value=0.5,
        limit=1.0,
        unit="ratio",
        message="excel evidence fixture",
        tbdy_ref="fixture",
        evaluation_level="DESIGN_LEVEL",
        action="",
        source="fixture",
        element_label="E1",
        story="S1",
        severity="HIGH",
        category="FIXTURE",
        report_section="fixture",
        legacy_contract_id="",
        governing_combo=None,
        combo_family=None,
        evidence=evidence,
    )


def _details_rows(path: Path):
    wb = load_workbook(path)
    ws = wb["Details"]
    rows = list(ws.iter_rows(values_only=True))
    header = list(rows[0])
    return header, [dict(zip(header, row)) for row in rows[1:]]


def test_excel_details_sheet_includes_evidence_column(tmp_path):
    path = tmp_path / "engine_report.xlsx"
    checks = [
        _check("column_pmm", {"governing_combo": "PMM_COMBO", "source": "column_pmm"}),
        _check("column_shear", {"Vx_case": "CASE_VX", "Vy_case": "CASE_VY"}),
    ]

    generated = ExcelReporter(write_history=False).generate(
        checks,
        {"errors": {}, "skipped": {}},
        output_path=str(path),
    )

    assert generated == str(path)
    header, rows = _details_rows(path)

    assert "evidence" in header
    assert header[-1] == "evidence"
    assert rows[0]["evidence"]
    assert "PMM_COMBO" in rows[0]["evidence"]
    assert "CASE_VX" in rows[1]["evidence"]


def test_excel_evidence_column_contains_stable_json_string(tmp_path):
    path = tmp_path / "engine_report.xlsx"
    checks = [
        _check(
            "beam_flexure",
            {
                "forces": {
                    "M_pos_case": "CASE_M_POS",
                    "M_neg_left_case": "CASE_M_LEFT",
                    "M_neg_right_case": "CASE_M_RIGHT",
                },
                "governing_combo": "BEAM_FLEX_COMBO",
            },
        ),
        _check(
            "column_rebar_minimum",
            {
                "As_total_mm2": 1200.0,
                "As_min_mm2": 900.0,
                "rho_pct": 1.33,
                "source": "real_rebar",
            },
        ),
    ]

    ExcelReporter(write_history=False).generate(
        checks,
        {"errors": {}, "skipped": {}},
        output_path=str(path),
    )
    _header, rows = _details_rows(path)

    beam_evidence = rows[0]["evidence"]
    rebar_evidence = rows[1]["evidence"]

    assert json.loads(beam_evidence)["forces"]["M_pos_case"] == "CASE_M_POS"
    assert json.loads(beam_evidence)["forces"]["M_neg_left_case"] == "CASE_M_LEFT"
    assert json.loads(rebar_evidence)["As_min_mm2"] == 900.0
    assert json.loads(rebar_evidence)["source"] == "real_rebar"

    assert beam_evidence == json.dumps(checks[0].evidence, ensure_ascii=False, sort_keys=True, default=str)
    assert rebar_evidence == json.dumps(checks[1].evidence, ensure_ascii=False, sort_keys=True, default=str)


def test_excel_evidence_column_is_empty_for_missing_evidence(tmp_path):
    path = tmp_path / "engine_report.xlsx"
    checks = [_check("column_geometry", None)]

    ExcelReporter(write_history=False).generate(
        checks,
        {"errors": {}, "skipped": {}},
        output_path=str(path),
    )
    _header, rows = _details_rows(path)

    assert rows[0]["evidence"] in (None, "")


def test_excel_existing_detail_columns_are_preserved_before_evidence(tmp_path):
    path = tmp_path / "engine_report.xlsx"
    ExcelReporter(write_history=False).generate(
        [_check("beam_shear", {"forces": {"V_max_case": "CASE_V_MAX", "V_at_support_case": "CASE_V_SUPPORT"}})],
        {"errors": {}, "skipped": {}},
        output_path=str(path),
    )
    header, rows = _details_rows(path)

    assert header[:-1] == [
        "check_id", "element_label", "story", "status", "ratio", "value", "limit", "unit",
        "message", "action", "tbdy_ref", "evaluation_level", "source", "severity", "category",
        "report_section", "legacy_contract_id",
    ]
    assert header[-1] == "evidence"
    assert "CASE_V_MAX" in rows[0]["evidence"]
    assert "CASE_V_SUPPORT" in rows[0]["evidence"]
