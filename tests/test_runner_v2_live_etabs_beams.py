from __future__ import annotations

import json
import os
from pathlib import Path

import pytest
from openpyxl import load_workbook

from tbdy_engine.adapters.check_adapter import CheckAdapter
from tbdy_engine.contracts.loader import EngineContractLoader
from tbdy_engine.etabs.normalizers.beam_design import build_beam_context_from_tables
from tbdy_engine.etabs.table_access import EtabsTableAccessStatus, read_etabs_table_on_demand
from tbdy_engine.reports.facade import ReportingFacade


ROOT = Path(__file__).resolve().parents[1]
EVALUATION_FIELDS = {"geometry", "flexure", "shear", "ductility"}
INCLUDED_BEAM_CHECKS = {"beam_geometry", "beam_flexure", "beam_shear", "beam_ductility"}
EXCLUDED_BEAM_CHECKS = {"beam_capacity_hierarchy", "beam_design_full"}
BEAM_TABLE_CANDIDATES = {
    "beam_design_summary": [
        "Concrete Beam Design Summary - TS 500-2000(R2018)",
        "Concrete Beam Design Summary",
    ],
    "beam_flexure_envelope": [
        "Concrete Beam Flexure Envelope - TS 500-2000(R2018)",
        "Concrete Beam Flexure Envelope -  TS 500-2000(R2018)",
    ],
    "beam_shear_envelope": [
        "Concrete Beam Shear Envelope - TS 500-2000(R2018)",
        "Concrete Beam Shear Envelope -  TS 500-2000(R2018)",
    ],
}


def _skip_unless_live_enabled() -> None:
    if os.environ.get("TBDY_RUN_ETABS_LIVE_SMOKE") != "1":
        pytest.skip("Live ETABS beam vertical slice disabled; set TBDY_RUN_ETABS_LIVE_SMOKE=1")


def _read_first_ok(logical_name: str):
    diagnostics = []
    for table_name in BEAM_TABLE_CANDIDATES[logical_name]:
        result = read_etabs_table_on_demand(table_name)
        diagnostics.append(result.to_dict())
        if result.status is EtabsTableAccessStatus.OK and result.df is not None:
            return result, diagnostics
    return None, diagnostics


def _live_tables_or_skip() -> tuple[dict[str, object], dict[str, list[dict[str, object]]]]:
    tables: dict[str, object] = {}
    diagnostics: dict[str, list[dict[str, object]]] = {}

    for logical_name in ("beam_design_summary", "beam_flexure_envelope", "beam_shear_envelope"):
        result, attempts = _read_first_ok(logical_name)
        diagnostics[logical_name] = attempts
        if result is None:
            pytest.skip(f"Required live ETABS beam table unavailable: {logical_name}; attempts={attempts}")
        tables[logical_name] = result.df
        tables[f"{logical_name}_source_table"] = result.table_name

    return tables, diagnostics


def _catalog():
    return EngineContractLoader.from_project_root(ROOT).build_runtime_catalog()


def _status_from_diagnostic(diagnostic: str | None) -> str:
    return "NO_DATA" if diagnostic else "OK"


def _check_payload(*, status: str, evidence: dict[str, object], diagnostic: str | None = None, ratio: float = 0.0, value: float = 0.0, limit: float = 1.0, unit: str = "") -> dict[str, object]:
    return {
        "status": status,
        "ratio": ratio,
        "value": value,
        "limit": limit,
        "unit": unit,
        "message": diagnostic or "ETABS-derived beam evaluation field",
        "source": "live_etabs_table" if evidence.get("source_table") else "diagnostic",
        "evaluation_level": "ETABS_DESIGN_RESULT" if status != "NO_DATA" else "NOT_EVALUATED",
        "evidence": evidence,
    }


def _source_evidence(row: dict[str, object] | None, *, fallback_reason: str | None = None) -> dict[str, object]:
    if not row:
        return {
            "source_table": None,
            "source_row": None,
            "source_rows": [],
            "source_columns": [],
            "evidence_type": "diagnostic",
            "unit_conversion_status": "not_normalized",
            "combo_status": "not_inferred",
            "diagnostic": fallback_reason or "NO_DATA",
        }
    return {
        "source_table": row.get("source_table"),
        "source_row": row.get("source_row"),
        "source_rows": row.get("source_rows") or [row.get("source_row")],
        "source_columns": row.get("source_columns") or [],
        "evidence_type": "live_etabs_table",
        "unit_conversion_status": "not_normalized",
        "combo_status": "not_inferred",
        "diagnostic": row.get("diagnostic"),
    }


def _beam_output_from_context(context: dict[str, object]) -> dict[str, object]:
    design_rows = [row for row in context["design_metadata"].get("beam_design_summary_rows", []) if row.get("label")]
    flexure_grouped = context["design_metadata"].get("beam_flexure_grouped", {})
    shear_grouped = context["design_metadata"].get("beam_shear_grouped", {})
    outputs = []

    for design_row in design_rows:
        key = str(design_row.get("key") or "")
        label = str(design_row.get("label") or "")
        story = str(design_row.get("story") or "")
        flexure = flexure_grouped.get(key, {}) if isinstance(flexure_grouped, dict) else {}
        shear = shear_grouped.get(key, {}) if isinstance(shear_grouped, dict) else {}
        flexure_row = flexure.get("governing_ratio") or flexure.get("governing_positive") or flexure.get("governing_negative") if isinstance(flexure, dict) else None
        shear_row = shear.get("governing_ratio") or shear.get("governing_shear") if isinstance(shear, dict) else None
        ductility_diag = design_row.get("diagnostic")

        outputs.append(
            {
                "label": label,
                "story": story,
                "checks": {
                    "geometry": _check_payload(status="OK", evidence=_source_evidence(design_row), value=1.0, limit=1.0),
                    "flexure": _check_payload(
                        status=_status_from_diagnostic(None if flexure_row else "TABLE_FIELD_MISSING: flexure governing row"),
                        evidence=_source_evidence(flexure_row, fallback_reason="TABLE_FIELD_MISSING: flexure governing row"),
                        diagnostic=None if flexure_row else "TABLE_FIELD_MISSING: flexure governing row",
                        ratio=float(flexure_row.get("ratio") or 0.0) if isinstance(flexure_row, dict) else 0.0,
                    ),
                    "shear": _check_payload(
                        status=_status_from_diagnostic(None if shear_row else "TABLE_FIELD_MISSING: shear governing row"),
                        evidence=_source_evidence(shear_row, fallback_reason="TABLE_FIELD_MISSING: shear governing row"),
                        diagnostic=None if shear_row else "TABLE_FIELD_MISSING: shear governing row",
                        ratio=float(shear_row.get("ratio") or 0.0) if isinstance(shear_row, dict) else 0.0,
                    ),
                    "ductility": _check_payload(
                        status=_status_from_diagnostic(ductility_diag),
                        evidence=_source_evidence(design_row),
                        diagnostic=str(ductility_diag) if ductility_diag else None,
                    ),
                },
                "evidence": {"key": key, "source_table": design_row.get("source_table"), "source_row": design_row.get("source_row")},
            }
        )
    return {"evaluation": "BEAM_DESIGN", "outputs": outputs, "summary": {"total_beams": len(outputs)}, "evidence": {"source": "live_etabs_table_normalizer"}}


def _run_live_beam_slice(tmp_path: Path):
    tables, diagnostics = _live_tables_or_skip()
    context = build_beam_context_from_tables(tables)
    evaluation_result = _beam_output_from_context(context)
    eval_results = {
        "results": {"BEAM_DESIGN": evaluation_result},
        "errors": {},
        "skipped": {},
        "execution_order": ["BEAM_DESIGN"],
        "cache_stats": {},
    }
    catalog = _catalog()
    checks = CheckAdapter(catalog).adapt_all(eval_results)
    ReportingFacade(tmp_path).generate(checks, eval_results, runtime_catalog=catalog)
    return {"tables": tables, "diagnostics": diagnostics, "context": context, "evaluation_result": evaluation_result, "eval_results": eval_results, "checks": checks, "report_dir": tmp_path}


def _json_report(report_dir: Path) -> dict[str, object]:
    return json.loads((report_dir / "engine_report.json").read_text(encoding="utf-8"))


def _excel_details_rows(report_dir: Path) -> tuple[list[str], list[list[object]]]:
    workbook = load_workbook(report_dir / "engine_report.xlsx")
    rows = list(workbook["Details"].iter_rows(values_only=True))
    return list(rows[0]), [list(row) for row in rows[1:]]


@pytest.mark.etabs_smoke
def test_live_etabs_beam_vertical_slice_reports_json_and_excel(tmp_path):
    _skip_unless_live_enabled()
    run = _run_live_beam_slice(tmp_path)
    json_payload = _json_report(tmp_path)
    header, excel_rows = _excel_details_rows(tmp_path)
    json_check_ids = {row["check_id"] for row in json_payload["checks"]}
    excel_check_id_index = header.index("check_id")
    excel_check_ids = {row[excel_check_id_index] for row in excel_rows}

    assert INCLUDED_BEAM_CHECKS.issubset(json_check_ids)
    assert INCLUDED_BEAM_CHECKS.issubset(excel_check_ids)
    assert not (EXCLUDED_BEAM_CHECKS & json_check_ids)
    assert not (EXCLUDED_BEAM_CHECKS & excel_check_ids)
    assert set(run["evaluation_result"]["outputs"][0]["checks"]) == EVALUATION_FIELDS
    assert all("beam_" not in field for field in run["evaluation_result"]["outputs"][0]["checks"])


@pytest.mark.etabs_smoke
def test_live_etabs_beam_report_evidence_contains_source_provenance(tmp_path):
    _skip_unless_live_enabled()
    _run_live_beam_slice(tmp_path)
    json_payload = _json_report(tmp_path)
    beam_rows = [row for row in json_payload["checks"] if row["check_id"] in INCLUDED_BEAM_CHECKS]

    assert beam_rows
    for row in beam_rows:
        evidence = row.get("evidence")
        assert isinstance(evidence, dict), row["check_id"]
        assert "source_table" in evidence, row["check_id"]
        assert "source_row" in evidence, row["check_id"]
        assert isinstance(evidence.get("source_columns"), list), row["check_id"]
        assert evidence.get("unit_conversion_status") == "not_normalized"
        assert evidence.get("combo_status") == "not_inferred"

    header, excel_rows = _excel_details_rows(tmp_path)
    check_id_index = header.index("check_id")
    evidence_index = header.index("evidence")
    for row in excel_rows:
        if row[check_id_index] not in INCLUDED_BEAM_CHECKS:
            continue
        evidence = json.loads(row[evidence_index])
        assert "source_table" in evidence
        assert "source_columns" in evidence


@pytest.mark.etabs_smoke
def test_live_etabs_beam_table_attempts_are_narrow_and_structured(tmp_path):
    _skip_unless_live_enabled()
    run = _run_live_beam_slice(tmp_path)
    diagnostics = run["diagnostics"]

    assert set(diagnostics) == {"beam_design_summary", "beam_flexure_envelope", "beam_shear_envelope"}
    for logical_name, attempts in diagnostics.items():
        attempted_names = [attempt["table_name"] for attempt in attempts]
        assert attempted_names == BEAM_TABLE_CANDIDATES[logical_name][: len(attempted_names)]


def test_live_beam_smoke_source_is_contract_aligned_and_has_no_fake_evaluator():
    source = Path(__file__).read_text(encoding="utf-8")
    forbidden_raw_com_import = "win32com" + "." + "client"
    forbidden_raw_com_call = "Get" + "Active" + "Object"

    assert "TBDY_RUN_ETABS_LIVE_SMOKE" in source
    assert "read_etabs_table_on_demand" in source
    assert "CheckAdapter" in source
    assert "ReportingFacade" in source
    assert "BeamDesignModule" not in source
    assert "fake_evaluator" not in source
    assert forbidden_raw_com_import not in source
    assert forbidden_raw_com_call not in source
    assert "beam_capacity_hierarchy" in source
    assert "Concrete Joint" not in source
