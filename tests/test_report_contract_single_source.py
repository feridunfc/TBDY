from __future__ import annotations

import json
from pathlib import Path

from openpyxl import load_workbook

from tbdy_engine.adapters.check_adapter import CheckResult
from tbdy_engine.contracts.loader import EngineContractLoader
from tbdy_engine.reports.excel_reporter import ExcelReporter
from tbdy_engine.reports.json_reporter import JSONReporter
from tbdy_engine.reports.report_plan import ReportPlanner


ROOT = Path(__file__).resolve().parents[1]
EXCEL_DETAILS_HEADER = [
    "check_id",
    "element_label",
    "story",
    "status",
    "ratio",
    "value",
    "limit",
    "unit",
    "message",
    "action",
    "tbdy_ref",
    "evaluation_level",
    "source",
    "severity",
    "category",
    "report_section",
    "legacy_contract_id",
    "evidence",
]
REQUIRED_INCLUDE_FIELDS = [field for field in EXCEL_DETAILS_HEADER if field != "legacy_contract_id"]
JSON_TOP_LEVEL_KEYS = [
    "report_metadata",
    "summary",
    "checks",
    "evaluation_errors",
    "evaluation_skipped",
    "execution_order",
    "cache_stats",
    "coverage",
    "distributions",
]
FORBIDDEN_SECOND_CONTRACT_FILES = [
    ROOT / "docs" / "workbook_manifest.yaml",
    ROOT / "docs" / "sheet_contracts.yaml",
    ROOT / "docs" / "unit_contract.yaml",
    ROOT / "docs" / "evidence_contract.yaml",
    ROOT / "tbdy_engine" / "contracts" / "workbook_manifest.yaml",
    ROOT / "tbdy_engine" / "contracts" / "sheet_contracts.yaml",
    ROOT / "tbdy_engine" / "contracts" / "unit_contract.yaml",
    ROOT / "tbdy_engine" / "contracts" / "evidence_contract.yaml",
]


def _catalog():
    return EngineContractLoader.from_project_root(ROOT).build_runtime_catalog()


def _full_engine_report():
    plan = ReportPlanner(_catalog().reports).plan()
    return plan.get("full_engine_report")


def _check(*, evidence: dict | None = None, combo_family: str | None = None) -> CheckResult:
    return CheckResult(
        check_id="column_axial",
        check_name="axial",
        evaluation="COLUMN_DESIGN",
        status="OK",
        ratio=0.5,
        value=50.0,
        limit=100.0,
        unit="kN",
        message="single contract evidence fixture",
        tbdy_ref="TBDY",
        evaluation_level="DESIGN_LEVEL",
        action="none",
        source="single-contract-test",
        element_label="C1",
        story="S1",
        severity="HIGH",
        category="AXIAL",
        report_section="columns",
        legacy_contract_id="",
        governing_combo="S_E_1",
        combo_family=combo_family,
        evidence=evidence,
    )


def _details_header(path: Path) -> list[str]:
    workbook = load_workbook(path)
    sheet = workbook["Details"]
    return [cell.value for cell in sheet[1]]


def _report_contract_rows(path: Path) -> dict[str, str]:
    workbook = load_workbook(path)
    sheet = workbook["Report_Contract"]
    return {sheet.cell(row=i, column=1).value: sheet.cell(row=i, column=2).value for i in range(2, 7)}


def _details_rows(path: Path):
    workbook = load_workbook(path)
    sheet = workbook["Details"]
    rows = list(sheet.iter_rows(values_only=True))
    header = list(rows[0])
    return header, [dict(zip(header, row)) for row in rows[1:]]


def test_full_engine_report_include_fields_cover_excel_details_contract_fields():
    include_fields = set(_full_engine_report().include_fields)

    for field in [
        "value",
        "limit",
        "unit",
        "source",
        "severity",
        "category",
        "report_section",
        "evidence",
        "check_id",
        "element_label",
        "story",
        "status",
        "ratio",
        "message",
        "action",
        "tbdy_ref",
        "evaluation_level",
    ]:
        assert field in include_fields

    assert "legacy_contract_id" not in include_fields


def test_excel_details_header_is_stable_and_contract_compatible(tmp_path):
    output_path = tmp_path / "engine_report.xlsx"
    ExcelReporter(write_history=False).generate(
        [_check(evidence={"source_workbook": "fixture.xlsx"})],
        {"errors": {}, "skipped": {}},
        output_path=str(output_path),
    )

    header = _details_header(output_path)
    include_fields = set(_full_engine_report().include_fields)

    assert header == EXCEL_DETAILS_HEADER
    for field in header:
        if field == "legacy_contract_id":
            continue
        assert field in include_fields


def test_json_top_level_shape_remains_stable_and_check_evidence_is_preserved():
    evidence = {"source_workbook": "fixture.xlsx", "source_sheet": "Details"}
    payload = JSONReporter(write_history=False).build_payload(
        [_check(evidence=evidence)],
        {"errors": {}, "skipped": {}, "execution_order": [], "cache_stats": {}},
    )

    assert list(payload.keys()) == JSON_TOP_LEVEL_KEYS
    assert payload["checks"][0]["evidence"] == evidence
    assert "excel" not in payload
    assert "workbook_manifest" not in payload


def test_report_contract_sheet_uses_reports_yaml_single_source(tmp_path):
    output_path = tmp_path / "engine_report.xlsx"
    planned_report = _full_engine_report()

    ExcelReporter(write_history=False).generate(
        [_check(evidence={"source_workbook": "fixture.xlsx"})],
        {"errors": {}, "skipped": {}},
        output_path=str(output_path),
        planned_report=planned_report,
    )

    rows = _report_contract_rows(output_path)

    assert set(rows) == {"report_id", "formats", "sections", "include_fields", "metrics"}
    assert rows["report_id"] == "full_engine_report"
    assert rows["formats"] == "json,excel"
    assert "evidence" in rows["sections"]
    for field in REQUIRED_INCLUDE_FIELDS:
        assert field in rows["include_fields"].split(",")
    assert "legacy_contract_id" not in rows["include_fields"].split(",")


def test_no_second_contract_system_files_exist():
    for path in FORBIDDEN_SECOND_CONTRACT_FILES:
        assert not path.exists(), str(path.relative_to(ROOT))


def test_evidence_policy_lives_in_checkresult_evidence_not_separate_contract(tmp_path):
    evidence = {
        "source_workbook": "B-BLOK_Revised.xlsx",
        "source_sheet": "Concrete Beam Flexure Envelope -  TS 500-2000(R2018)",
        "source_row": 42,
        "source_columns": ["MomentTop", "AsTopCombo"],
        "evidence_type": "design_envelope",
        "confidence": "fixture",
        "unit_conversion_status": "not_required",
        "combo_family_status": "not_proven",
    }
    check = _check(evidence=evidence)
    json_path = tmp_path / "engine_report.json"
    excel_path = tmp_path / "engine_report.xlsx"

    JSONReporter(write_history=False).generate(
        [check],
        {"errors": {}, "skipped": {}, "execution_order": [], "cache_stats": {}},
        output_path=str(json_path),
    )
    ExcelReporter(write_history=False).generate(
        [check],
        {"errors": {}, "skipped": {}},
        output_path=str(excel_path),
    )

    json_payload = json.loads(json_path.read_text(encoding="utf-8"))
    _header, excel_rows = _details_rows(excel_path)
    excel_evidence = json.loads(excel_rows[0]["evidence"])

    assert json_payload["checks"][0]["evidence"] == evidence
    assert excel_evidence == evidence
    for path in FORBIDDEN_SECOND_CONTRACT_FILES:
        assert not path.exists(), str(path.relative_to(ROOT))


def test_no_combo_family_inference_or_uses_combo_copying():
    evidence = {
        "source_workbook": "fixture.xlsx",
        "combo_family_status": "not_proven",
    }
    payload = JSONReporter(write_history=False).build_payload(
        [_check(evidence=evidence, combo_family=None)],
        {"errors": {}, "skipped": {}, "execution_order": [], "cache_stats": {}},
    )
    row = payload["checks"][0]

    assert row["combo_family"] is None
    assert row["evidence"]["combo_family_status"] == "not_proven"
    assert "uses_combo" not in row["evidence"]
    assert "combo_family" not in row["evidence"]
