from __future__ import annotations

import json
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook

from tbdy_engine.adapters.check_adapter import CheckAdapter
from tbdy_engine.contracts.loader import EngineContractLoader
from tbdy_engine.etabs.normalizers.beam_design import (
    make_beam_diagnostic_evidence,
    make_beam_evidence,
)
from tbdy_engine.reports.facade import ReportingFacade


ROOT = Path(__file__).resolve().parents[1]
REQUIRED_EVIDENCE_KEYS = [
    "source_table",
    "source_row",
    "source_rows",
    "source_columns",
    "evidence_type",
    "confidence",
    "unit_conversion_status",
    "combo_family_status",
    "logical_table",
    "attempted_candidates",
    "notes",
]
INCLUDED_BEAM_CHECKS = {"beam_geometry", "beam_flexure", "beam_shear", "beam_ductility"}
FORBIDDEN_SECOND_CONTRACT_FILES = (
    ROOT / "docs" / "workbook_manifest.yaml",
    ROOT / "docs" / "sheet_contracts.yaml",
    ROOT / "docs" / "unit_contract.yaml",
    ROOT / "docs" / "evidence_contract.yaml",
    ROOT / "tbdy_engine" / "contracts" / "workbook_manifest.yaml",
    ROOT / "tbdy_engine" / "contracts" / "sheet_contracts.yaml",
    ROOT / "tbdy_engine" / "contracts" / "unit_contract.yaml",
    ROOT / "tbdy_engine" / "contracts" / "evidence_contract.yaml",
)
EXPECTED_EXCEL_SHEETS = ["Summary", "Details", "Eval_Skipped", "Eval_Errors", "Report_Contract"]


def _catalog():
    return EngineContractLoader.from_project_root(ROOT).build_runtime_catalog()


def _assert_standard_evidence(evidence: dict[str, object]) -> None:
    assert list(evidence) == REQUIRED_EVIDENCE_KEYS
    assert evidence["evidence_type"] in {"live_etabs_table", "diagnostic_helper"}
    assert evidence["confidence"] in {"HIGH", "MEDIUM", "LOW"}
    assert evidence["unit_conversion_status"] in {
        "not_required",
        "not_required_ratio",
        "not_normalized",
        "blocked_until_unit_contract",
        "unknown",
    }
    assert evidence["combo_family_status"] in {
        "not_applicable",
        "not_classified",
        "combo_name_present_family_unclassified",
        "heuristic_deferred",
    }
    assert isinstance(evidence["source_columns"], list)
    assert isinstance(evidence["attempted_candidates"], list)
    assert isinstance(evidence["notes"], list)


def _example_evidence(logical_table: str, table_name: str, row: int = 12) -> dict[str, object]:
    return make_beam_evidence(
        source_table=table_name,
        source_row=row,
        source_rows=[row],
        source_columns=["Beam", "Story", "Location", "Combo", "Moment"],
        logical_table=logical_table,
        attempted_candidates=[table_name],
        combo="EQX",
        notes=[],
    )


def _evaluation_results() -> dict[str, object]:
    geometry = _example_evidence("beam_design_summary", "Concrete Beam Design Summary - TS 500-2000(R2018)", 3)
    flexure = _example_evidence("beam_flexure_envelope", "Concrete Beam Flexure Envelope - TS 500-2000(R2018)", 12)
    shear = _example_evidence("beam_shear_envelope", "Concrete Beam Shear Envelope - TS 500-2000(R2018)", 8)
    ductility = make_beam_diagnostic_evidence(
        logical_table="beam_design_summary",
        reason="TABLE_FIELD_MISSING: beam design summary rebar/status fields",
        attempted_candidates=["Concrete Beam Design Summary - TS 500-2000(R2018)"],
        source_table="Concrete Beam Design Summary - TS 500-2000(R2018)",
        source_row=3,
        source_columns=["Beam", "Story", "Status"],
        confidence="LOW",
    )
    return {
        "results": {
            "BEAM_DESIGN": {
                "evaluation": "BEAM_DESIGN",
                "outputs": [
                    {
                        "label": "B23",
                        "story": "5",
                        "checks": {
                            "geometry": _check_payload("OK", geometry),
                            "flexure": _check_payload("OK", flexure),
                            "shear": _check_payload("OK", shear),
                            "ductility": _check_payload("NO_DATA", ductility, message="TABLE_FIELD_MISSING: beam design summary rebar/status fields"),
                        },
                        "evidence": {"source_table": geometry["source_table"], "source_row": geometry["source_row"]},
                    }
                ],
                "summary": {"total_beams": 1},
                "evidence": {"source": "beam evidence contract test"},
            }
        },
        "errors": {},
        "skipped": {},
        "execution_order": ["BEAM_DESIGN"],
        "cache_stats": {},
    }


def _check_payload(status: str, evidence: dict[str, object], *, message: str = "ETABS-derived beam evidence") -> dict[str, object]:
    return {
        "status": status,
        "ratio": 0.0,
        "value": 0.0,
        "limit": 1.0,
        "unit": "",
        "message": message,
        "source": "live_etabs_table" if evidence["evidence_type"] == "live_etabs_table" else "diagnostic",
        "evaluation_level": "ETABS_DESIGN_RESULT" if status != "NO_DATA" else "NOT_EVALUATED",
        "evidence": evidence,
    }


def test_example_beam_evidence_payload_shape_matches_contract():
    evidence = _example_evidence("beam_flexure_envelope", "Concrete Beam Flexure Envelope - TS 500-2000(R2018)")

    assert evidence == {
        "source_table": "Concrete Beam Flexure Envelope - TS 500-2000(R2018)",
        "source_row": 12,
        "source_rows": [12],
        "source_columns": ["Beam", "Story", "Location", "Combo", "Moment"],
        "evidence_type": "live_etabs_table",
        "confidence": "HIGH",
        "unit_conversion_status": "not_normalized",
        "combo_family_status": "combo_name_present_family_unclassified",
        "logical_table": "beam_flexure_envelope",
        "attempted_candidates": ["Concrete Beam Flexure Envelope - TS 500-2000(R2018)"],
        "notes": [],
    }


def test_diagnostic_evidence_payload_shape_matches_contract():
    evidence = make_beam_diagnostic_evidence(
        logical_table="beam_design_summary",
        reason="TABLE_FIELD_MISSING: beam design summary rebar/status fields",
        attempted_candidates=["Concrete Beam Design Summary - TS 500-2000(R2018)"],
        source_table="Concrete Beam Design Summary - TS 500-2000(R2018)",
        source_row=4,
        source_columns=["Beam", "Story"],
    )

    _assert_standard_evidence(evidence)
    assert evidence["evidence_type"] == "diagnostic_helper"
    assert evidence["confidence"] == "LOW"
    assert evidence["combo_family_status"] == "not_applicable"
    assert evidence["notes"] == ["TABLE_FIELD_MISSING: beam design summary rebar/status fields"]


def test_check_adapter_preserves_standard_beam_evidence():
    rows = CheckAdapter(_catalog()).adapt_all(_evaluation_results())
    by_id = {row.check_id: row for row in rows if row.check_id in INCLUDED_BEAM_CHECKS}

    assert set(by_id) == INCLUDED_BEAM_CHECKS
    for row in by_id.values():
        _assert_standard_evidence(row.evidence)
    assert by_id["beam_flexure"].evidence["logical_table"] == "beam_flexure_envelope"
    assert by_id["beam_ductility"].status == "NO_DATA"
    assert by_id["beam_ductility"].evidence["evidence_type"] == "diagnostic_helper"


def test_json_and_excel_reports_preserve_standard_beam_evidence(tmp_path):
    catalog = _catalog()
    eval_results = _evaluation_results()
    rows = CheckAdapter(catalog).adapt_all(eval_results)

    ReportingFacade(tmp_path).generate(rows, eval_results, runtime_catalog=catalog)

    json_payload = json.loads((tmp_path / "engine_report.json").read_text(encoding="utf-8"))
    json_rows = [row for row in json_payload["checks"] if row["check_id"] in INCLUDED_BEAM_CHECKS]
    assert {row["check_id"] for row in json_rows} == INCLUDED_BEAM_CHECKS
    for row in json_rows:
        _assert_standard_evidence(row["evidence"])

    workbook = load_workbook(tmp_path / "engine_report.xlsx")
    assert workbook.sheetnames == EXPECTED_EXCEL_SHEETS
    details_rows = list(workbook["Details"].iter_rows(values_only=True))
    header = list(details_rows[0])
    check_id_index = header.index("check_id")
    evidence_index = header.index("evidence")
    excel_rows = [row for row in details_rows[1:] if row[check_id_index] in INCLUDED_BEAM_CHECKS]
    assert {row[check_id_index] for row in excel_rows} == INCLUDED_BEAM_CHECKS
    for row in excel_rows:
        evidence = json.loads(row[evidence_index])
        _assert_standard_evidence(evidence)


def test_beam_design_output_still_uses_evaluation_fields_only():
    output = _evaluation_results()["results"]["BEAM_DESIGN"]["outputs"][0]

    assert set(output["checks"]) == {"geometry", "flexure", "shear", "ductility"}
    assert not any(name.startswith("beam_") for name in output["checks"])


def test_no_second_contract_files_exist():
    for path in FORBIDDEN_SECOND_CONTRACT_FILES:
        assert not path.exists(), str(path.relative_to(ROOT))
