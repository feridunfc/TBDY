from __future__ import annotations

import ast
import json
from pathlib import Path
from typing import Mapping

from openpyxl import load_workbook

import tbdy_engine.runner_v2 as runner_v2
from tbdy_engine.runner_v2 import TBDYEngineV2


ROOT = Path(__file__).resolve().parents[1]
FORBIDDEN_IMPORT_PREFIXES = (
    "tbdy_engine.design",
    "tbdy_engine.etabs",
    "tbdy_engine.engine.context_builder",
)


def _engine(tmp_path: Path) -> TBDYEngineV2:
    return TBDYEngineV2(ctx={"fixture": "context"}, report_dir=tmp_path)


def _column_payload() -> Mapping[str, object]:
    return {
        "outputs": [
            {
                "label": "C1",
                "story": "S1",
                "checks": {
                    "geometry": {
                        "status": "OK",
                        "ratio": 0.10,
                        "value": 0.10,
                        "limit": 1.0,
                        "unit": "ratio",
                        "message": "execute/report fixture",
                        "source": "runner_v2_execute_report_test",
                        "evaluation_level": "DESIGN_LEVEL",
                        "evidence": {
                            "source": "runner_v2_execute_report_test",
                            "evidence_type": "diagnostic_helper",
                            "confidence": "HIGH",
                            "unit_conversion_status": "not_required",
                            "combo_family_status": "not_applicable",
                        },
                    }
                },
            }
        ]
    }


def _ok_evaluator(context: object) -> Mapping[str, object]:
    return _column_payload()


def _failing_evaluator(context: object) -> Mapping[str, object]:
    raise RuntimeError("boom")


def _patch_evaluators(monkeypatch, mapping: dict[str, object]) -> None:
    monkeypatch.setattr(
        runner_v2.TBDYEngineV2,
        "_build_evaluators",
        lambda self, catalog: mapping,
    )


def _json_payload(report_dir: Path) -> dict[str, object]:
    return json.loads((report_dir / "engine_report.json").read_text(encoding="utf-8"))


def _details_rows(report_dir: Path) -> tuple[list[str], list[list[object]]]:
    workbook = load_workbook(report_dir / "engine_report.xlsx")
    sheet = workbook["Details"]
    rows = list(sheet.iter_rows(values_only=True))
    return list(rows[0]), [list(row) for row in rows[1:]]


def test_run_executes_fake_evaluator_and_writes_reports(tmp_path, monkeypatch):
    _patch_evaluators(monkeypatch, {"COLUMN_DESIGN": _ok_evaluator})

    result = _engine(tmp_path).run()

    assert (tmp_path / "engine_report.json").exists()
    assert (tmp_path / "engine_report.xlsx").exists()
    assert result["status"] in {"OK", "PARTIAL"}
    assert set(result["reports"]) == {
        "json",
        "json_snapshot",
        "excel",
        "excel_snapshot",
        "action_summary",
    }
    assert result["reports"]["json"].endswith("engine_report.json")
    assert result["reports"]["excel"].endswith("engine_report.xlsx")


def test_json_report_contains_adapted_check_and_evidence(tmp_path, monkeypatch):
    _patch_evaluators(monkeypatch, {"COLUMN_DESIGN": _ok_evaluator})

    _engine(tmp_path).run()
    payload = _json_payload(tmp_path)

    assert set(payload) >= {
        "report_metadata",
        "summary",
        "checks",
        "evaluation_errors",
        "evaluation_skipped",
        "execution_order",
        "cache_stats",
        "coverage",
        "distributions",
    }
    column_checks = [check for check in payload["checks"] if check["check_id"] == "column_geometry"]
    assert column_checks
    assert column_checks[0]["evidence"]["evidence_type"] == "diagnostic_helper"


def test_excel_report_contains_details_evidence_column(tmp_path, monkeypatch):
    _patch_evaluators(monkeypatch, {"COLUMN_DESIGN": _ok_evaluator})

    _engine(tmp_path).run()
    workbook = load_workbook(tmp_path / "engine_report.xlsx")
    assert set(workbook.sheetnames) >= {"Summary", "Details", "Eval_Skipped", "Eval_Errors"}

    header, rows = _details_rows(tmp_path)
    assert "evidence" in header
    check_id_index = header.index("check_id")
    evidence_index = header.index("evidence")
    matching_rows = [row for row in rows if row[check_id_index] == "column_geometry"]

    assert matching_rows
    evidence = json.loads(matching_rows[0][evidence_index])
    assert evidence["evidence_type"] == "diagnostic_helper"


def test_skipped_evaluator_reaches_eval_results_and_reports(tmp_path, monkeypatch):
    _patch_evaluators(monkeypatch, {"COLUMN_DESIGN": _ok_evaluator})

    result = _engine(tmp_path).run()
    payload = _json_payload(tmp_path)
    workbook = load_workbook(tmp_path / "engine_report.xlsx")
    skipped_sheet = workbook["Eval_Skipped"]

    assert result["evaluation_skipped"]
    assert payload["evaluation_skipped"]
    assert skipped_sheet.max_row > 1


def test_evaluator_exception_yields_partial_error_report_and_continues(tmp_path, monkeypatch):
    _patch_evaluators(
        monkeypatch,
        {
            "COLUMN_DESIGN": _ok_evaluator,
            "BEAM_DESIGN": _failing_evaluator,
        },
    )

    result = _engine(tmp_path).run()
    payload = _json_payload(tmp_path)
    workbook = load_workbook(tmp_path / "engine_report.xlsx")
    errors_sheet = workbook["Eval_Errors"]

    assert result["status"] == "PARTIAL"
    assert payload["evaluation_errors"].get("BEAM_DESIGN") == "boom"
    assert errors_sheet.max_row > 1
    assert any(row[0].value == "BEAM_DESIGN" and row[1].value == "boom" for row in errors_sheet.iter_rows(min_row=2))
    assert any(check["check_id"] == "column_geometry" for check in payload["checks"])


def test_execution_order_is_preserved_in_json_report(tmp_path, monkeypatch):
    _patch_evaluators(monkeypatch, {"COLUMN_DESIGN": _ok_evaluator})

    _engine(tmp_path).run()
    payload = _json_payload(tmp_path)

    assert isinstance(payload["execution_order"], list)
    assert "COLUMN_DESIGN" in payload["execution_order"]


def test_dry_run_remains_plan_only(tmp_path, monkeypatch):
    def forbidden_run(self, context, *, enabled_only=True):
        raise AssertionError("dry_run must not call RuntimeScheduler.run")

    def forbidden_build_evaluators(self, catalog):
        raise AssertionError("dry_run must not build evaluators")

    monkeypatch.setattr(runner_v2.RuntimeScheduler, "run", forbidden_run)
    monkeypatch.setattr(TBDYEngineV2, "_build_evaluators", forbidden_build_evaluators)

    _engine(tmp_path).dry_run()

    assert not (tmp_path / "engine_report.json").exists()
    assert not (tmp_path / "engine_report.xlsx").exists()


def test_report_filenames_unchanged(tmp_path, monkeypatch):
    _patch_evaluators(monkeypatch, {"COLUMN_DESIGN": _ok_evaluator})

    result = _engine(tmp_path).run()

    assert Path(result["reports"]["json"]).name == "engine_report.json"
    assert Path(result["reports"]["excel"]).name == "engine_report.xlsx"
    assert (tmp_path / "engine_report.json").exists()
    assert (tmp_path / "engine_report.xlsx").exists()


def test_no_forbidden_imports_or_schema_drift_sources():
    for relative_path in ("tbdy_engine/runner_v2.py", "tbdy_engine/runtime/scheduler.py"):
        tree = ast.parse((ROOT / relative_path).read_text(encoding="utf-8"))
        imported_modules: list[str] = []
        for node in ast.walk(tree):
            if isinstance(node, ast.Import):
                imported_modules.extend(alias.name for alias in node.names)
            elif isinstance(node, ast.ImportFrom) and node.module:
                imported_modules.append(node.module)
        forbidden_imports = sorted(
            module_name
            for module_name in imported_modules
            if any(
                module_name == prefix or module_name.startswith(prefix + ".")
                for prefix in FORBIDDEN_IMPORT_PREFIXES
            )
        )
        assert forbidden_imports == []

    combined = (ROOT / "tbdy_engine" / "runner_v2.py").read_text(encoding="utf-8")
    combined += "\n" + (ROOT / "tbdy_engine" / "runtime" / "scheduler.py").read_text(encoding="utf-8")
    assert "combo_family" not in combined
    assert "uses_combo" not in combined
    assert "message_text" not in combined
    assert ".message" not in combined
