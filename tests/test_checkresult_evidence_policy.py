from __future__ import annotations

import json
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from tbdy_engine.adapters.check_adapter import CheckResult
from tbdy_engine.reports.excel_reporter import ExcelReporter
from tbdy_engine.reports.json_reporter import JSONReporter


ROOT = Path(__file__).resolve().parents[1]

REQUIRED_PROVENANCE_FIELDS = {
    "source_workbook",
    "source_sheet",
    "source_row",
    "source_columns",
}
REQUIRED_AUDIT_FIELDS = {
    "evidence_type",
    "confidence",
    "unit_conversion_status",
    "combo_family_status",
}
ALLOWED_EVIDENCE_TYPES = {
    "raw_excel_table",
    "clean_excel_data_table",
    "excel_computed",
    "excel_computed_with_raw_data_link",
    "lookup_table",
    "diagnostic_helper",
    "heuristic",
    "etabs_shaped_synthetic",
}
ALLOWED_CONFIDENCE = {"HIGH", "MEDIUM", "LOW"}
ALLOWED_UNIT_CONVERSION_STATUS = {
    "not_required",
    "not_required_ratio",
    "converted_with_audit",
    "unknown",
    "blocked_until_unit_contract",
    "mixed_requires_contract",
}
ALLOWED_COMBO_FAMILY_STATUS = {
    "explicit",
    "not_classified",
    "combo_name_present_family_unclassified",
    "response_spectrum_candidate_unclassified",
    "heuristic_deferred",
    "not_applicable",
}
FORBIDDEN_SECOND_CONTRACT_FILES = [
    ROOT / "docs" / "workbook_manifest.yaml",
    ROOT / "docs" / "sheet_contracts.yaml",
    ROOT / "docs" / "unit_contract.yaml",
    ROOT / "docs" / "evidence_contract.yaml",
    ROOT / "tbdy_engine" / "contracts" / "workbook_manifest.yaml",
    ROOT / "tbdy_engine" / "contracts" / "sheet_contracts.yaml",
    ROOT / "tbdy_engine" / "contracts" / "unit_contract.yaml",
    ROOT / "tbdy_engine" / "contracts" / "evidence_contract.yaml",
]


def _validate_evidence_policy(evidence: dict[str, Any]) -> tuple[bool, list[str]]:
    errors: list[str] = []

    for field in sorted(REQUIRED_PROVENANCE_FIELDS | REQUIRED_AUDIT_FIELDS):
        if field not in evidence:
            errors.append(f"missing:{field}")

    evidence_type = evidence.get("evidence_type")
    if evidence_type not in ALLOWED_EVIDENCE_TYPES:
        errors.append(f"invalid:evidence_type:{evidence_type}")

    confidence = evidence.get("confidence")
    if confidence not in ALLOWED_CONFIDENCE:
        errors.append(f"invalid:confidence:{confidence}")

    unit_status = evidence.get("unit_conversion_status")
    if unit_status not in ALLOWED_UNIT_CONVERSION_STATUS:
        errors.append(f"invalid:unit_conversion_status:{unit_status}")

    combo_status = evidence.get("combo_family_status")
    if combo_status not in ALLOWED_COMBO_FAMILY_STATUS:
        errors.append(f"invalid:combo_family_status:{combo_status}")

    source_row = evidence.get("source_row")
    notes = evidence.get("notes") or []
    if source_row is None and not notes:
        errors.append("source_row_none_requires_notes")

    source_columns = evidence.get("source_columns")
    if "source_columns" in evidence and not isinstance(source_columns, list):
        errors.append("source_columns_must_be_list")

    combo_family = evidence.get("combo_family")
    if combo_family not in (None, "") and combo_status not in {"explicit", "heuristic_deferred"}:
        errors.append("combo_family_requires_explicit_or_heuristic_deferred_status")

    if combo_status == "combo_name_present_family_unclassified" and combo_family not in (None, ""):
        errors.append("unclassified_combo_name_must_not_set_combo_family")

    if "message" in evidence or "message_text" in evidence:
        errors.append("message_text_must_not_be_evidence")

    return not errors, errors


def _valid_workbook_evidence(**overrides: Any) -> dict[str, Any]:
    evidence: dict[str, Any] = {
        "source_workbook": "08_KIRIS_KESME_KONTROLLERI.xlsx",
        "source_sheet": "KIRIS_KESME_KONTROL_OZET",
        "source_row": 7,
        "source_columns": ["Kat", "Kiriş", "Tasarım Kuvveti (Vd)", "Vr", "Oran", "Durum"],
        "evidence_type": "excel_computed_with_raw_data_link",
        "confidence": "HIGH",
        "unit_conversion_status": "unknown",
        "combo_name": "ENVE_EUD-EUHD",
        "combo_family": None,
        "combo_family_status": "combo_name_present_family_unclassified",
        "notes": ["combo family intentionally not inferred"],
    }
    evidence.update(overrides)
    return evidence


def _check(check_id: str = "column_axial", evidence: dict[str, Any] | None = None, combo_family: str | None = None) -> CheckResult:
    return CheckResult(
        check_id=check_id,
        check_name=check_id,
        evaluation="EVIDENCE_POLICY_TEST",
        status="OK",
        ratio=0.5,
        value=0.5,
        limit=1.0,
        unit="ratio",
        message="evidence policy fixture",
        tbdy_ref="fixture",
        evaluation_level="DESIGN_LEVEL",
        action="none",
        source="evidence-policy-test",
        element_label="E1",
        story="S1",
        severity="HIGH",
        category="EVIDENCE",
        report_section="audit",
        legacy_contract_id="",
        governing_combo=None,
        combo_family=combo_family,
        evidence=evidence,
    )


def _excel_evidence(path: Path) -> dict[str, Any]:
    workbook = load_workbook(path)
    sheet = workbook["Details"]
    rows = list(sheet.iter_rows(values_only=True))
    header = list(rows[0])
    evidence_value = dict(zip(header, rows[1]))["evidence"]
    return json.loads(evidence_value)


def test_valid_workbook_evidence_policy_payload_passes_local_validator():
    evidence = _valid_workbook_evidence()

    valid, errors = _validate_evidence_policy(evidence)

    assert valid is True
    assert errors == []


def test_invalid_enum_values_fail_validator():
    for field, value in [
        ("evidence_type", "spreadsheet_magic"),
        ("confidence", "CERTAIN"),
        ("unit_conversion_status", "converted_silently"),
        ("combo_family_status", "guessed"),
    ]:
        evidence = _valid_workbook_evidence(**{field: value})
        valid, errors = _validate_evidence_policy(evidence)

        assert valid is False
        assert any(error.startswith(f"invalid:{field}:") for error in errors)


def test_combo_family_cannot_be_inferred_from_combo_name():
    invalid = _valid_workbook_evidence(
        combo_name="ENVE_EUD-EUHD",
        combo_family="S_E",
        combo_family_status="combo_name_present_family_unclassified",
    )
    valid_invalid, errors = _validate_evidence_policy(invalid)

    assert valid_invalid is False
    assert "unclassified_combo_name_must_not_set_combo_family" in errors

    valid = _valid_workbook_evidence(
        combo_name="ENVE_EUD-EUHD",
        combo_family=None,
        combo_family_status="combo_name_present_family_unclassified",
    )
    valid_valid, valid_errors = _validate_evidence_policy(valid)

    assert valid_valid is True
    assert valid_errors == []


def test_explicit_combo_family_requires_explicit_status():
    explicit = _valid_workbook_evidence(
        combo_name="SOME_EXPLICIT_COMBO",
        combo_family="S_E",
        combo_family_status="explicit",
    )
    valid_explicit, explicit_errors = _validate_evidence_policy(explicit)

    assert valid_explicit is True
    assert explicit_errors == []

    invalid = _valid_workbook_evidence(
        combo_family="S_E",
        combo_family_status="not_classified",
    )
    valid_invalid, errors = _validate_evidence_policy(invalid)

    assert valid_invalid is False
    assert "combo_family_requires_explicit_or_heuristic_deferred_status" in errors


def test_source_row_none_requires_explanatory_note():
    valid = _valid_workbook_evidence(
        source_row=None,
        notes=["summary-level evidence; no single source row"],
    )
    valid_result, valid_errors = _validate_evidence_policy(valid)

    assert valid_result is True
    assert valid_errors == []

    invalid = _valid_workbook_evidence(source_row=None, notes=[])
    invalid_result, errors = _validate_evidence_policy(invalid)

    assert invalid_result is False
    assert "source_row_none_requires_notes" in errors


def test_reporter_roundtrip_preserves_policy_fields(tmp_path):
    evidence = _valid_workbook_evidence(
        raw_unit="kN",
        normalized_unit="kN",
        sample_values={"Vd": 123.4, "Vr": 180.0},
    )
    check = _check(evidence=evidence, combo_family=None)
    json_path = tmp_path / "engine_report.json"
    excel_path = tmp_path / "engine_report.xlsx"

    JSONReporter(write_history=False).generate(
        [check],
        {"errors": {}, "skipped": {}, "execution_order": [], "cache_stats": {}},
        output_path=str(json_path),
    )
    ExcelReporter(write_history=False).generate(
        [check],
        {"errors": {}, "skipped": {}},
        output_path=str(excel_path),
    )

    json_payload = json.loads(json_path.read_text(encoding="utf-8"))
    json_evidence = json_payload["checks"][0]["evidence"]
    excel_evidence = _excel_evidence(excel_path)

    assert json_evidence == evidence
    assert excel_evidence == evidence
    assert json_payload["checks"][0]["combo_family"] is None
    assert "uses_combo" not in json_evidence
    assert "uses_combo" not in excel_evidence


def test_no_second_contract_files_exist():
    for path in FORBIDDEN_SECOND_CONTRACT_FILES:
        assert not path.exists(), str(path.relative_to(ROOT))


def test_real_etabs_evidence_audit_synthetic_marker_remains_explicit():
    from tests import test_real_etabs_evidence_audit as audit

    assert audit.FIXTURE_STATUS_SYNTHETIC == "ETABS_SHAPED_SYNTHETIC_NO_REAL_FIXTURE"
    assert audit.FIXTURE_STATUS_REAL == "REAL_ETABS_FIXTURE_AVAILABLE"

    status = audit._fixture_status()
    assert status in {audit.FIXTURE_STATUS_SYNTHETIC, audit.FIXTURE_STATUS_REAL}
    if status == audit.FIXTURE_STATUS_SYNTHETIC:
        assert audit._real_fixture_candidates() == []
    else:
        assert audit._real_fixture_candidates()


def test_reports_yaml_does_not_define_workbook_specific_contract():
    reports_yaml = (ROOT / "tbdy_engine" / "contracts" / "reports.yaml").read_text(encoding="utf-8")
    checks_yaml = (ROOT / "tbdy_engine" / "contracts" / "checks.yaml").read_text(encoding="utf-8")

    for forbidden_key in ["workbook_manifest", "sheet_contracts", "unit_contract", "evidence_contract"]:
        assert forbidden_key not in reports_yaml

    assert "evidence" in reports_yaml
    assert "full_engine_report" in reports_yaml
    assert "column_axial" not in reports_yaml
    assert "beam_shear" not in reports_yaml
    assert "column_axial" in checks_yaml
    assert "beam_shear" in checks_yaml


def test_evidence_policy_examples_align_with_checkresult_visible_fields(tmp_path):
    checks = [
        _check(
            "column_axial",
            _valid_workbook_evidence(
                source_workbook="column_forces.xlsx",
                source_sheet="Column Forces",
                source_row=12,
                source_columns=["N", "Combo"],
                evidence_type="clean_excel_data_table",
                unit_conversion_status="converted_with_audit",
                combo_family_status="explicit",
                combo_name="S_E_1",
                combo_family="S_E",
            ),
        ),
        _check(
            "beam_shear",
            _valid_workbook_evidence(
                source_workbook="08_KIRIS_KESME_KONTROLLERI.xlsx",
                source_sheet="KIRIS_KESME_KONTROL_OZET",
                source_row=7,
                source_columns=["Vd", "Vr", "Oran"],
                evidence_type="excel_computed_with_raw_data_link",
                unit_conversion_status="unknown",
                combo_family_status="combo_name_present_family_unclassified",
                combo_name="ENVE_EUD-EUHD",
                combo_family=None,
            ),
        ),
        _check(
            "story_drift",
            _valid_workbook_evidence(
                source_workbook="story_drift.xlsx",
                source_sheet="Drift Summary",
                source_row=3,
                source_columns=["Story", "Drift", "Limit"],
                evidence_type="clean_excel_data_table",
                unit_conversion_status="not_required_ratio",
                combo_family_status="not_applicable",
                combo_name=None,
                combo_family=None,
            ),
        ),
    ]

    for check in checks:
        valid, errors = _validate_evidence_policy(check.evidence)
        assert valid is True, errors

    json_path = tmp_path / "engine_report.json"
    excel_path = tmp_path / "engine_report.xlsx"
    JSONReporter(write_history=False).generate(
        checks,
        {"errors": {}, "skipped": {}, "execution_order": [], "cache_stats": {}},
        output_path=str(json_path),
    )
    ExcelReporter(write_history=False).generate(
        checks,
        {"errors": {}, "skipped": {}},
        output_path=str(excel_path),
    )

    payload = json.loads(json_path.read_text(encoding="utf-8"))
    json_rows = {row["check_id"]: row for row in payload["checks"]}
    workbook = load_workbook(excel_path)
    sheet = workbook["Details"]
    rows = list(sheet.iter_rows(values_only=True))
    header = list(rows[0])
    excel_rows = {row[header.index("check_id")]: dict(zip(header, row)) for row in rows[1:]}

    for check_id in ["column_axial", "beam_shear", "story_drift"]:
        json_evidence = json_rows[check_id]["evidence"]
        excel_evidence = json.loads(excel_rows[check_id]["evidence"])
        assert json_evidence["source_workbook"] == excel_evidence["source_workbook"]
        assert json_evidence["source_sheet"] == excel_evidence["source_sheet"]
        assert json_evidence["source_row"] == excel_evidence["source_row"]
        assert json_evidence["unit_conversion_status"] == excel_evidence["unit_conversion_status"]
        assert json_evidence["combo_family_status"] == excel_evidence["combo_family_status"]
