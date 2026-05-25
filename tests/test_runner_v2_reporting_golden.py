from __future__ import annotations

import ast
import json
from pathlib import Path
from typing import Mapping

from openpyxl import load_workbook

import tbdy_engine.runner_v2 as runner_v2
from tbdy_engine.runner_v2 import TBDYEngineV2


ROOT = Path(__file__).resolve().parents[1]
REPORT_GOLDEN = ROOT / "tests" / "golden" / "runner_v2_report_shape.json"
EXCEL_GOLDEN = ROOT / "tests" / "golden" / "runner_v2_excel_shape.json"
FORBIDDEN_SECOND_CONTRACT_FILES = [
    ROOT / "docs" / "workbook_manifest.yaml",
    ROOT / "docs" / "sheet_contracts.yaml",
    ROOT / "docs" / "unit_contract.yaml",
    ROOT / "docs" / "evidence_contract.yaml",
    ROOT / "tbdy_engine" / "contracts" / "workbook_manifest.yaml",
    ROOT / "tbdy_engine" / "contracts" / "sheet_contracts.yaml",
    ROOT / "tbdy_engine" / "contracts" / "unit_contract.yaml",
    ROOT / "tbdy_engine" / "contracts" / "evidence_contract.yaml",
]
FORBIDDEN_IMPORT_PREFIXES = (
    "tbdy_engine.design",
    "tbdy_engine.etabs",
    "tbdy_engine.engine.context_builder",
)


def _engine(tmp_path: Path) -> TBDYEngineV2:
    return TBDYEngineV2(ctx={"fixture": "context"}, report_dir=tmp_path)


def _column_payload() -> Mapping[str, object]:
    return {
        "outputs": [
            {
                "label": "C1",
                "story": "S1",
                "checks": {
                    "geometry": {
                        "status": "OK",
                        "ratio": 0.10,
                        "value": 0.10,
                        "limit": 1.0,
                        "unit": "ratio",
                        "message": "golden/report fixture",
                        "source": "runner_v2_reporting_golden_test",
                        "evaluation_level": "DESIGN_LEVEL",
                        "evidence": {
                            "source": "runner_v2_reporting_golden_test",
                            "evidence_type": "diagnostic_helper",
                            "confidence": "HIGH",
                            "unit_conversion_status": "not_required",
                            "combo_family_status": "not_applicable",
                        },
                    }
                },
            }
        ]
    }


def _ok_evaluator(context: object) -> Mapping[str, object]:
    return _column_payload()


def _failing_evaluator(context: object) -> Mapping[str, object]:
    raise RuntimeError("boom")


def _patch_evaluators(monkeypatch) -> None:
    monkeypatch.setattr(
        runner_v2.TBDYEngineV2,
        "_build_evaluators",
        lambda self, catalog: {
            "COLUMN_DESIGN": _ok_evaluator,
            "BEAM_DESIGN": _failing_evaluator,
        },
    )


def _run(tmp_path: Path, monkeypatch):
    _patch_evaluators(monkeypatch)
    engine = _engine(tmp_path)
    runner_payload = engine.run()
    json_payload = json.loads((tmp_path / "engine_report.json").read_text(encoding="utf-8"))
    workbook = load_workbook(tmp_path / "engine_report.xlsx")
    return runner_payload, json_payload, workbook


def _golden(path: Path) -> dict[str, object]:
    return json.loads(path.read_text(encoding="utf-8"))


def _column_check(json_payload: dict[str, object]) -> dict[str, object]:
    checks = json_payload["checks"]
    for check in checks:
        if check["check_id"] == "column_geometry":
            return check
    raise AssertionError("column_geometry check not found")


def _details_header(workbook) -> list[str]:
    return [cell.value for cell in workbook["Details"][1]]


def _details_rows(workbook) -> list[list[object]]:
    return [list(row) for row in workbook["Details"].iter_rows(min_row=2, values_only=True)]


def test_runner_v2_returned_payload_shape_is_stable(tmp_path, monkeypatch):
    runner_payload, _json_payload, _workbook = _run(tmp_path, monkeypatch)

    assert list(runner_payload) == _golden(REPORT_GOLDEN)["runner_return_keys"]


def test_json_top_level_keys_are_stable(tmp_path, monkeypatch):
    _runner_payload, json_payload, _workbook = _run(tmp_path, monkeypatch)

    assert list(json_payload) == _golden(REPORT_GOLDEN)["json_top_level_keys"]


def test_json_check_row_fields_are_stable(tmp_path, monkeypatch):
    _runner_payload, json_payload, _workbook = _run(tmp_path, monkeypatch)
    check = _column_check(json_payload)

    assert list(check) == _golden(REPORT_GOLDEN)["json_check_row_keys"]


def test_evidence_is_preserved_in_json(tmp_path, monkeypatch):
    _runner_payload, json_payload, _workbook = _run(tmp_path, monkeypatch)
    evidence = _column_check(json_payload)["evidence"]

    assert list(evidence) == _golden(REPORT_GOLDEN)["evidence_keys"]
    assert evidence["evidence_type"] == "diagnostic_helper"
    assert evidence["confidence"] == "HIGH"
    assert evidence["unit_conversion_status"] == "not_required"
    assert evidence["combo_family_status"] == "not_applicable"


def test_excel_sheet_names_are_stable(tmp_path, monkeypatch):
    _runner_payload, _json_payload, workbook = _run(tmp_path, monkeypatch)

    assert workbook.sheetnames == _golden(EXCEL_GOLDEN)["excel_sheetnames"]


def test_excel_details_header_is_stable(tmp_path, monkeypatch):
    _runner_payload, _json_payload, workbook = _run(tmp_path, monkeypatch)

    assert _details_header(workbook) == _golden(EXCEL_GOLDEN)["excel_details_header"]


def test_excel_details_evidence_parses(tmp_path, monkeypatch):
    _runner_payload, _json_payload, workbook = _run(tmp_path, monkeypatch)
    header = _details_header(workbook)
    rows = _details_rows(workbook)
    check_id_index = header.index("check_id")
    evidence_index = header.index("evidence")
    column_rows = [row for row in rows if row[check_id_index] == "column_geometry"]

    assert column_rows
    evidence = json.loads(column_rows[0][evidence_index])
    assert evidence["evidence_type"] == "diagnostic_helper"


def test_eval_skipped_and_eval_errors_are_stable(tmp_path, monkeypatch):
    _runner_payload, json_payload, workbook = _run(tmp_path, monkeypatch)

    assert json_payload["evaluation_skipped"]
    assert json_payload["evaluation_errors"].get("BEAM_DESIGN") == "boom"
    assert workbook["Eval_Skipped"].max_row > 1
    assert workbook["Eval_Errors"].max_row > 1


def test_report_contract_metadata_is_stable(tmp_path, monkeypatch):
    _runner_payload, _json_payload, workbook = _run(tmp_path, monkeypatch)
    sheet = workbook["Report_Contract"]
    keys = [row[0].value for row in sheet.iter_rows(min_row=2, max_col=1)]

    assert keys == _golden(EXCEL_GOLDEN)["report_contract_keys"]


def test_normalized_golden_files_match_current_shape(tmp_path, monkeypatch):
    runner_payload, json_payload, workbook = _run(tmp_path, monkeypatch)
    check = _column_check(json_payload)
    evidence = check["evidence"]

    current_report_shape = {
        "runner_return_keys": list(runner_payload),
        "json_top_level_keys": list(json_payload),
        "json_check_row_keys": list(check),
        "evidence_keys": list(evidence),
    }
    current_excel_shape = {
        "excel_sheetnames": workbook.sheetnames,
        "excel_details_header": _details_header(workbook),
        "report_contract_keys": [
            row[0].value for row in workbook["Report_Contract"].iter_rows(min_row=2, max_col=1)
        ],
    }

    assert current_report_shape == _golden(REPORT_GOLDEN)
    assert current_excel_shape == _golden(EXCEL_GOLDEN)


def test_dry_run_still_writes_no_files(tmp_path):
    _engine(tmp_path).dry_run()

    assert not (tmp_path / "engine_report.json").exists()
    assert not (tmp_path / "engine_report.xlsx").exists()


def test_no_forbidden_imports_or_schema_drift_sources():
    for relative_path in ("tbdy_engine/runner_v2.py", "tbdy_engine/runtime/scheduler.py"):
        tree = ast.parse((ROOT / relative_path).read_text(encoding="utf-8"))
        imported_modules: list[str] = []
        for node in ast.walk(tree):
            if isinstance(node, ast.Import):
                imported_modules.extend(alias.name for alias in node.names)
            elif isinstance(node, ast.ImportFrom) and node.module:
                imported_modules.append(node.module)
        forbidden_imports = sorted(
            module_name
            for module_name in imported_modules
            if any(
                module_name == prefix or module_name.startswith(prefix + ".")
                for prefix in FORBIDDEN_IMPORT_PREFIXES
            )
        )
        assert forbidden_imports == []

    combined = (ROOT / "tbdy_engine" / "runner_v2.py").read_text(encoding="utf-8")
    combined += "\n" + (ROOT / "tbdy_engine" / "runtime" / "scheduler.py").read_text(encoding="utf-8")
    assert "combo_family" not in combined
    assert "uses_combo" not in combined
    assert "message_text" not in combined
    assert ".message" not in combined

    for path in FORBIDDEN_SECOND_CONTRACT_FILES:
        assert not path.exists(), str(path.relative_to(ROOT))


def test_execution_order_is_deterministic(tmp_path, monkeypatch):
    _runner_payload_a, json_payload_a, _workbook_a = _run(tmp_path / "a", monkeypatch)
    _runner_payload_b, json_payload_b, _workbook_b = _run(tmp_path / "b", monkeypatch)

    assert json_payload_a["execution_order"] == json_payload_b["execution_order"]
    assert "COLUMN_DESIGN" in json_payload_a["execution_order"]
    assert "BEAM_DESIGN" in json_payload_a["execution_order"]


def test_report_filenames_unchanged(tmp_path, monkeypatch):
    runner_payload, _json_payload, _workbook = _run(tmp_path, monkeypatch)

    assert Path(runner_payload["reports"]["json"]).name == "engine_report.json"
    assert Path(runner_payload["reports"]["excel"]).name == "engine_report.xlsx"
    assert (tmp_path / "engine_report.json").exists()
    assert (tmp_path / "engine_report.xlsx").exists()
