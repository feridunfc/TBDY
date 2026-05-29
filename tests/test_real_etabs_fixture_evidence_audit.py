from __future__ import annotations

import pytest

pytestmark = [
    pytest.mark.legacy_evidence_audit,
    pytest.mark.skip(
        reason="Legacy evidence audit uses pre-closure adapter/report contracts; archived from BEAM_RUNTIME_CLOSURE proof."
    ),
]

import csv
import json
import re
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from tbdy_engine.adapters.check_adapter import CheckAdapter
from tbdy_engine.contracts.loader import EngineContractLoader
from tbdy_engine.reports.excel_reporter import ExcelReporter
from tbdy_engine.reports.json_reporter import JSONReporter


ROOT = Path(__file__).resolve().parents[1]
FIXTURE_DIR = ROOT / "tests" / "fixtures" / "etabs_anonymized"
FIXTURE_STATUS = "ETABS_EXPORT_SHAPED_ANONYMIZED"
ALLOWED_FIXTURE_STATUSES = {
    "REAL_ETABS_EXPORT_ANONYMIZED",
    "ETABS_EXPORT_SHAPED_ANONYMIZED",
    "ETABS_SHAPED_SYNTHETIC_NO_REAL_FIXTURE",
}
FORBIDDEN_IDENTIFIERS = [
    "client",
    "project",
    "address",
    "company",
    "email",
    "phone",
    "vergi",
]
TC_ID_PATTERN = re.compile(r"\b\d{11}\b")
FORBIDDEN_SECOND_CONTRACT_FILES = [
    ROOT / "docs" / "workbook_manifest.yaml",
    ROOT / "docs" / "sheet_contracts.yaml",
    ROOT / "docs" / "unit_contract.yaml",
    ROOT / "docs" / "evidence_contract.yaml",
    ROOT / "tbdy_engine" / "contracts" / "workbook_manifest.yaml",
    ROOT / "tbdy_engine" / "contracts" / "sheet_contracts.yaml",
    ROOT / "tbdy_engine" / "contracts" / "unit_contract.yaml",
    ROOT / "tbdy_engine" / "contracts" / "evidence_contract.yaml",
]


def _catalog():
    return EngineContractLoader.from_project_root(ROOT).build_runtime_catalog()


def _fixture_path(name: str) -> Path:
    return FIXTURE_DIR / name


def _csv_rows(path: Path) -> list[dict[str, str]]:
    with path.open("r", encoding="utf-8", newline="") as handle:
        return list(csv.DictReader(handle))


def _json_fixture_rows(path: Path) -> list[dict[str, Any]]:
    payload = json.loads(path.read_text(encoding="utf-8"))

    if isinstance(payload, list):
        return payload

    if isinstance(payload, dict):
        if isinstance(payload.get("checks"), list):
            return payload["checks"]
        if isinstance(payload.get("rows"), list):
            return payload["rows"]

    raise AssertionError(f"Unsupported JSON fixture shape: {path}")


def _as_float(value: Any) -> float:
    return float(value)


def _base_evidence(
    *,
    source_file: str,
    table_name: str,
    source_row: int | None,
    source_columns: list[str],
    evidence_type: str = "clean_excel_data_table",
    confidence: str = "MEDIUM",
    unit_conversion_status: str = "unknown",
    combo_name: str | None = None,
    notes: list[str] | None = None,
) -> dict[str, Any]:
    evidence = {
        "fixture_status": FIXTURE_STATUS,
        "source_file": source_file,
        "source_workbook": source_file,
        "table_name": table_name,
        "source_sheet": table_name,
        "source_row": source_row,
        "source_columns": source_columns,
        "evidence_type": evidence_type,
        "confidence": confidence,
        "unit_conversion_status": unit_conversion_status,
        "combo_name": combo_name,
        "combo_family": None,
        "combo_family_status": "combo_name_present_family_unclassified" if combo_name else "not_classified",
        "notes": notes or ["anonymized ETABS-export-shaped fixture; not claimed as real ETABS export"],
    }
    return evidence


def _check_payload(
    *,
    status: str = "OK",
    ratio: float = 0.5,
    value: float = 0.5,
    limit: float = 1.0,
    unit: str = "ratio",
    evidence: dict[str, Any] | None = None,
    governing_combo: str | None = None,
) -> dict[str, Any]:
    payload = {
        "status": status,
        "ratio": ratio,
        "value": value,
        "limit": limit,
        "unit": unit,
        "message": "anonymized ETABS fixture evidence audit",
        "tbdy_ref": "fixture",
        "evaluation_level": "DESIGN_LEVEL",
        "source": "anonymized_fixture",
        "combo_family": None,
    }
    if evidence is not None:
        payload["evidence"] = evidence
    if governing_combo is not None:
        payload["governing_combo"] = governing_combo
    return payload


def _column_outputs() -> dict[str, Any]:
    force = _csv_rows(_fixture_path("column_forces.csv"))[0]
    rebar = _json_fixture_rows(_fixture_path("column_rebar_minimum.json"))[0]
    label = force["element_label"]
    story = force["story"]

    axial_evidence = _base_evidence(
        source_file="column_forces.csv",
        table_name="column_forces",
        source_row=int(force["source_row"]),
        source_columns=["N_kn", "N_case"],
        combo_name=force["N_case"],
    )
    axial_evidence.update(
        {
            "component_case": force["N_case"],
            "governing_combo": force["N_case"],
            "N_kn": _as_float(force["N_kn"]),
        }
    )

    shear_evidence = _base_evidence(
        source_file="column_forces.csv",
        table_name="column_forces",
        source_row=int(force["source_row"]),
        source_columns=["Vx_kn", "Vy_kn", "Vx_case", "Vy_case"],
        combo_name=None,
        notes=["column shear has component cases but no single top-level governing combo in this audit fixture"],
    )
    shear_evidence.update(
        {
            "force": "max(abs(Vx_kn), abs(Vy_kn))",
            "Vx_kn": _as_float(force["Vx_kn"]),
            "Vy_kn": _as_float(force["Vy_kn"]),
            "Vx_case": force["Vx_case"],
            "Vy_case": force["Vy_case"],
        }
    )

    rebar_evidence = _base_evidence(
        source_file="column_rebar_minimum.json",
        table_name="column_rebar_minimum",
        source_row=int(rebar["source_row"]),
        source_columns=[
            "As_total_mm2",
            "As_min_mm2",
            "rho_pct",
            "rho_min_pct",
            "n_bars_total",
            "bar_diameter_mm",
        ],
        evidence_type="lookup_table",
        combo_name=None,
        notes=["rebar minimum fixture is anonymized export-shaped data"],
    )
    rebar_evidence.update(rebar)

    return {
        "label": label,
        "story": story,
        "checks": {
            "axial": _check_payload(
                ratio=0.72,
                value=_as_float(force["N_kn"]),
                limit=1000.0,
                unit="kN",
                evidence=axial_evidence,
                governing_combo=force["N_case"],
            ),
            "shear": _check_payload(
                ratio=0.62,
                value=max(abs(_as_float(force["Vx_kn"])), abs(_as_float(force["Vy_kn"]))),
                limit=500.0,
                unit="kN",
                evidence=shear_evidence,
            ),
            "rebar_minimum": _check_payload(
                ratio=0.75,
                value=_as_float(rebar["rho_pct"]),
                limit=_as_float(rebar["rho_min_pct"]),
                unit="%",
                evidence=rebar_evidence,
            ),
        },
    }


def _beam_semantic_evidence() -> tuple[dict[str, Any], dict[str, Any]]:
    flexure_rows = _csv_rows(_fixture_path("concrete_beam_flexure_envelope.csv"))
    shear_rows = _csv_rows(_fixture_path("concrete_beam_shear_envelope.csv"))

    end_i = next(row for row in flexure_rows if row["Location"] == "End-I")
    middle = next(row for row in flexure_rows if row["Location"] == "Middle")
    end_j = next(row for row in flexure_rows if row["Location"] == "End-J")
    max_shear = max(shear_rows, key=lambda row: abs(_as_float(row["Shear"])))
    support_shear = max((row for row in shear_rows if row["Location"] in {"End-I", "End-J"}), key=lambda row: abs(_as_float(row["Shear"])))

    flexure_evidence = _base_evidence(
        source_file="concrete_beam_flexure_envelope.csv",
        table_name="Concrete Beam Flexure Envelope -  TS 500-2000(R2018)",
        source_row=2,
        source_columns=["Location", "MomentTop", "MomentBot", "AsTopCombo", "AsBotCombo"],
        evidence_type="excel_computed_with_raw_data_link",
        combo_name=None,
        notes=["beam flexure semantic mapping uses End-I/Middle/End-J rows from anonymized export-shaped fixture"],
    )
    flexure_evidence["forces"] = {
        "M_pos_knm": abs(_as_float(middle["MomentBot"])),
        "M_neg_left_knm": abs(_as_float(end_i["MomentTop"])),
        "M_neg_right_knm": abs(_as_float(end_j["MomentTop"])),
        "M_pos_case": middle["AsBotCombo"],
        "M_neg_left_case": end_i["AsTopCombo"],
        "M_neg_right_case": end_j["AsTopCombo"],
    }

    shear_evidence = _base_evidence(
        source_file="concrete_beam_shear_envelope.csv",
        table_name="Concrete Beam Shear Envelope -  TS 500-2000(R2018)",
        source_row=2,
        source_columns=["Location", "Shear", "VCombo"],
        evidence_type="excel_computed_with_raw_data_link",
        combo_name=max_shear["VCombo"],
    )
    shear_evidence["forces"] = {
        "V_max_kn": abs(_as_float(max_shear["Shear"])),
        "V_at_support_kn": abs(_as_float(support_shear["Shear"])),
        "V_max_case": max_shear["VCombo"],
        "V_at_support_case": support_shear["VCombo"],
    }
    return flexure_evidence, shear_evidence


def _beam_outputs() -> dict[str, Any]:
    flexure_evidence, shear_evidence = _beam_semantic_evidence()
    return {
        "label": "B_A",
        "story": "S_A",
        "checks": {
            "flexure": _check_payload(
                ratio=0.76,
                value=0.76,
                limit=1.0,
                evidence=flexure_evidence,
                governing_combo="BEAM_FLEX_COMBO",
            ),
            "shear": _check_payload(
                ratio=0.69,
                value=shear_evidence["forces"]["V_max_kn"],
                limit=200.0,
                unit="kN",
                evidence=shear_evidence,
                governing_combo=shear_evidence["forces"]["V_max_case"],
            ),
        },
    }


def _eval_results() -> dict[str, Any]:
    return {
        "results": {
            "COLUMN_DESIGN": {"outputs": [_column_outputs()]},
            "BEAM_DESIGN": {"outputs": [_beam_outputs()]},
        },
        "errors": {},
        "skipped": {},
        "execution_order": ["COLUMN_DESIGN", "BEAM_DESIGN"],
        "cache_stats": {},
    }


def _adapted_rows():
    rows = CheckAdapter(_catalog()).adapt_all(_eval_results())
    return {row.check_id: row for row in rows}


def _json_rows(path: Path):
    payload = json.loads(path.read_text(encoding="utf-8"))
    return {row["check_id"]: row for row in payload["checks"]}


def _excel_rows(path: Path):
    workbook = load_workbook(path)
    sheet = workbook["Details"]
    rows = list(sheet.iter_rows(values_only=True))
    header = list(rows[0])
    return header, {row[header.index("check_id")]: dict(zip(header, row)) for row in rows[1:]}


def test_fixture_status_is_explicit():
    assert FIXTURE_STATUS in ALLOWED_FIXTURE_STATUSES
    assert FIXTURE_STATUS == "ETABS_EXPORT_SHAPED_ANONYMIZED"
    assert FIXTURE_STATUS != "REAL_ETABS_EXPORT_ANONYMIZED"


def test_committed_fixture_files_are_small_and_anonymized():
    assert FIXTURE_DIR.exists()
    files = [path for path in FIXTURE_DIR.iterdir() if path.is_file()]
    assert files

    for path in files:
        assert path.stat().st_size < 1_000_000
        if path.suffix.lower() in {".csv", ".json", ".txt"}:
            text = path.read_text(encoding="utf-8").lower()
            for identifier in FORBIDDEN_IDENTIFIERS:
                assert identifier not in text
            assert TC_ID_PATTERN.search(text) is None


def test_column_fixture_evidence_reaches_adapter_json_excel(tmp_path):
    checks = list(_adapted_rows().values())
    json_path = tmp_path / "engine_report.json"
    excel_path = tmp_path / "engine_report.xlsx"
    JSONReporter(write_history=False).generate(checks, _eval_results(), output_path=str(json_path))
    ExcelReporter(write_history=False).generate(checks, _eval_results(), output_path=str(excel_path))

    adapter_rows = _adapted_rows()
    json_rows = _json_rows(json_path)
    _header, excel_rows = _excel_rows(excel_path)

    assert adapter_rows["column_axial"].evidence["component_case"] == "S_E_COL_N"
    assert json_rows["column_axial"]["evidence"]["governing_combo"] == "S_E_COL_N"
    assert json.loads(excel_rows["column_axial"]["evidence"])["component_case"] == "S_E_COL_N"

    assert adapter_rows["column_shear"].evidence["Vx_case"] == "K_E_COL_VX"
    assert json_rows["column_shear"]["evidence"]["Vy_case"] == "K_E_COL_VY"
    assert json.loads(excel_rows["column_shear"]["evidence"])["Vx_case"] == "K_E_COL_VX"

    assert adapter_rows["column_rebar_minimum"].evidence["As_min_mm2"] == 900.0
    assert json_rows["column_rebar_minimum"]["evidence"]["rho_pct"] == 1.3333333333
    assert json.loads(excel_rows["column_rebar_minimum"]["evidence"])["source"] == "anonymized_export_shape"


def test_beam_fixture_evidence_reaches_adapter_json_excel(tmp_path):
    checks = list(_adapted_rows().values())
    json_path = tmp_path / "engine_report.json"
    excel_path = tmp_path / "engine_report.xlsx"
    JSONReporter(write_history=False).generate(checks, _eval_results(), output_path=str(json_path))
    ExcelReporter(write_history=False).generate(checks, _eval_results(), output_path=str(excel_path))

    adapter_rows = _adapted_rows()
    json_rows = _json_rows(json_path)
    _header, excel_rows = _excel_rows(excel_path)

    assert adapter_rows["beam_flexure"].evidence["forces"]["M_pos_case"] == "Crack_SeisX"
    assert json_rows["beam_flexure"]["evidence"]["forces"]["M_neg_left_case"] == "Crack_SeisX_Soil"
    assert json.loads(excel_rows["beam_flexure"]["evidence"])["forces"]["M_neg_right_case"] == "Crack_SeisX"

    assert adapter_rows["beam_shear"].evidence["forces"]["V_max_case"] == "Crack_SeisY_Soil"
    assert json_rows["beam_shear"]["evidence"]["forces"]["V_at_support_case"] == "Crack_SeisY_Soil"
    assert json.loads(excel_rows["beam_shear"]["evidence"])["forces"]["V_max_case"] == "Crack_SeisY_Soil"


def test_source_table_provenance_survives_for_column_and_beam_checks():
    rows = _adapted_rows()

    for check_id in ["column_axial", "beam_flexure"]:
        evidence = rows[check_id].evidence
        assert evidence["fixture_status"] == FIXTURE_STATUS
        assert evidence["source_file"]
        assert evidence["source_workbook"] == evidence["source_file"]
        assert evidence["table_name"]
        assert evidence["source_sheet"] == evidence["table_name"]
        assert evidence["source_row"] is not None
        assert evidence["source_columns"]
        assert evidence["evidence_type"]
        assert evidence["unit_conversion_status"]
        assert evidence["combo_family_status"]


def test_excel_evidence_cells_are_parseable_json(tmp_path):
    checks = list(_adapted_rows().values())
    excel_path = tmp_path / "engine_report.xlsx"
    ExcelReporter(write_history=False).generate(checks, _eval_results(), output_path=str(excel_path))

    header, rows = _excel_rows(excel_path)
    assert "evidence" in header

    for check_id in ["column_axial", "column_shear", "column_rebar_minimum", "beam_flexure", "beam_shear"]:
        parsed = json.loads(rows[check_id]["evidence"])
        assert parsed["fixture_status"] == FIXTURE_STATUS
        assert parsed["source_file"]
        assert parsed["source_columns"]


def test_no_unsafe_inference_or_message_derived_evidence():
    rows = _adapted_rows()

    for row in rows.values():
        if row.evidence:
            assert row.evidence.get("combo_family") is None
            assert row.evidence.get("combo_family_status") in {
                "combo_name_present_family_unclassified",
                "not_classified",
            }
            assert "uses_combo" not in row.evidence
            assert "message" not in row.evidence
            assert "message_text" not in row.evidence
        assert row.combo_family is None


def test_real_etabs_evidence_audit_status_remains_honest():
    try:
        from tests import test_real_etabs_evidence_audit as audit
    except ImportError:
        return

    assert audit.FIXTURE_STATUS_SYNTHETIC == "ETABS_SHAPED_SYNTHETIC_NO_REAL_FIXTURE"
    assert audit.FIXTURE_STATUS_REAL == "REAL_ETABS_FIXTURE_AVAILABLE"
    assert audit.FIXTURE_STATUS_SYNTHETIC != FIXTURE_STATUS


def test_single_contract_system_remains_intact():
    for path in FORBIDDEN_SECOND_CONTRACT_FILES:
        assert not path.exists(), str(path.relative_to(ROOT))
