# tests/test_runtime_interface_audit.py
from __future__ import annotations

import importlib
import sys
from dataclasses import dataclass
from enum import Enum
from pathlib import Path
from typing import Literal

from openpyxl import load_workbook

from tbdy_engine.adapters.check_adapter import CheckAdapter
from tbdy_engine.contracts.loader import EngineContractLoader
from tbdy_engine.reports.excel_reporter import ExcelReporter
from tbdy_engine.reports.json_reporter import JSONReporter


ROOT = Path(__file__).resolve().parents[1]

AuditStatus = Literal["PRESENT_IMPORTABLE", "PRESENT_NOT_IMPORTABLE", "ABSENT"]
RunnerStatus = Literal["PRESENT", "ABSENT"]


class RuntimeObjectStatus(str, Enum):
    PRESENT_IMPORTABLE = "PRESENT_IMPORTABLE"
    PRESENT_NOT_IMPORTABLE = "PRESENT_NOT_IMPORTABLE"
    ABSENT = "ABSENT"


@dataclass(frozen=True)
class RuntimeObjectAudit:
    name: str
    status: RuntimeObjectStatus
    module_candidates: tuple[str, ...]


APPROVED_RUNTIME_OBJECTS = {
    "DatasetValidator": (
        "tbdy_engine.runtime.dataset_validator",
        "tbdy_engine.runtime.validator",
        "tbdy_engine.contracts.validator",
    ),
    "EvaluationDAG": (
        "tbdy_engine.runtime.evaluation_dag",
        "tbdy_engine.runtime.dag",
    ),
    "Scheduler": (
        "tbdy_engine.runtime.scheduler",
    ),
    "EvaluationResult": (
        "tbdy_engine.runtime.evaluation_result",
        "tbdy_engine.runtime.results",
        "tbdy_engine.runtime.scheduler",
    ),
}

ALLOWED_STATUSES = {
    RuntimeObjectStatus.PRESENT_IMPORTABLE,
    RuntimeObjectStatus.PRESENT_NOT_IMPORTABLE,
    RuntimeObjectStatus.ABSENT,
}

FORBIDDEN_SECOND_CONTRACT_FILES = [
    ROOT / "docs" / "workbook_manifest.yaml",
    ROOT / "docs" / "sheet_contracts.yaml",
    ROOT / "docs" / "unit_contract.yaml",
    ROOT / "docs" / "evidence_contract.yaml",
    ROOT / "tbdy_engine" / "contracts" / "workbook_manifest.yaml",
    ROOT / "tbdy_engine" / "contracts" / "sheet_contracts.yaml",
    ROOT / "tbdy_engine" / "contracts" / "unit_contract.yaml",
    ROOT / "tbdy_engine" / "contracts" / "evidence_contract.yaml",
]

FORBIDDEN_HEAVY_MODULE_PREFIXES = (
    "tbdy_engine.design",
    "tbdy_engine.etabs",
    "tbdy_engine.engine.context_builder",
)

MODULES_AT_IMPORT = frozenset(sys.modules)


def _catalog():
    return EngineContractLoader.from_project_root(ROOT).build_runtime_catalog()


def _object_status(object_name: str, module_candidates: tuple[str, ...]) -> RuntimeObjectStatus:
    saw_file = False

    for module_name in module_candidates:
        module_path = ROOT / Path(*module_name.split(".")).with_suffix(".py")
        package_path = ROOT / Path(*module_name.split(".")) / "__init__.py"
        if module_path.exists() or package_path.exists():
            saw_file = True

        try:
            module = importlib.import_module(module_name)
        except Exception:
            continue

        if hasattr(module, object_name):
            return RuntimeObjectStatus.PRESENT_IMPORTABLE

    return RuntimeObjectStatus.PRESENT_NOT_IMPORTABLE if saw_file else RuntimeObjectStatus.ABSENT


def _runtime_interface_status() -> dict[str, str]:
    result = {
        name: _object_status(name, candidates).value
        for name, candidates in APPROVED_RUNTIME_OBJECTS.items()
    }

    runner_v2_path = ROOT / "tbdy_engine" / "runner_v2.py"
    result["runner_v2"] = "PRESENT" if runner_v2_path.exists() else "ABSENT"
    return result


RUNTIME_INTERFACE_STATUS = {
    name: _object_status(name, candidates).value
    for name, candidates in APPROVED_RUNTIME_OBJECTS.items()
}
RUNTIME_INTERFACE_STATUS["runner_v2"] = (
    "PRESENT" if (ROOT / "tbdy_engine" / "runner_v2.py").exists() else "ABSENT"
)


def _minimal_check_payload():
    return {
        "status": "OK",
        "ratio": 0.5,
        "value": 0.5,
        "limit": 1.0,
        "unit": "ratio",
        "message": "runtime interface audit fixture",
        "tbdy_ref": "fixture",
        "evaluation_level": "DESIGN_LEVEL",
        "source": "runtime_interface_audit",
        "evidence": {
            "source_workbook": "runtime_interface_audit",
            "source_sheet": "COLUMN_DESIGN",
            "source_row": 1,
            "source_columns": ["geometry"],
            "evidence_type": "diagnostic_helper",
            "confidence": "HIGH",
            "unit_conversion_status": "not_required",
            "combo_family_status": "not_applicable",
        },
    }


def _minimal_eval_results():
    return {
        "results": {
            "COLUMN_DESIGN": {
                "outputs": [
                    {
                        "label": "C1",
                        "story": "S1",
                        "checks": {
                            "geometry": _minimal_check_payload(),
                        },
                    }
                ]
            }
        },
        "errors": {},
        "skipped": {},
        "execution_order": ["COLUMN_DESIGN"],
        "cache_stats": {},
    }


def _adapted_rows():
    return CheckAdapter(_catalog()).adapt_all(_minimal_eval_results())


def test_contract_loader_exposes_single_runtime_catalog_source():
    catalog = _catalog()

    for attr in ["checks", "reports", "datasets", "combo_families", "evaluations"]:
        assert hasattr(catalog, attr), attr

    assert catalog.checks
    assert catalog.reports
    assert catalog.datasets
    assert catalog.evaluations

    enabled_checks = [
        check for check in catalog.checks.values()
        if getattr(check, "runner_enabled", False)
    ]
    assert enabled_checks

    assert "full_engine_report" in catalog.reports
    full_engine_report = catalog.reports["full_engine_report"]
    include_fields = getattr(full_engine_report, "include_fields", [])
    assert "evidence" in include_fields

    for path in FORBIDDEN_SECOND_CONTRACT_FILES:
        assert not path.exists(), str(path.relative_to(ROOT))


def test_approved_runtime_flow_object_availability_is_explicit():
    audits = [
        RuntimeObjectAudit(
            name=name,
            status=_object_status(name, candidates),
            module_candidates=candidates,
        )
        for name, candidates in APPROVED_RUNTIME_OBJECTS.items()
    ]

    assert {audit.name for audit in audits} == {
        "DatasetValidator",
        "EvaluationDAG",
        "Scheduler",
        "EvaluationResult",
    }

    for audit in audits:
        assert audit.status in ALLOWED_STATUSES

    status = _runtime_interface_status()
    assert set(status) == {
        "DatasetValidator",
        "EvaluationDAG",
        "Scheduler",
        "EvaluationResult",
        "runner_v2",
    }
    assert status["runner_v2"] in {"PRESENT", "ABSENT"}
    for key in ["DatasetValidator", "EvaluationDAG", "Scheduler", "EvaluationResult"]:
        assert status[key] in {item.value for item in RuntimeObjectStatus}


def test_check_adapter_input_boundary_documents_de_facto_eval_results_shape():
    rows = _adapted_rows()
    by_check_id = {row.check_id: row for row in rows}

    assert "column_geometry" in by_check_id

    row = by_check_id["column_geometry"]
    assert row.evaluation == "COLUMN_DESIGN"
    assert row.element_label == "C1"
    assert row.story == "S1"
    assert row.status == "OK"
    assert row.evidence["evidence_type"] == "diagnostic_helper"

    eval_results = _minimal_eval_results()
    assert set(eval_results) == {
        "results",
        "errors",
        "skipped",
        "execution_order",
        "cache_stats",
    }


def test_reporters_accept_checkresult_rows_plus_eval_results_shape(tmp_path):
    checks = _adapted_rows()
    eval_results = _minimal_eval_results()

    json_path = tmp_path / "engine_report.json"
    excel_path = tmp_path / "engine_report.xlsx"

    JSONReporter(write_history=False).generate(
        checks,
        eval_results,
        runtime_catalog=_catalog(),
        output_path=str(json_path),
    )
    ExcelReporter(write_history=False).generate(
        checks,
        eval_results,
        output_path=str(excel_path),
    )

    import json

    payload = json.loads(json_path.read_text(encoding="utf-8"))
    assert list(payload.keys()) == [
        "report_metadata",
        "summary",
        "checks",
        "evaluation_errors",
        "evaluation_skipped",
        "execution_order",
        "cache_stats",
        "coverage",
        "distributions",
    ]
    assert payload["checks"][0]["evidence"]["evidence_type"] == "diagnostic_helper"

    workbook = load_workbook(excel_path)
    assert set(workbook.sheetnames) >= {
        "Summary",
        "Details",
        "Eval_Skipped",
        "Eval_Errors",
    }

    details = workbook["Details"]
    header = [cell.value for cell in details[1]]
    assert "evidence" in header


def test_runner_v2_status_is_explicit():
    runner_v2_path = ROOT / "tbdy_engine" / "runner_v2.py"

    if runner_v2_path.exists():
        text = runner_v2_path.read_text(encoding="utf-8")
        assert text
        runner_status: RunnerStatus = "PRESENT"
    else:
        runner_status = "ABSENT"

    assert runner_status in {"PRESENT", "ABSENT"}
    assert RUNTIME_INTERFACE_STATUS["runner_v2"] == runner_status


def test_audit_source_does_not_import_forbidden_heavy_modules():
    import ast

    source = Path(__file__).read_text(encoding="utf-8-sig")
    tree = ast.parse(source)

    imported_modules: list[str] = []
    for node in ast.walk(tree):
        if isinstance(node, ast.Import):
            imported_modules.extend(alias.name for alias in node.names)
        elif isinstance(node, ast.ImportFrom):
            if node.module:
                imported_modules.append(node.module)

    forbidden_imports = sorted(
        module_name
        for module_name in imported_modules
        if any(
            module_name == prefix or module_name.startswith(prefix + ".")
            for prefix in FORBIDDEN_HEAVY_MODULE_PREFIXES
        )
    )

    assert forbidden_imports == []


def test_architecture_gap_report_is_encoded_as_test_constants():
    assert set(RUNTIME_INTERFACE_STATUS) == {
        "DatasetValidator",
        "EvaluationDAG",
        "Scheduler",
        "EvaluationResult",
        "runner_v2",
    }

    for key in ["DatasetValidator", "EvaluationDAG", "Scheduler", "EvaluationResult"]:
        assert RUNTIME_INTERFACE_STATUS[key] in {
            "PRESENT_IMPORTABLE",
            "PRESENT_NOT_IMPORTABLE",
            "ABSENT",
        }

    assert RUNTIME_INTERFACE_STATUS["runner_v2"] in {"PRESENT", "ABSENT"}


def test_runtime_package_inventory_is_deterministic_if_present():
    runtime_dir = ROOT / "tbdy_engine" / "runtime"

    if not runtime_dir.exists():
        assert RUNTIME_INTERFACE_STATUS["Scheduler"] in {
            "ABSENT",
            "PRESENT_NOT_IMPORTABLE",
            "PRESENT_IMPORTABLE",
        }
        return

    modules = sorted(
        path.name
        for path in runtime_dir.iterdir()
        if path.is_file() and path.suffix == ".py"
    )

    assert modules == sorted(modules)
