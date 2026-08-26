from __future__ import annotations

from hashlib import sha256
from importlib.util import module_from_spec, spec_from_file_location
from io import BytesIO
import json
from pathlib import Path
import sys
from zipfile import ZipFile

import pytest
from openpyxl import load_workbook
from pypdf import PdfReader

import tbdy_engine.product_reports.building_report_package as package_module
from tbdy_engine.product_reports.building_report_html import (
    HtmlRenderOptions,
    render_building_report_html,
)
from tbdy_engine.product_reports.building_report_json import (
    export_building_report_model_json,
)
from tbdy_engine.product_reports.building_report_pdf import (
    PdfRenderIntegrityError,
)
from tbdy_engine.product_reports.building_report_projection import ReportView
from tbdy_engine.product_reports.building_report_xlsx import (
    SUPPORTED_ET_XMLFILE_VERSION,
    SUPPORTED_OPENPYXL_VERSION,
    render_building_report_xlsx,
    xlsx_toolchain_identity,
)
from tbdy_engine.product_reports.building_report_package import (
    DEFAULT_PACKAGE_FILENAME,
    DEFAULT_PACKAGE_MEMBERS,
    build_building_report_package,
    build_report_delivery_artifacts,
    verify_building_report_package,
)
from tbdy_engine.product_reports.report_artifact import ReportArtifact
from tbdy_engine.product_reports.report_manifest import (
    DEFAULT_PAYLOAD_FILENAMES,
    build_report_manifest,
    verify_manifest_payloads,
)
from tbdy_engine.product_reports.report_presentation_selection import (
    ReportPresentationSelection,
)


_HELPER_MODULE_NAME = "_ur_1c_html_test_helpers_for_ur_1e"
_HELPER_PATH = Path(__file__).with_name("test_building_report_html.py")
_HELPER_SPEC = spec_from_file_location(_HELPER_MODULE_NAME, _HELPER_PATH)
if _HELPER_SPEC is None or _HELPER_SPEC.loader is None:  # pragma: no cover
    raise RuntimeError("could not load UR-1C HTML test helpers")
_HELPERS = module_from_spec(_HELPER_SPEC)
sys.modules[_HELPER_MODULE_NAME] = _HELPERS
_HELPER_SPEC.loader.exec_module(_HELPERS)
_model = _HELPERS._model
_projection = _HELPERS._projection


def _workbook(artifact: ReportArtifact):
    return load_workbook(BytesIO(artifact.content), data_only=False, keep_links=True)


def _rows_by_header(ws):
    headers = [cell.value for cell in ws[1]]
    result = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        result.append(dict(zip(headers, row)))
    return result


def _artifact_map(artifacts):
    return {item.filename: item for item in artifacts}


def _dummy_payloads(*, xlsx_content: bytes = b"xlsx-v1") -> tuple[ReportArtifact, ...]:
    model_content = b'{"model":"canonical"}\n'
    model_sha = sha256(model_content).hexdigest()
    artifacts = []
    for filename in DEFAULT_PAYLOAD_FILENAMES:
        if filename == "building_report_model.json":
            content = model_content
            fmt = "JSON"
            media = "application/json"
            view = None
            source_model = model_sha
        elif filename.endswith(".html"):
            content = ("<!doctype html>" + filename).encode()
            fmt = "HTML"
            media = "text/html; charset=utf-8"
            view = "AUDIT" if filename.startswith("audit.") else "ENGINEERING"
            source_model = None
        elif filename.endswith(".pdf"):
            content = ("%PDF-1.7\n" + filename).encode()
            fmt = "PDF"
            media = "application/pdf"
            view = "AUDIT" if filename.startswith("audit.") else "ENGINEERING"
            source_model = None
        else:
            content = xlsx_content if filename == "engineering.xlsx" else b"xlsx-audit"
            fmt = "XLSX"
            media = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            view = "AUDIT" if filename.startswith("audit.") else "ENGINEERING"
            source_model = None
        artifacts.append(
            ReportArtifact(
                logical_role="TEST_PAYLOAD",
                format=fmt,
                media_type=media,
                filename=filename,
                content=content,
                view=view,
                source_report_id="REPORT:1",
                source_project_id="PROJECT:1",
                source_model_sha256=source_model,
            )
        )
    return tuple(artifacts)


def test_canonical_json_is_full_model_deterministic_and_viewless() -> None:
    model = _model((("A", "PASS"), ("B", "FAIL")))
    first = export_building_report_model_json(model)
    second = export_building_report_model_json(model)

    assert first.filename == "building_report_model.json"
    assert first.view is None
    assert first.content == model.to_json().encode("utf-8")
    assert first.content == second.content
    assert first.sha256 == second.sha256
    assert first.artifact_id == second.artifact_id
    assert first.source_model_sha256 == sha256(first.content).hexdigest()

    payload = json.loads(first.content)
    assert len(payload["contributions"]) == 2
    assert payload["coverage_reconciliation"] == model.reconciliation.as_dict()
    assert "view" not in payload
    assert "ENGINEERING" not in payload.get("artifact_type", "")


def test_canonical_json_is_unaffected_by_presentation_selection() -> None:
    model = _model((("A", "PASS"), ("B", "FAIL")))
    canonical_before = export_building_report_model_json(model)
    selection = ReportPresentationSelection(statuses=("FAIL",))
    projection = _projection(model, ReportView.ENGINEERING)
    selected_xlsx = render_building_report_xlsx(projection, selection=selection)

    canonical_after = export_building_report_model_json(model)
    assert canonical_before.content == canonical_after.content
    assert canonical_before.artifact_id == canonical_after.artifact_id
    workbook = _workbook(selected_xlsx)
    rows = _rows_by_header(workbook["Contributions"])
    assert [row["status"] for row in rows] == ["FAIL"]


def test_xlsx_toolchain_is_exact_and_pinned() -> None:
    identity = json.loads(xlsx_toolchain_identity())
    assert identity["openpyxl"] == SUPPORTED_OPENPYXL_VERSION == "3.1.5"
    assert identity["et_xmlfile"] == SUPPORTED_ET_XMLFILE_VERSION == "2.0.0"


@pytest.mark.parametrize(
    ("view", "expected_sheets"),
    [
        (
            ReportView.ENGINEERING,
            [
                "Identity",
                "Project Basis",
                "Coverage",
                "Analysis Basis",
                "Status Facets",
                "Components",
                "Contributions",
                "Fields",
                "Calculations",
                "Table Rows",
                "Actions",
                "Warnings",
            ],
        ),
        (
            ReportView.AUDIT,
            [
                "Identity",
                "Project Basis",
                "Coverage",
                "Coverage Trace",
                "Analysis Basis",
                "Report Bindings",
                "Source Manifest",
                "Status Facets",
                "Components",
                "Contributions",
                "Fields",
                "Calculations",
                "Table Rows",
                "Actions",
                "Warnings",
            ],
        ),
    ],
)
def test_xlsx_reopens_with_exact_stable_sheet_order_and_no_formulas(view, expected_sheets) -> None:
    model = _model(
        (("A", "PASS"),),
        formula_rule="A",
        formula="Ndm <= 0.40 Ac fck",
        value=123.4567890123,
        unit="kN",
    )
    projection = _projection(model, view)
    first = render_building_report_xlsx(projection)
    second = render_building_report_xlsx(projection)

    assert first.content == second.content
    assert first.sha256 == second.sha256
    assert first.artifact_id == second.artifact_id
    assert first.view == view.value
    assert first.filename == f"{view.value.lower()}.xlsx"

    workbook = _workbook(first)
    assert workbook.sheetnames == expected_sheets
    assert not getattr(workbook, "_external_links", ())
    assert getattr(workbook, "vba_archive", None) is None
    assert all(
        cell.data_type != "f"
        for ws in workbook.worksheets
        for row in ws.iter_rows()
        for cell in row
    )

    contributions = _rows_by_header(workbook["Contributions"])
    assert contributions[0]["status"] == "PASS"
    fields = _rows_by_header(workbook["Fields"])
    summary = next(row for row in fields if row["field_scope"] == "SUMMARY")
    assert summary["value"] == 123.4567890123
    assert summary["unit"] == "kN"
    calculations = _rows_by_header(workbook["Calculations"])
    assert calculations[0]["formula_text"] == "Ndm <= 0.40 Ac fck"
    assert calculations[0]["governing_ref"] == "GOV:ROW:1"

    if view is ReportView.AUDIT:
        source_manifest = _rows_by_header(workbook["Source Manifest"])
        assert source_manifest[0]["fingerprint"] == "sha256:toy"
        assert source_manifest[0]["locator"] == "TOY §1"
        bindings = _rows_by_header(workbook["Report Bindings"])
        assert bindings[0]["source_ref"]
        assert bindings[0]["contribution_ref"]


@pytest.mark.parametrize("active_text", ["=1+1", "+SUM(A1:A2)", "-1+2", "@danger"])
def test_xlsx_spreadsheet_active_text_is_literal_not_formula(active_text: str) -> None:
    model = _model(
        (("A", "PASS"),),
        formula_rule="A",
        formula=active_text,
        title=active_text,
    )
    artifact = render_building_report_xlsx(
        _projection(model, ReportView.ENGINEERING)
    )
    workbook = _workbook(artifact)

    identity_rows = _rows_by_header(workbook["Identity"])
    assert next(row["value"] for row in identity_rows if row["key"] == "title") == active_text
    calculation_rows = _rows_by_header(workbook["Calculations"])
    assert calculation_rows[0]["formula_text"] == active_text
    assert all(
        cell.data_type != "f"
        for ws in workbook.worksheets
        for row in ws.iter_rows()
        for cell in row
    )


def test_xlsx_has_no_external_relationships_or_macros() -> None:
    artifact = render_building_report_xlsx(
        _projection(_model(), ReportView.AUDIT)
    )
    with ZipFile(BytesIO(artifact.content), "r") as archive:
        names = archive.namelist()
        assert not any("externalLinks/" in name for name in names)
        assert not any("vbaProject" in name for name in names)
        for name in names:
            if name.endswith((".xml", ".rels")):
                data = archive.read(name)
                assert b'TargetMode="External"' not in data


def test_truthful_empty_state_sheet_is_not_silently_omitted() -> None:
    selection = ReportPresentationSelection(include_results=False)
    artifact = render_building_report_xlsx(
        _projection(_model(), ReportView.ENGINEERING),
        selection=selection,
    )
    workbook = _workbook(artifact)
    assert "Contributions" in workbook.sheetnames
    assert workbook["Contributions"]["A2"].value == "Excluded from presentation selection."


def test_manifest_accounts_exactly_seven_payloads_without_recursive_self_hash() -> None:
    payloads = _dummy_payloads()
    first = build_report_manifest(payloads)
    second = build_report_manifest(tuple(reversed(payloads)))

    assert first.filename == "manifest.json"
    assert first.view is None
    assert first.content == second.content
    assert first.sha256 == second.sha256
    assert first.artifact_id == second.artifact_id
    verify_manifest_payloads(first, payloads)

    document = json.loads(first.content)
    assert document["payload_count"] == 7
    assert document["manifest_self_hash_in_payload_inventory"] is False
    assert [item["filename"] for item in document["payloads"]] == list(
        DEFAULT_PAYLOAD_FILENAMES
    )
    assert "manifest.json" not in [item["filename"] for item in document["payloads"]]
    assert first.sha256 not in first.content.decode("utf-8")


def test_one_payload_mutation_changes_manifest_and_zip_identity() -> None:
    first_payloads = _dummy_payloads(xlsx_content=b"xlsx-v1")
    second_payloads = _dummy_payloads(xlsx_content=b"xlsx-v2")
    first_manifest = build_report_manifest(first_payloads)
    second_manifest = build_report_manifest(second_payloads)

    assert first_manifest.content != second_manifest.content
    assert first_manifest.sha256 != second_manifest.sha256
    assert first_manifest.artifact_id != second_manifest.artifact_id

    first_members = tuple(sorted(first_payloads + (first_manifest,), key=lambda x: x.filename))
    second_members = tuple(sorted(second_payloads + (second_manifest,), key=lambda x: x.filename))
    first_zip = package_module._package_zip_bytes(first_members)
    second_zip = package_module._package_zip_bytes(second_members)
    assert first_zip != second_zip
    assert sha256(first_zip).hexdigest() != sha256(second_zip).hexdigest()


def test_delivery_members_are_exact_and_html_bytes_reuse_ur_1c_renderer() -> None:
    model = _model()
    artifacts = build_report_delivery_artifacts(model)
    by_name = _artifact_map(artifacts)

    assert tuple(sorted(by_name)) == DEFAULT_PACKAGE_MEMBERS
    engineering_projection = _projection(model, ReportView.ENGINEERING)
    audit_projection = _projection(model, ReportView.AUDIT)
    assert by_name["engineering.html"].content == render_building_report_html(
        engineering_projection,
        options=HtmlRenderOptions(),
    ).encode("utf-8")
    assert by_name["audit.html"].content == render_building_report_html(
        audit_projection,
        options=HtmlRenderOptions(),
    ).encode("utf-8")
    assert by_name["engineering.pdf"].view == "ENGINEERING"
    assert by_name["audit.pdf"].view == "AUDIT"
    assert by_name["engineering.pdf"].source_projection_sha256
    assert by_name["engineering.pdf"].presentation_selection_sha256
    assert by_name["engineering.pdf"].source_render_sha256
    assert by_name["engineering.pdf"].options_sha256
    assert by_name["engineering.pdf"].renderer_toolchain_fingerprint


def test_pdf_delivery_calls_accepted_ur_1d_path(monkeypatch) -> None:
    model = _model()
    calls = []

    def fake_pdf(projection, *, selection=None, pdf_options=None):
        calls.append((projection.view.value, selection, pdf_options))
        content = b"%PDF-1.7\naccepted-stub"
        return ReportArtifact(
            logical_role="UNIFIED_ENGINEERING_REVIEW",
            format="PDF",
            media_type="application/pdf",
            filename=f"stub_{projection.view.value.lower()}.pdf",
            content=content,
            view=projection.view.value,
            source_report_id=projection.report_id,
            source_project_id=projection.project_id,
            source_projection_sha256="1" * 64,
            presentation_selection_sha256="2" * 64,
            source_render_sha256="3" * 64,
            options_sha256="4" * 64,
            renderer_toolchain_fingerprint="accepted-ur-1d-stub",
        )

    monkeypatch.setattr(package_module, "render_building_report_pdf", fake_pdf)
    artifacts = package_module.build_report_delivery_artifacts(model)
    by_name = _artifact_map(artifacts)

    assert [item[0] for item in calls] == ["ENGINEERING", "AUDIT"]
    assert by_name["engineering.pdf"].content == b"%PDF-1.7\naccepted-stub"
    assert by_name["engineering.pdf"].source_render_sha256 == "3" * 64
    assert by_name["engineering.pdf"].renderer_toolchain_fingerprint == "accepted-ur-1d-stub"


def test_final_package_is_byte_deterministic_reopens_and_manifest_verifies_every_payload() -> None:
    model = _model(
        (("A", "PASS"), ("B", "FAIL")),
        formula_rule="A",
        formula="Ndm <= 0.40 Ac fck",
    )
    first = build_building_report_package(model)
    second = build_building_report_package(model)

    assert first.filename == DEFAULT_PACKAGE_FILENAME
    assert first.view is None
    assert first.content == second.content
    assert first.sha256 == second.sha256
    assert first.artifact_id == second.artifact_id
    verify_building_report_package(first)

    with ZipFile(BytesIO(first.content), "r") as archive:
        assert archive.namelist() == list(DEFAULT_PACKAGE_MEMBERS)
        assert len(archive.namelist()) == 8
        content = {name: archive.read(name) for name in archive.namelist()}

    json.loads(content["building_report_model.json"])
    assert content["engineering.html"].startswith(b"<!doctype html>")
    assert content["audit.html"].startswith(b"<!doctype html>")
    assert len(PdfReader(BytesIO(content["engineering.pdf"])).pages) > 0
    assert len(PdfReader(BytesIO(content["audit.pdf"])).pages) > 0
    engineering_wb = load_workbook(BytesIO(content["engineering.xlsx"]), data_only=False)
    audit_wb = load_workbook(BytesIO(content["audit.xlsx"]), data_only=False)
    assert "Contributions" in engineering_wb.sheetnames
    assert "Source Manifest" in audit_wb.sheetnames

    manifest = json.loads(content["manifest.json"])
    entries = {item["filename"]: item for item in manifest["payloads"]}
    assert set(entries) == set(DEFAULT_PAYLOAD_FILENAMES)
    for filename, entry in entries.items():
        assert entry["byte_length"] == len(content[filename])
        assert entry["sha256"] == sha256(content[filename]).hexdigest()


def test_model_and_presentation_mutations_change_only_applicable_identities() -> None:
    base_model = _model((("A", "PASS"), ("B", "FAIL")))
    changed_model = _model(
        (("A", "PASS"), ("B", "FAIL")),
        title="Changed canonical model title",
    )
    base_json = export_building_report_model_json(base_model)
    changed_json = export_building_report_model_json(changed_model)
    assert base_json.artifact_id != changed_json.artifact_id

    filtered = ReportPresentationSelection(statuses=("FAIL",))
    default_package = build_building_report_package(base_model)
    filtered_package = build_building_report_package(
        base_model,
        engineering_selection=filtered,
    )
    assert default_package.artifact_id != filtered_package.artifact_id

    with ZipFile(BytesIO(default_package.content), "r") as default_zip:
        default_canonical = default_zip.read("building_report_model.json")
    with ZipFile(BytesIO(filtered_package.content), "r") as filtered_zip:
        filtered_canonical = filtered_zip.read("building_report_model.json")
    assert default_canonical == filtered_canonical == base_json.content


def test_upstream_pdf_integrity_error_is_not_swallowed(monkeypatch) -> None:
    def blocked(*args, **kwargs):
        raise PdfRenderIntegrityError("upstream integrity blocked")

    monkeypatch.setattr(package_module, "render_building_report_pdf", blocked)
    with pytest.raises(PdfRenderIntegrityError, match="upstream integrity blocked"):
        build_report_delivery_artifacts(_model())


def test_ur_1e_does_not_introduce_global_compliance_claims() -> None:
    artifacts = build_report_delivery_artifacts(_model())
    manifest = json.loads(_artifact_map(artifacts)["manifest.json"].content)
    assert manifest["integrity_scope"] == "DELIVERY_ONLY"
    assert manifest["global_compliance_verdict_emitted"] is False

    engineering = _workbook(_artifact_map(artifacts)["engineering.xlsx"])
    identity = _rows_by_header(engineering["Identity"])
    values = {row["key"]: row["value"] for row in identity}
    assert values["global_compliance_verdict_emitted"] is False
    assert "compliance_percentage" not in values
    assert "project_pass" not in values
    assert "project_fail" not in values
