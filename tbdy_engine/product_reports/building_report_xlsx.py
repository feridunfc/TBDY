"""Deterministic renderer-neutral XLSX delivery for UR-1E.

The workbook consumes only BuildingReportProjection plus presentation/export
options. It copies literal projection truth and never evaluates engineering,
regulatory, coverage, applicability, governing-selection, compliance, or
remediation logic.
"""
from __future__ import annotations

from dataclasses import dataclass
from datetime import datetime
from hashlib import sha256
from importlib.metadata import PackageNotFoundError, version as package_version
from io import BytesIO
import json
import re
from zipfile import ZIP_DEFLATED, ZipFile, ZipInfo

from tbdy_engine.product_reports.building_report_projection import (
    BuildingReportProjection,
    ReportView,
)
from tbdy_engine.product_reports.report_artifact import ReportArtifact
from tbdy_engine.product_reports.report_presentation_selection import (
    ReportPresentationSelection,
    resolve_presentation_selection,
)


SUPPORTED_OPENPYXL_VERSION = "3.1.5"
SUPPORTED_ET_XMLFILE_VERSION = "2.0.0"
_XLSX_MEDIA_TYPE = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
_FIXED_DT = datetime(1980, 1, 1, 0, 0, 0)
_FIXED_ZIP_DT = (1980, 1, 1, 0, 0, 0)
_SAFE_STEM_RE = re.compile(r"[^A-Za-z0-9._-]+")


class XlsxRenderIntegrityError(ValueError):
    """Raised when truthful deterministic XLSX export cannot be guaranteed."""


class XlsxBackendUnavailableError(RuntimeError):
    """Raised when the pinned deterministic XLSX backend is unavailable."""


@dataclass(frozen=True, slots=True)
class XlsxRenderOptions:
    """Bounded export-only options. Defaults preserve frozen package filenames."""

    filename_stem: str | None = None

    def __post_init__(self) -> None:
        if self.filename_stem is not None:
            if (
                not isinstance(self.filename_stem, str)
                or not self.filename_stem.strip()
                or self.filename_stem != self.filename_stem.strip()
            ):
                raise XlsxRenderIntegrityError(
                    "filename_stem must be None or a nonblank exact string"
                )

    def as_dict(self) -> dict[str, object]:
        return {"filename_stem": self.filename_stem}


def _canonical_json_bytes(payload: object) -> bytes:
    return (
        json.dumps(
            payload,
            ensure_ascii=True,
            sort_keys=True,
            separators=(",", ":"),
        )
        + "\n"
    ).encode("utf-8")


def _canonical_sha256(payload: object) -> str:
    return sha256(_canonical_json_bytes(payload)).hexdigest()


def xlsx_toolchain_identity() -> str:
    """Return the exact pinned XLSX exporter/backend identity."""

    try:
        openpyxl_version = package_version("openpyxl")
        et_xmlfile_version = package_version("et_xmlfile")
    except PackageNotFoundError as exc:
        raise XlsxBackendUnavailableError(
            "required XLSX backend package is not installed"
        ) from exc
    if openpyxl_version != SUPPORTED_OPENPYXL_VERSION:
        raise XlsxBackendUnavailableError(
            f"unsupported openpyxl version: {openpyxl_version} != {SUPPORTED_OPENPYXL_VERSION}"
        )
    if et_xmlfile_version != SUPPORTED_ET_XMLFILE_VERSION:
        raise XlsxBackendUnavailableError(
            f"unsupported et_xmlfile version: {et_xmlfile_version} != {SUPPORTED_ET_XMLFILE_VERSION}"
        )
    return json.dumps(
        {
            "profile": "tbdy-report-xlsx-toolchain.ur_1e.v1",
            "openpyxl": openpyxl_version,
            "et_xmlfile": et_xmlfile_version,
            "ooxml_normalization": "sorted-members-fixed-zip-metadata-deflate9",
        },
        ensure_ascii=True,
        sort_keys=True,
        separators=(",", ":"),
    )


def _safe_stem(value: str) -> str:
    stem = _SAFE_STEM_RE.sub("_", value).strip("._-")
    if not stem:
        raise XlsxRenderIntegrityError("filename_stem has no safe filename characters")
    return stem[:120]


def _artifact_filename(
    projection: BuildingReportProjection,
    options: XlsxRenderOptions,
) -> str:
    if options.filename_stem is None:
        return f"{projection.view.value.lower()}.xlsx"
    return f"{_safe_stem(options.filename_stem)}_{projection.view.value.lower()}.xlsx"


def _json_text(value: object) -> str:
    return json.dumps(
        value,
        ensure_ascii=True,
        sort_keys=True,
        separators=(",", ":"),
    )


def _display_value(value: object) -> object:
    if isinstance(value, (dict, list, tuple)):
        return _json_text(value)
    return value


def _set_cell(cell: object, value: object) -> None:
    """Write source/user strings as literal text, including spreadsheet-active prefixes."""

    from openpyxl.cell.cell import Cell

    if not isinstance(cell, Cell):
        raise TypeError("cell must be openpyxl Cell")
    value = _display_value(value)
    if isinstance(value, str):
        cell.value = value
        cell.data_type = "s"
    else:
        cell.value = value


def _write_sheet(
    workbook: object,
    name: str,
    headers: tuple[str, ...],
    rows: list[tuple[object, ...]],
    *,
    empty_message: str,
) -> None:
    ws = workbook.create_sheet(title=name)
    for column, header in enumerate(headers, start=1):
        _set_cell(ws.cell(row=1, column=column), header)
    if not rows:
        rows = [(empty_message,) + (None,) * (len(headers) - 1)]
    for row_index, row in enumerate(rows, start=2):
        if len(row) != len(headers):
            raise XlsxRenderIntegrityError(
                f"worksheet {name!r} row width does not match header width"
            )
        for column, value in enumerate(row, start=1):
            _set_cell(ws.cell(row=row_index, column=column), value)


def _projection_contributions(
    payload: dict[str, object],
    selected_refs: set[str],
) -> list[dict[str, object]]:
    values = payload.get("contributions", [])
    if not isinstance(values, list):
        raise XlsxRenderIntegrityError("projection contributions must be a list")
    result: list[dict[str, object]] = []
    for item in values:
        if not isinstance(item, dict) or "contribution_ref" not in item:
            raise XlsxRenderIntegrityError("projection contribution identity is invalid")
        if str(item["contribution_ref"]) in selected_refs:
            result.append(item)
    return result


def _identity_rows(
    projection: BuildingReportProjection,
    selected_count: int,
    canonical_count: int,
) -> list[tuple[object, ...]]:
    payload = projection.as_dict()
    keys = (
        "schema_version",
        "artifact_type",
        "view",
        "report_id",
        "project_id",
        "title",
        "report_integrity_status",
    )
    rows = [(key, payload.get(key)) for key in keys]
    rows.extend(
        (
            ("presentation_selected_contribution_count", selected_count),
            ("canonical_contribution_count", canonical_count),
            ("presentation_only", True),
            ("global_compliance_verdict_emitted", False),
        )
    )
    return rows


def _basis_rows(payload: dict[str, object], included: bool) -> list[tuple[object, ...]]:
    if not included:
        return []
    ledger = payload.get("project_basis")
    if not isinstance(ledger, dict):
        return []
    entries = ledger.get("entries")
    if not isinstance(entries, list):
        return []
    rows: list[tuple[object, ...]] = []
    for item in entries:
        if isinstance(item, dict):
            rows.append(
                (
                    item.get("key"),
                    item.get("label"),
                    item.get("value"),
                    item.get("unit"),
                    item.get("source_ids"),
                    item.get("note"),
                )
            )
    return rows


def _coverage_rows(payload: dict[str, object], included: bool) -> list[tuple[object, ...]]:
    if not included:
        return []
    summary = payload.get("coverage_summary")
    if not isinstance(summary, dict):
        return []
    return [(key, summary[key]) for key in sorted(summary)]


def _status_rows(payload: dict[str, object], included: bool) -> list[tuple[object, ...]]:
    if not included:
        return []
    values = payload.get("status_facets")
    if not isinstance(values, list):
        return []
    return [
        (item.get("status"), item.get("count"))
        for item in values
        if isinstance(item, dict)
    ]


def _component_rows(
    payload: dict[str, object],
    selected_refs: set[str],
    included: bool,
) -> list[tuple[object, ...]]:
    if not included:
        return []
    values = payload.get("component_facets")
    if not isinstance(values, list):
        return []
    rows: list[tuple[object, ...]] = []
    for item in values:
        if not isinstance(item, dict):
            continue
        refs = item.get("contribution_refs")
        if not isinstance(refs, list):
            refs = []
        selected = sorted(str(ref) for ref in refs if str(ref) in selected_refs)
        if not selected:
            continue
        rows.append(
            (
                item.get("component_type"),
                item.get("component_id"),
                item.get("contribution_count"),
                selected,
            )
        )
    return rows


def _contribution_rows(
    contributions: list[dict[str, object]],
    included: bool,
) -> list[tuple[object, ...]]:
    if not included:
        return []
    return [
        (
            item.get("contribution_ref"),
            item.get("slice_id"),
            item.get("contribution_kind"),
            item.get("component_type"),
            item.get("component_id"),
            item.get("title"),
            item.get("status"),
            item.get("report_source_refs"),
            item.get("authority_refs"),
            item.get("evidence_refs"),
            item.get("warnings"),
        )
        for item in contributions
    ]


def _field_rows(
    contributions: list[dict[str, object]],
    included: bool,
) -> list[tuple[object, ...]]:
    if not included:
        return []
    rows: list[tuple[object, ...]] = []
    for item in contributions:
        ref = item.get("contribution_ref")
        fields = item.get("summary_fields")
        if isinstance(fields, list):
            for field in fields:
                if isinstance(field, dict):
                    rows.append(
                        (
                            ref,
                            "SUMMARY",
                            None,
                            field.get("key"),
                            field.get("label"),
                            field.get("value"),
                            field.get("unit"),
                            field.get("role"),
                            field.get("note"),
                        )
                    )
        calculations = item.get("calculations")
        if isinstance(calculations, list):
            for calc in calculations:
                if not isinstance(calc, dict):
                    continue
                calculation_id = calc.get("calculation_id")
                for scope, key in (("CALC_INPUT", "inputs"), ("CALC_OUTPUT", "outputs")):
                    values = calc.get(key)
                    if not isinstance(values, list):
                        continue
                    for field in values:
                        if isinstance(field, dict):
                            rows.append(
                                (
                                    ref,
                                    scope,
                                    calculation_id,
                                    field.get("key"),
                                    field.get("label"),
                                    field.get("value"),
                                    field.get("unit"),
                                    field.get("role"),
                                    field.get("note"),
                                )
                            )
    return rows


def _calculation_rows(
    contributions: list[dict[str, object]],
    included: bool,
) -> list[tuple[object, ...]]:
    if not included:
        return []
    rows: list[tuple[object, ...]] = []
    for item in contributions:
        calculations = item.get("calculations")
        if not isinstance(calculations, list):
            continue
        for calc in calculations:
            if isinstance(calc, dict):
                rows.append(
                    (
                        item.get("contribution_ref"),
                        calc.get("calculation_id"),
                        calc.get("title"),
                        calc.get("formula"),
                        calc.get("governing_ref"),
                        calc.get("authority_refs"),
                        calc.get("evidence_refs"),
                    )
                )
    return rows


def _table_rows(
    contributions: list[dict[str, object]],
    included: bool,
) -> list[tuple[object, ...]]:
    if not included:
        return []
    result: list[tuple[object, ...]] = []
    for item in contributions:
        tables = item.get("tables")
        if not isinstance(tables, list):
            continue
        for table in tables:
            if not isinstance(table, dict):
                continue
            columns = table.get("columns")
            rows = table.get("rows")
            if not isinstance(columns, list) or not isinstance(rows, list):
                continue
            for row_index, row in enumerate(rows):
                if not isinstance(row, dict):
                    continue
                for column in columns:
                    result.append(
                        (
                            item.get("contribution_ref"),
                            table.get("table_id"),
                            table.get("title"),
                            table.get("purpose"),
                            row_index,
                            column,
                            row.get(str(column)),
                        )
                    )
    return result


def _warning_rows(
    contributions: list[dict[str, object]],
    included: bool,
) -> list[tuple[object, ...]]:
    if not included:
        return []
    rows: list[tuple[object, ...]] = []
    for item in contributions:
        values = item.get("warnings")
        if isinstance(values, list):
            for warning in values:
                rows.append((item.get("contribution_ref"), warning))
    return rows


def _analysis_rows(
    payload: dict[str, object],
    projection: BuildingReportProjection,
    included: bool,
) -> list[tuple[object, ...]]:
    if not included:
        return []
    key = (
        "analysis_basis_refs"
        if projection.view is ReportView.AUDIT
        else "analysis_basis_warnings"
    )
    values = payload.get(key)
    if not isinstance(values, list):
        return []
    return [
        (item.get("instance_id"), item.get("status"), item.get("source_ref"))
        for item in values
        if isinstance(item, dict)
    ]


def _action_rows(payload: dict[str, object], included: bool) -> list[tuple[object, ...]]:
    if not included:
        return []
    reconciliation = payload.get("coverage_reconciliation")
    if not isinstance(reconciliation, dict):
        return []
    rows: list[tuple[object, ...]] = []
    for key in (
        "required_action_finding_ids",
        "missing_action_finding_ids",
        "duplicate_action_finding_ids",
        "orphan_action_binding_finding_ids",
    ):
        values = reconciliation.get(key)
        if isinstance(values, list):
            rows.extend((key, value) for value in values)
    return rows


def _audit_trace_rows(payload: dict[str, object], included: bool) -> list[tuple[object, ...]]:
    if not included:
        return []
    reconciliation = payload.get("coverage_reconciliation")
    if not isinstance(reconciliation, dict):
        return []
    return [(key, reconciliation[key]) for key in sorted(reconciliation)]


def _binding_rows(payload: dict[str, object], included: bool) -> list[tuple[object, ...]]:
    if not included:
        return []
    values = payload.get("report_bindings")
    if not isinstance(values, list):
        return []
    return [
        (item.get("source_ref"), item.get("contribution_ref"))
        for item in values
        if isinstance(item, dict)
    ]


def _manifest_rows(payload: dict[str, object], included: bool) -> list[tuple[object, ...]]:
    if not included:
        return []
    manifest = payload.get("source_manifest")
    if not isinstance(manifest, dict):
        return []
    values = manifest.get("entries")
    if not isinstance(values, list):
        return []
    return [
        (
            item.get("source_id"),
            item.get("source_kind"),
            item.get("title"),
            item.get("fingerprint"),
            item.get("locator"),
            item.get("authority_refs"),
            item.get("evidence_refs"),
        )
        for item in values
        if isinstance(item, dict)
    ]


def _build_workbook_bytes(
    projection: BuildingReportProjection,
    selection: ReportPresentationSelection,
) -> bytes:
    try:
        from openpyxl import Workbook
        from openpyxl.packaging.core import DocumentProperties
    except Exception as exc:  # pragma: no cover - environment-specific import failure
        raise XlsxBackendUnavailableError("openpyxl could not be imported") from exc

    payload = projection.as_dict()
    values = payload.get("contributions")
    if not isinstance(values, list):
        raise XlsxRenderIntegrityError("projection contributions must be a list")
    canonical_refs = {
        str(item["contribution_ref"])
        for item in values
        if isinstance(item, dict) and "contribution_ref" in item
    }
    if len(canonical_refs) != len(values):
        raise XlsxRenderIntegrityError(
            "projection contribution_ref identities must be exact and unique"
        )
    selected_refs = set(selection.selected_contribution_refs(projection))
    contributions = _projection_contributions(payload, selected_refs)

    workbook = Workbook()
    workbook.remove(workbook.active)
    workbook.iso_dates = True
    workbook.properties = DocumentProperties(
        creator="tbdy_engine",
        lastModifiedBy="tbdy_engine",
        created=_FIXED_DT,
        modified=_FIXED_DT,
        title="TBDY Building Report Projection",
        subject="Deterministic UR-1E delivery workbook",
        description="Presentation-only workbook; canonical engineering truth remains upstream.",
        keywords="TBDY,report,projection",
        category="engineering-report-delivery",
    )

    _write_sheet(
        workbook,
        "Identity",
        ("key", "value"),
        _identity_rows(projection, len(selected_refs), len(canonical_refs)),
        empty_message="No canonical identity values are available.",
    )
    _write_sheet(
        workbook,
        "Project Basis",
        ("key", "label", "value", "unit", "source_ids", "note"),
        _basis_rows(payload, selection.include_overview),
        empty_message=(
            "Excluded from presentation selection."
            if not selection.include_overview
            else "No canonical project basis entries are available."
        ),
    )
    _write_sheet(
        workbook,
        "Coverage",
        ("key", "value"),
        _coverage_rows(payload, selection.include_coverage),
        empty_message=(
            "Excluded from presentation selection."
            if not selection.include_coverage
            else "No canonical coverage summary entries are available."
        ),
    )

    audit_evidence = projection.view is ReportView.AUDIT and selection.include_evidence
    if projection.view is ReportView.AUDIT:
        _write_sheet(
            workbook,
            "Coverage Trace",
            ("key", "value"),
            _audit_trace_rows(payload, audit_evidence),
            empty_message=(
                "Excluded from presentation selection."
                if not audit_evidence
                else "No canonical coverage reconciliation trace is available."
            ),
        )

    analysis_included = (
        selection.include_overview
        if projection.view is ReportView.ENGINEERING
        else (selection.include_overview or selection.include_evidence)
    )
    _write_sheet(
        workbook,
        "Analysis Basis",
        ("instance_id", "status", "source_ref"),
        _analysis_rows(payload, projection, analysis_included),
        empty_message=(
            "Excluded from presentation selection."
            if not analysis_included
            else "No analysis-basis warning/trace rows are available."
        ),
    )

    if projection.view is ReportView.AUDIT:
        _write_sheet(
            workbook,
            "Report Bindings",
            ("source_ref", "contribution_ref"),
            _binding_rows(payload, audit_evidence),
            empty_message=(
                "Excluded from presentation selection."
                if not audit_evidence
                else "No report binding rows are available."
            ),
        )
        _write_sheet(
            workbook,
            "Source Manifest",
            (
                "source_id",
                "source_kind",
                "title",
                "fingerprint",
                "locator",
                "authority_refs",
                "evidence_refs",
            ),
            _manifest_rows(payload, audit_evidence),
            empty_message=(
                "Excluded from presentation selection."
                if not audit_evidence
                else "No source manifest rows are available."
            ),
        )

    _write_sheet(
        workbook,
        "Status Facets",
        ("status", "count"),
        _status_rows(payload, selection.include_overview),
        empty_message=(
            "Excluded from presentation selection."
            if not selection.include_overview
            else "No canonical status facets are available."
        ),
    )
    _write_sheet(
        workbook,
        "Components",
        ("component_type", "component_id", "canonical_contribution_count", "selected_refs"),
        _component_rows(payload, selected_refs, selection.include_components),
        empty_message=(
            "Excluded from presentation selection."
            if not selection.include_components
            else "No components are included by this presentation selection."
        ),
    )
    _write_sheet(
        workbook,
        "Contributions",
        (
            "contribution_ref",
            "slice_id",
            "contribution_kind",
            "component_type",
            "component_id",
            "title",
            "status",
            "report_source_refs",
            "authority_refs",
            "evidence_refs",
            "warnings",
        ),
        _contribution_rows(contributions, selection.include_results),
        empty_message=(
            "Excluded from presentation selection."
            if not selection.include_results
            else "No contributions are included by this presentation selection."
        ),
    )
    _write_sheet(
        workbook,
        "Fields",
        (
            "contribution_ref",
            "field_scope",
            "calculation_id",
            "key",
            "label",
            "value",
            "unit",
            "role",
            "note",
        ),
        _field_rows(contributions, selection.include_results),
        empty_message=(
            "Excluded from presentation selection."
            if not selection.include_results
            else "No resolved fields are available for selected contributions."
        ),
    )
    _write_sheet(
        workbook,
        "Calculations",
        (
            "contribution_ref",
            "calculation_id",
            "title",
            "formula_text",
            "governing_ref",
            "authority_refs",
            "evidence_refs",
        ),
        _calculation_rows(contributions, selection.include_results),
        empty_message=(
            "Excluded from presentation selection."
            if not selection.include_results
            else "No resolved calculation displays are available."
        ),
    )
    _write_sheet(
        workbook,
        "Table Rows",
        (
            "contribution_ref",
            "table_id",
            "title",
            "purpose",
            "row_index",
            "column",
            "value",
        ),
        _table_rows(contributions, selection.include_results),
        empty_message=(
            "Excluded from presentation selection."
            if not selection.include_results
            else "No resolved table rows are available."
        ),
    )
    _write_sheet(
        workbook,
        "Actions",
        ("action_reconciliation_key", "finding_id"),
        _action_rows(payload, selection.include_actions),
        empty_message=(
            "Excluded from presentation selection."
            if not selection.include_actions
            else "No canonical action records are available in this projection."
        ),
    )
    _write_sheet(
        workbook,
        "Warnings",
        ("contribution_ref", "warning"),
        _warning_rows(contributions, selection.include_results),
        empty_message=(
            "Excluded from presentation selection."
            if not selection.include_results
            else "No upstream warnings are available for selected contributions."
        ),
    )

    if getattr(workbook, "_external_links", ()):
        raise XlsxRenderIntegrityError("external workbook links are forbidden")

    buffer = BytesIO()
    workbook.save(buffer)
    return _normalize_ooxml(buffer.getvalue())


def _normalize_ooxml(raw_bytes: bytes) -> bytes:
    """Canonicalize OOXML payload plus ZIP metadata/member order after serialization."""

    source = BytesIO(raw_bytes)
    output = BytesIO()
    with ZipFile(source, "r") as archive:
        names = archive.namelist()
        if len(names) != len(set(names)):
            raise XlsxRenderIntegrityError("backend produced duplicate OOXML member names")
        with ZipFile(
            output,
            "w",
            compression=ZIP_DEFLATED,
            compresslevel=9,
            allowZip64=False,
        ) as normalized:
            for name in sorted(names):
                if name.startswith("/") or ".." in name.split("/"):
                    raise XlsxRenderIntegrityError("unsafe OOXML member path")
                data = archive.read(name)
                if name == "docProps/core.xml":
                    data, count = re.subn(
                        rb"(<dcterms:modified\b[^>]*>)[^<]*(</dcterms:modified>)",
                        rb"\g<1>1980-01-01T00:00:00Z\g<2>",
                        data,
                        count=1,
                    )
                    if count != 1:
                        raise XlsxRenderIntegrityError(
                            "OOXML core properties must contain exactly one modified timestamp"
                        )
                info = ZipInfo(name, date_time=_FIXED_ZIP_DT)
                info.compress_type = ZIP_DEFLATED
                info.create_system = 3
                info.external_attr = (0o100644 & 0xFFFF) << 16
                info.flag_bits = 0x800
                normalized.writestr(
                    info,
                    data,
                    compress_type=ZIP_DEFLATED,
                    compresslevel=9,
                )
    normalized_bytes = output.getvalue()
    _assert_safe_ooxml(normalized_bytes)
    return normalized_bytes


def _assert_safe_ooxml(content: bytes) -> None:
    with ZipFile(BytesIO(content), "r") as archive:
        names = archive.namelist()
        lowered = [name.lower() for name in names]
        if any("vbaproject" in name or "macros" in name for name in lowered):
            raise XlsxRenderIntegrityError("macros/VBA are forbidden")
        if any("externallinks/" in name for name in lowered):
            raise XlsxRenderIntegrityError("external workbook links are forbidden")
        for name in names:
            if not name.endswith((".xml", ".rels")):
                continue
            data = archive.read(name)
            if b'TargetMode="External"' in data or b"TargetMode='External'" in data:
                raise XlsxRenderIntegrityError("external OOXML relationships are forbidden")


def render_building_report_xlsx(
    projection: BuildingReportProjection,
    *,
    selection: ReportPresentationSelection | None = None,
    options: XlsxRenderOptions | None = None,
) -> ReportArtifact:
    """Render one Engineering/Audit projection as deterministic formula-free XLSX."""

    if not isinstance(projection, BuildingReportProjection):
        raise TypeError("projection must be BuildingReportProjection")
    if projection.view not in (ReportView.ENGINEERING, ReportView.AUDIT):
        raise XlsxRenderIntegrityError("UR-1E supports only ENGINEERING and AUDIT")
    if projection.report_integrity_status != "RECONCILED":
        raise XlsxRenderIntegrityError(
            "report_integrity_status must be canonical RECONCILED before XLSX export"
        )
    if options is None:
        options = XlsxRenderOptions()
    if not isinstance(options, XlsxRenderOptions):
        raise TypeError("options must be XlsxRenderOptions or None")

    toolchain = xlsx_toolchain_identity()
    resolved_selection = resolve_presentation_selection(projection, selection)
    content = _build_workbook_bytes(projection, resolved_selection)

    projection_sha256 = sha256(projection.to_json().encode("utf-8")).hexdigest()
    selection_sha256 = _canonical_sha256(resolved_selection.as_dict())
    options_sha256 = _canonical_sha256(options.as_dict())

    return ReportArtifact(
        logical_role="UNIFIED_ENGINEERING_REVIEW",
        format="XLSX",
        media_type=_XLSX_MEDIA_TYPE,
        filename=_artifact_filename(projection, options),
        content=content,
        view=projection.view.value,
        source_report_id=projection.report_id,
        source_project_id=projection.project_id,
        source_projection_sha256=projection_sha256,
        presentation_selection_sha256=selection_sha256,
        options_sha256=options_sha256,
        renderer_toolchain_fingerprint=toolchain,
    )


__all__ = [
    "SUPPORTED_ET_XMLFILE_VERSION",
    "SUPPORTED_OPENPYXL_VERSION",
    "XlsxBackendUnavailableError",
    "XlsxRenderIntegrityError",
    "XlsxRenderOptions",
    "render_building_report_xlsx",
    "xlsx_toolchain_identity",
]