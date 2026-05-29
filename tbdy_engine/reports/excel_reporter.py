from __future__ import annotations

import json
from collections import Counter
from dataclasses import asdict, is_dataclass
from pathlib import Path
from typing import Any, Iterable, Mapping, Sequence


NO_DATA = "NO_DATA"

SHEETS = [
    "Summary",
    "Kiriş Kesme",
    "Kiriş Donatı Seçimi",
    "Beam Checks",
    "Evidence",
]

SHEAR_COLUMNS = [
    "Kat",
    "Kiriş",
    "Kesit",
    "Tasarım Kuvveti Vd (kN)",
    "Eksenel Kuvvet P (kN)",
    "Minimum Kol Adedi",
    "Seçilen Kol Adedi",
    "Kesme Donatı Çapı (mm)",
    "Kol Adet - Çap",
    "B (m)",
    "H (m)",
    "d (m)",
    "Asmin (cm²)",
    "Asw (cm²)",
    "Vmax (kN)",
    "Kesit Kontrol (%)",
    "Vc (kN)",
    "Vw (kN)",
    "Vr (kN)",
    "Oran (%)",
    "Durum",
    "Check ID",
]

FLEXURE_COLUMNS = [
    "Kat",
    "Kiriş",
    "Kesit",
    "I Üst - Seçilen Donatı",
    "I Üst - Gerekli Alan (cm²)",
    "I Üst - Seçilen Alan (cm²)",
    "Üst Açıklık - Seçilen Donatı",
    "Üst Açıklık - Gerekli Alan (cm²)",
    "Üst Açıklık - Seçilen Alan (cm²)",
    "J Üst - Seçilen Donatı",
    "J Üst - Gerekli Alan (cm²)",
    "J Üst - Seçilen Alan (cm²)",
    "Alt - Seçilen Donatı",
    "Alt - Gerekli Alan (cm²)",
    "Alt - Seçilen Alan (cm²)",
    "B (m)",
    "H (m)",
    "L (m)",
    "Toplam Gerekli Alan (cm²)",
    "Seçilen Toplam Alan (cm²)",
    "Fark (%)",
    "Durum",
    "Check ID",
]

BEAM_CHECK_COLUMNS = [
    "Component",
    "Story",
    "Section",
    "Check Type",
    "Status",
    "Demand",
    "Capacity",
    "Ratio",
    "Unit",
    "Code Ref",
    "Messages",
    "Check ID",
]

EVIDENCE_COLUMNS = [
    "Check ID",
    "Component",
    "Check Type",
    "Source Table",
    "Source Row",
    "Source Columns",
    "Evidence Key",
    "Raw Evidence JSON",
]

STATUS_DISPLAY = {
    "OK": "✓",
    "FAIL": "✗",
    "WARNING": "!",
    "NO_DATA": "NO_DATA",
    "ERROR": "ERROR",
}

CHECK_TYPE_DISPLAY = {
    "beam_geometry": "Geometry",
    "beam_flexure": "Flexure",
    "beam_shear": "Shear",
}


class ExcelReporter:
    def generate(self, check_results: Sequence[Any], output_path="engine_report.xlsx"):
        try:
            import openpyxl
            from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
        except Exception:
            return None

        checks = [_check_to_dict(check) for check in list(check_results)]
        styles = _Styles(openpyxl, Alignment, Border, Font, PatternFill, Side)

        wb = openpyxl.Workbook()
        summary_ws = wb.active
        summary_ws.title = "Summary"
        _write_summary(summary_ws, checks, styles)
        _write_shear_sheet(wb.create_sheet("Kiriş Kesme"), checks, styles)
        _write_flexure_sheet(wb.create_sheet("Kiriş Donatı Seçimi"), checks, styles)
        _write_beam_checks_sheet(wb.create_sheet("Beam Checks"), checks, styles)
        _write_evidence_sheet(wb.create_sheet("Evidence"), checks, styles)

        for sheet_name in SHEETS:
            _finish_sheet(wb[sheet_name], styles)

        path = Path(output_path)
        path.parent.mkdir(parents=True, exist_ok=True)
        wb.save(path)
        return str(path)


class _Styles:
    def __init__(self, openpyxl, Alignment, Border, Font, PatternFill, Side):
        self.alignment_center = Alignment(horizontal="center", vertical="center", wrap_text=True)
        self.alignment_left = Alignment(horizontal="left", vertical="center", wrap_text=True)
        self.alignment_right = Alignment(horizontal="right", vertical="center", wrap_text=True)
        self.title_font = Font(bold=True, size=14)
        self.header_font = Font(bold=True, color="FFFFFF")
        self.bold_font = Font(bold=True)
        self.header_fill = PatternFill("solid", fgColor="666666")
        self.block_fill = PatternFill("solid", fgColor="D9EAF7")
        self.no_data_fill = PatternFill("solid", fgColor="D9D9D9")
        self.ok_fill = PatternFill("solid", fgColor="C6EFCE")
        self.fail_fill = PatternFill("solid", fgColor="FFC7CE")
        self.warning_fill = PatternFill("solid", fgColor="FFEB9C")
        side = Side(style="thin", color="B7B7B7")
        self.border = Border(left=side, right=side, top=side, bottom=side)


def _write_summary(ws, checks: list[dict[str, Any]], styles: _Styles) -> None:
    _title(ws, "ENGINE REPORT SUMMARY", 1, 1, 2, styles)
    ws.append(["Metric", "Value"])
    _style_header_row(ws, 2, 1, 2, styles)
    for metric, value in _summary_items(checks):
        ws.append([metric, value])


def _write_shear_sheet(ws, checks: list[dict[str, Any]], styles: _Styles) -> None:
    _title(ws, "KİRİŞ KESME KAPASİTE HESABI", 1, 1, len(SHEAR_COLUMNS), styles)
    summary_labels = [
        ("Beton Sınıfı", _evidence_summary_value(checks, "beam_shear", "concrete_class")),
        ("Donatı Sınıfı", _evidence_summary_value(checks, "beam_shear", "rebar_class")),
        ("Paspayı (m)", _evidence_summary_value(checks, "beam_shear", "cover_m")),
        ("Minimum Kol Adedi", _evidence_summary_value(checks, "beam_shear", "min_leg_count")),
        ("Etriye Aralığı (m)", _evidence_summary_value(checks, "beam_shear", "stirrup_spacing_m")),
        ("Etriye Çapı (mm)", _evidence_summary_value(checks, "beam_shear", "stirrup_diameter")),
        ("Deprem Katkısı Dikkate Alındı", _evidence_summary_value(checks, "beam_shear", "earthquake_contribution_considered")),
        ("Birim Notu", "Kuvvet birimi: kN, uzunluk birimi: m"),
        ("TOPLAM", _count_type(checks, "beam_shear")),
        ("YETERLİ", _count_type_status(checks, "beam_shear", "OK")),
        ("YETERSİZ", _count_type_status(checks, "beam_shear", "FAIL")),
    ]
    _summary_block(ws, summary_labels, 3, styles)
    header_row = 16
    _write_table(ws, header_row, SHEAR_COLUMNS, [_shear_row(c) for c in checks if c.get("check_type") == "beam_shear"], styles)


def _write_flexure_sheet(ws, checks: list[dict[str, Any]], styles: _Styles) -> None:
    _title(ws, "KİRİŞ DONATI SEÇİMİ", 1, 1, len(FLEXURE_COLUMNS), styles)
    summary_labels = [
        ("Beton Sınıfı", _evidence_summary_value(checks, "beam_flexure", "concrete_class")),
        ("Donatı Sınıfı", _evidence_summary_value(checks, "beam_flexure", "rebar_class")),
        ("Seçim Fazlası Oranı", _evidence_summary_value(checks, "beam_flexure", "selection_excess_ratio")),
        ("Unique Name", _evidence_summary_value(checks, "beam_flexure", "unique_name")),
        ("Montaj Akstan Al", _evidence_summary_value(checks, "beam_flexure", "montaj_akstan_al")),
        ("Benzer Kiriş Kullan", _evidence_summary_value(checks, "beam_flexure", "use_similar_beam")),
        ("Birim Notu", "Alan birimi: cm², uzunluk birimi: m, çap birimi: mm"),
        ("TOPLAM", _count_type(checks, "beam_flexure")),
        ("YETERLİ", _count_type_status(checks, "beam_flexure", "OK")),
        ("YETERSİZ", _count_type_status(checks, "beam_flexure", "FAIL")),
    ]
    _summary_block(ws, summary_labels, 3, styles)
    header_row = 15
    _write_table(ws, header_row, FLEXURE_COLUMNS, [_flexure_row(c) for c in checks if c.get("check_type") == "beam_flexure"], styles)


def _write_beam_checks_sheet(ws, checks: list[dict[str, Any]], styles: _Styles) -> None:
    _title(ws, "BEAM CHECKS", 1, 1, len(BEAM_CHECK_COLUMNS), styles)
    rows = []
    for check in checks:
        rows.append([
            _value(check.get("component")),
            _value(check.get("story")),
            _value(check.get("section")),
            CHECK_TYPE_DISPLAY.get(str(check.get("check_type") or ""), _value(check.get("check_type"))),
            _status_display(check),
            _value(check.get("demand")),
            _value(check.get("capacity")),
            _ratio_percent(check.get("ratio")),
            _value(check.get("unit")),
            _value(check.get("code_ref")),
            _messages(check.get("messages")),
            _value(check.get("id")),
        ])
    _write_table(ws, 3, BEAM_CHECK_COLUMNS, rows, styles)


def _write_evidence_sheet(ws, checks: list[dict[str, Any]], styles: _Styles) -> None:
    _title(ws, "EVIDENCE", 1, 1, len(EVIDENCE_COLUMNS), styles)
    rows = []
    for check in checks:
        evidence = _evidence(check)
        if not evidence:
            source_table = source_row = source_columns = evidence_key = raw = NO_DATA
        else:
            source_table = _value(_first(evidence, "source_table", "table_name"))
            source_row = _value(evidence.get("source_row"))
            source_columns = _json_or_no_data(evidence.get("source_columns"))
            evidence_key = _value(_first(evidence, "key", "evidence_key"))
            raw = json.dumps(evidence, ensure_ascii=False, sort_keys=True, default=str)
        rows.append([
            _value(check.get("id")),
            _value(check.get("component")),
            _value(check.get("check_type")),
            source_table,
            source_row,
            source_columns,
            evidence_key,
            raw,
        ])
    _write_table(ws, 3, EVIDENCE_COLUMNS, rows, styles)


def _summary_items(checks: list[dict[str, Any]]) -> list[tuple[str, Any]]:
    ids = [check.get("id") for check in checks if check.get("id")]
    counts = Counter(ids)
    return [
        ("Total Checks", len(checks)),
        ("OK", sum(1 for c in checks if c.get("status") == "OK")),
        ("FAIL", sum(1 for c in checks if c.get("status") == "FAIL")),
        ("WARNING", sum(1 for c in checks if c.get("status") == "WARNING")),
        ("NO_DATA", sum(1 for c in checks if c.get("status") == "NO_DATA")),
        ("ERROR", sum(1 for c in checks if c.get("status") == "ERROR")),
        ("Unique Components", len({c.get("component") for c in checks if c.get("component")})),
        ("Duplicate Check IDs", sum(count - 1 for count in counts.values() if count > 1)),
        ("Beam Shear Checks", _count_type(checks, "beam_shear")),
        ("Beam Flexure Checks", _count_type(checks, "beam_flexure")),
        ("Beam Geometry Checks", _count_type(checks, "beam_geometry")),
    ]


def _shear_row(check: dict[str, Any]) -> list[Any]:
    evidence = _evidence(check)
    return [
        _value(check.get("story")),
        _value(check.get("component")),
        _value(check.get("section")),
        _kN_value(_first(evidence, "Vd", "demand"), check.get("demand"), check.get("unit")),
        _value(_first(evidence, "P", "axial_force")),
        _value(evidence.get("min_leg_count")),
        _value(evidence.get("selected_leg_count")),
        _value(evidence.get("stirrup_diameter")),
        _value(evidence.get("leg_diameter_label")),
        _value(_first(evidence, "B", "b_m")),
        _value(_first(evidence, "H", "h_m")),
        _value(_first(evidence, "d", "d_m")),
        _value(_first(evidence, "Asmin", "Asmin_cm2")),
        _value(_first(evidence, "Asw", "Asw_cm2")),
        _value(_first(evidence, "Vmax", "vmax")),
        _ratio_percent(_first(evidence, "section_control_ratio", "section_control")),
        _value(evidence.get("Vc")),
        _value(evidence.get("Vw")),
        _kN_value(_first(evidence, "Vr"), check.get("capacity"), check.get("unit")),
        _ratio_percent(check.get("ratio")),
        _status_display(check),
        _value(check.get("id")),
    ]


def _flexure_row(check: dict[str, Any]) -> list[Any]:
    evidence = _evidence(check)
    return [
        _value(check.get("story")),
        _value(check.get("component")),
        _value(check.get("section")),
        _value(evidence.get("i_top_selected_rebar")),
        _value(evidence.get("i_top_required_area")),
        _value(evidence.get("i_top_selected_area")),
        _value(evidence.get("span_top_selected_rebar")),
        _value(evidence.get("span_top_required_area")),
        _value(evidence.get("span_top_selected_area")),
        _value(evidence.get("j_top_selected_rebar")),
        _value(evidence.get("j_top_required_area")),
        _value(evidence.get("j_top_selected_area")),
        _value(evidence.get("bottom_selected_rebar")),
        _value(evidence.get("bottom_required_area")),
        _value(evidence.get("bottom_selected_area")),
        _value(_first(evidence, "B", "b_m")),
        _value(_first(evidence, "H", "h_m")),
        _value(_first(evidence, "L", "l_m")),
        _value(evidence.get("total_required_area")),
        _value(evidence.get("total_selected_area")),
        _ratio_percent(_first(evidence, "excess_ratio") if _first(evidence, "excess_ratio") is not None else check.get("ratio")),
        _status_display(check),
        _value(check.get("id")),
    ]


def _title(ws, title: str, row: int, start_col: int, end_col: int, styles: _Styles) -> None:
    ws.merge_cells(start_row=row, start_column=start_col, end_row=row, end_column=end_col)
    cell = ws.cell(row=row, column=start_col, value=title)
    cell.font = styles.title_font
    cell.alignment = styles.alignment_center
    cell.fill = styles.block_fill


def _summary_block(ws, items: list[tuple[str, Any]], start_row: int, styles: _Styles) -> None:
    row = start_row
    for label, value in items:
        ws.cell(row=row, column=1, value=label)
        ws.cell(row=row, column=2, value=_value(value))
        ws.cell(row=row, column=1).font = styles.bold_font
        ws.cell(row=row, column=1).fill = styles.block_fill
        row += 1


def _write_table(ws, header_row: int, headers: list[str], rows: Iterable[list[Any]], styles: _Styles) -> None:
    for col, header in enumerate(headers, start=1):
        ws.cell(row=header_row, column=col, value=header)
    _style_header_row(ws, header_row, 1, len(headers), styles)
    for row_values in rows:
        ws.append(row_values)
    if ws.max_row == header_row:
        ws.append([NO_DATA] + [None] * (len(headers) - 1))
    ws.auto_filter.ref = ws.dimensions
    ws.freeze_panes = ws.cell(row=header_row + 1, column=1).coordinate


def _style_header_row(ws, row: int, start_col: int, end_col: int, styles: _Styles) -> None:
    for col in range(start_col, end_col + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = styles.header_font
        cell.fill = styles.header_fill
        cell.alignment = styles.alignment_center
        cell.border = styles.border


def _finish_sheet(ws, styles: _Styles) -> None:
    for row in ws.iter_rows():
        for cell in row:
            cell.border = styles.border
            if cell.value == NO_DATA:
                cell.fill = styles.no_data_fill
            if isinstance(cell.value, (int, float)):
                cell.alignment = styles.alignment_right
            else:
                cell.alignment = styles.alignment_center if cell.value in STATUS_DISPLAY.values() else styles.alignment_left
            if cell.value == "✓":
                cell.fill = styles.ok_fill
            elif cell.value == "✗" or cell.value == "ERROR":
                cell.fill = styles.fail_fill
            elif cell.value == "!":
                cell.fill = styles.warning_fill
    for column_cells in ws.columns:
        values = [str(cell.value) for cell in column_cells if cell.value is not None]
        width = max([len(value) for value in values] + [10])
        ws.column_dimensions[column_cells[0].column_letter].width = min(max(width + 2, 12), 38)


def _count_type(checks: list[dict[str, Any]], check_type: str) -> int:
    return sum(1 for c in checks if c.get("check_type") == check_type)


def _count_type_status(checks: list[dict[str, Any]], check_type: str, status: str) -> int:
    return sum(1 for c in checks if c.get("check_type") == check_type and c.get("status") == status)


def _evidence_summary_value(checks: list[dict[str, Any]], check_type: str, key: str) -> Any:
    for check in checks:
        if check.get("check_type") != check_type:
            continue
        evidence = _evidence(check)
        if key in evidence and evidence[key] not in (None, ""):
            return evidence[key]
    return NO_DATA


def _status_display(check: dict[str, Any]) -> str:
    return STATUS_DISPLAY.get(str(check.get("status") or ""), str(check.get("status") or NO_DATA))


def _ratio_percent(value: Any) -> Any:
    if value in (None, ""):
        return NO_DATA
    try:
        return float(value) * 100.0
    except Exception:
        return value


def _kN_value(evidence_value: Any, fallback_value: Any, unit: Any) -> Any:
    if evidence_value not in (None, ""):
        return evidence_value
    if str(unit or "") == "kN" and fallback_value not in (None, ""):
        return fallback_value
    return NO_DATA


def _value(value: Any) -> Any:
    if value in (None, ""):
        return NO_DATA
    if isinstance(value, (dict, list, tuple)):
        return json.dumps(value, ensure_ascii=False, sort_keys=True, default=str)
    return value


def _json_or_no_data(value: Any) -> Any:
    if value in (None, "", [], (), {}):
        return NO_DATA
    return json.dumps(value, ensure_ascii=False, sort_keys=True, default=str)


def _first(mapping: Mapping[str, Any], *keys: str) -> Any:
    for key in keys:
        value = mapping.get(key)
        if value not in (None, ""):
            return value
    return None


def _messages(value: Any) -> str:
    if value in (None, ""):
        return NO_DATA
    if isinstance(value, str):
        return value
    if isinstance(value, (list, tuple)):
        return " | ".join(str(item) for item in value if str(item)) or NO_DATA
    return str(value)


def _evidence(check: dict[str, Any]) -> Mapping[str, Any]:
    evidence = check.get("evidence")
    return evidence if isinstance(evidence, Mapping) else {}


def _check_to_dict(check: Any) -> dict[str, Any]:
    if hasattr(check, "to_dict") and callable(check.to_dict):
        value = check.to_dict()
        if isinstance(value, dict):
            return dict(value)
    if is_dataclass(check):
        return asdict(check)
    if isinstance(check, dict):
        return dict(check)
    if hasattr(check, "__dict__"):
        return dict(vars(check))
    raise TypeError(f"Unsupported CheckResult object: {type(check).__name__}")
