from __future__ import annotations

import json
import os
import subprocess
from datetime import datetime, timezone
from pathlib import Path
from typing import Mapping, Sequence

from tbdy_engine.design.beams.etabs_live_smoke_harness import run_etabs_beamcore_smoke_from_provider
from tbdy_engine.design.beams.etabs_single_beam_frameforce_runner import (
    SingleBeamFrameForceError,
    _attach_to_open_etabs,
    _frame_section,
    _frame_story,
    _safe_model_name,
    build_existing_p4_payload_from_frameforce,
    extract_frameforce_envelope,
    _r21a_raw_signed_action_fields,
    _r21a_raw_signed_governing_evidence,
)

class StoryBeamBatchError(RuntimeError):
    def __init__(self, stage: str, message: str):
        super().__init__(f"{stage}: {message}")
        self.stage = stage
        self.message = message


class _PayloadProvider:
    def __init__(self, payload: Mapping[str, object]) -> None:
        self._payload = payload

    def get_beam_payload(self) -> Mapping[str, object]:
        return self._payload


def run_live_etabs_story_beam_batch(
    *,
    story: str,
    combos: Sequence[str],
    output_dir: Path,
    sap_model: object | None = None,
    min_beams: int = 3,
    max_beams: int | None = None,
) -> Mapping[str, object]:
    """Run BeamCore checks for a selected ETABS story batch using real FrameForce actions.

    This uses the existing P4-compatible mapping from R7A. It does not introduce a new
    payload type/schema and does not modify BeamCore calculators.
    """
    if not story:
        raise StoryBeamBatchError("story_beam_discovery", "selected story is required")
    combos = [combo for combo in combos if combo]
    if len(combos) < 2:
        raise StoryBeamBatchError("force_extract", "at least two combos are required")
    if min_beams < 1:
        raise StoryBeamBatchError("story_beam_discovery", "min_beams must be positive")

    output_dir = Path(output_dir)
    output_dir.mkdir(parents=True, exist_ok=True)

    sap_model = sap_model if sap_model is not None else _attach_to_open_etabs()
    discovered = discover_story_beams(sap_model=sap_model, story=story)
    if len(discovered) < min_beams:
        raise StoryBeamBatchError(
            "story_beam_discovery",
            f"selected story has {len(discovered)} beams; minimum required is {min_beams}",
        )

    selected = discovered[:max_beams] if max_beams is not None else discovered
    processed: list[dict[str, object]] = []
    failures: list[dict[str, object]] = []

    for beam in selected:
        try:
            envelope = extract_frameforce_envelope(sap_model=sap_model, beam_name=beam["object_name"], combos=combos)
            payload = build_existing_p4_payload_from_frameforce(
                beam_name=beam["object_name"],
                beam_info={"story": beam["story"], "section": beam["section"]},
                combos=combos,
                envelope=envelope,
                sap_model=sap_model,
            )
            # Preserve ETABS label/object metadata without changing the accepted P4-compatible payload shape.
            payload["source"]["object_name"] = beam["object_name"]
            payload["source"]["label"] = beam["label"]
            payload["source"]["story"] = beam["story"]

            beam_output = output_dir / _safe_name(beam["object_name"])
            result = run_etabs_beamcore_smoke_from_provider(provider=_PayloadProvider(payload), output_dir=beam_output)
            actions = dict(payload["actions"])
            actions.update(_r21a_raw_signed_action_fields(dict(envelope)))
            governing = _r21a_raw_signed_governing_evidence(dict(envelope))
            _write_raw_signed_evidence_sheet(Path(result["xlsx_path"]), governing)
            processed.append(
                {
                    "object_name": beam["object_name"],
                    "label": beam["label"],
                    "story": beam["story"],
                    "section": beam["section"],
                    "actions_source": "etabs_results",
                    "Ve_source": "etabs_results_envelope",
                    "actions": actions,
                    "governing": governing,
                    "BeamCoreResult produced": True,
                    "BeamCore checks executed": True,
                    "beam_core_status": result["beam_core_status"],
                    "check_count": result["check_count"],
                    "capacity_design_check_statuses": _capacity_design_check_statuses(result["check_types"]),
                    "artifact_paths": {
                        "json": str(result["json_path"]),
                        "xlsx": str(result["xlsx_path"]),
                    },
                }
            )
        except Exception as exc:
            stage = getattr(exc, "stage", "beam_core")
            failures.append(
                {
                    "object_name": beam["object_name"],
                    "label": beam.get("label"),
                    "story": beam.get("story"),
                    "stage": stage,
                    "error": str(exc),
                }
            )

    if len(processed) < min_beams:
        stage = "batch_minimum_processed" if processed else (failures[0]["stage"] if failures else "batch_minimum_processed")
        raise StoryBeamBatchError(
            str(stage),
            f"processed {len(processed)} beams; minimum required is {min_beams}; failures={failures}",
        )

    summary = {
        "run_timestamp": datetime.now(timezone.utc).isoformat(),
        "branch": _git(["rev-parse", "--abbrev-ref", "HEAD"]),
        "commit": _git(["rev-parse", "--short", "HEAD"]),
        "selected_story": story,
        "selected_combos": list(combos),
        "units": {"force": "kN", "moment": "kNm", "length": "mm"},
        "min_beam_requirement": min_beams,
        "beam_count_discovered": len(discovered),
        "beam_count_processed": len(processed),
        "beam_count_failed": len(failures),
        "actions_source": "etabs_results",
        "envelope_rules_document": "docs/beam_core_etabs_envelope_selection_rules.md",
        "BeamCore checks executed": True,
        "BeamCoreResult produced": True,
        "ETABS FrameForce bridge observed": True,
        "Selected story batch smoke observed": True,
        "beams": processed,
        "failures": failures,
        "forbidden_claims": [
            "ETABS_VALIDATED = TRUE",
            "DESIGN_ENGINE_VALIDATED = TRUE",
            "ETABS_BRIDGE = PROVEN_FOR_ALL_MODELS",
            "PRODUCTION_READY = TRUE",
            "RELEASE_READY = TRUE",
            "CODE_COMPLIANCE_PROVEN = TRUE",
        ],
    }

    json_path = output_dir / "story_beam_batch_summary.json"
    md_path = output_dir / "story_beam_batch_summary.md"
    json_path.write_text(json.dumps(summary, indent=2, ensure_ascii=False), encoding="utf-8")
    md_path.write_text(_render_markdown_summary(summary), encoding="utf-8")

    return {
        "status": "OK",
        "selected_story": story,
        "selected_combos": list(combos),
        "actions_source": "etabs_results",
        "beam_count_discovered": len(discovered),
        "beam_count_processed": len(processed),
        "beam_count_failed": len(failures),
        "json_path": json_path,
        "md_path": md_path,
        "summary": summary,
    }


def run_live_etabs_story_beam_batch_from_env() -> Mapping[str, object]:
    _require_env("TBDY_RUN_LIVE_ETABS_SMOKE", "1", "env_gate")
    _require_env("TBDY_LIVE_ETABS_COM_PROVIDER", "1", "env_gate")
    _require_env("TBDY_LIVE_ETABS_USE_OPEN_MODEL", "1", "env_gate")

    story = os.environ.get("TBDY_LIVE_ETABS_STORY")
    if not story:
        raise StoryBeamBatchError("story_beam_discovery", "TBDY_LIVE_ETABS_STORY is required")

    combos_raw = os.environ.get("TBDY_LIVE_ETABS_COMBOS")
    if not combos_raw:
        raise StoryBeamBatchError("force_extract", "TBDY_LIVE_ETABS_COMBOS is required")
    combos = [combo.strip() for combo in combos_raw.split(",") if combo.strip()]
    min_beams = int(os.environ.get("TBDY_LIVE_ETABS_MIN_BEAMS", "3"))
    max_beams = _optional_int_env("TBDY_LIVE_ETABS_MAX_BEAMS")
    output_dir = Path(os.environ.get("TBDY_LIVE_ETABS_OUTPUT_DIR", "_local/live_etabs_story_beam_batch"))

    return run_live_etabs_story_beam_batch(
        story=story,
        combos=combos,
        output_dir=output_dir,
        min_beams=min_beams,
        max_beams=max_beams,
    )


def discover_story_beams(*, sap_model: object, story: str) -> list[dict[str, str]]:
    names = _frame_names(sap_model)
    beams: list[dict[str, str]] = []
    for object_name in names:
        label, object_story = _frame_label_and_story(sap_model, object_name)
        if object_story != story:
            continue
        section = _frame_section(sap_model, object_name)
        if not section:
            raise StoryBeamBatchError("selected_beam_lookup", f"section not found for frame object {object_name}")
        beams.append(
            {
                "object_name": object_name,
                "label": label or object_name,
                "story": object_story,
                "section": section,
            }
        )
    return beams


def _frame_names(sap_model: object) -> list[str]:
    frame_obj = getattr(sap_model, "FrameObj", None)
    if frame_obj is None or not hasattr(frame_obj, "GetNameList"):
        raise StoryBeamBatchError("story_beam_discovery", "SapModel.FrameObj.GetNameList unavailable")

    try:
        raw = frame_obj.GetNameList()
    except Exception as exc:
        raise StoryBeamBatchError("story_beam_discovery", f"GetNameList failed: {exc}") from exc

    if isinstance(raw, (list, tuple)):
        for item in raw:
            if (
                isinstance(item, (list, tuple))
                and item
                and all(isinstance(value, str) for value in item)
            ):
                return [str(value) for value in item]

    if isinstance(raw, (list, tuple)) and raw and all(isinstance(value, str) for value in raw):
        return [str(value) for value in raw]

    raise StoryBeamBatchError("story_beam_discovery", "unsupported GetNameList return shape")


def _frame_label_and_story(sap_model: object, object_name: str) -> tuple[str | None, str | None]:
    frame_obj = getattr(sap_model, "FrameObj", None)
    if frame_obj is None or not hasattr(frame_obj, "GetLabelFromName"):
        return (None, None)
    try:
        raw = frame_obj.GetLabelFromName(object_name)
    except Exception:
        return (None, None)

    values = raw if isinstance(raw, (list, tuple)) else (raw,)
    strings = [value for value in values if isinstance(value, str) and value]

    if len(strings) >= 2:
        return (strings[0], strings[1])
    if len(strings) == 1:
        story = _frame_story(sap_model, object_name)
        return (strings[0], story)

    return (None, None)



def _raw_signed_evidence_rows(governing: Mapping[str, object]) -> list[dict[str, object]]:
    rows: list[dict[str, object]] = []
    for action, evidence in governing.items():
        if not isinstance(evidence, Mapping):
            continue
        if "etabs_raw_signed_value" not in evidence and "design_demand_magnitude" not in evidence:
            continue
        rows.append(
            {
                "action": action,
                "combo": evidence.get("combo"),
                "station": evidence.get("station"),
                "etabs_local_axis_component": evidence.get("etabs_local_axis_component"),
                "etabs_raw_signed_value": evidence.get("etabs_raw_signed_value"),
                "design_demand_magnitude": evidence.get("design_demand_magnitude"),
                "sign_convention": evidence.get("sign_convention"),
            }
        )
    return rows


def _write_raw_signed_evidence_sheet(xlsx_path: Path, governing: Mapping[str, object]) -> None:
    rows = _raw_signed_evidence_rows(governing)
    if not rows:
        return

    try:
        from openpyxl import load_workbook
    except Exception:
        return

    workbook = load_workbook(xlsx_path)
    sheet_name = "ETABS_Raw_Evidence"

    if sheet_name in workbook.sheetnames:
        del workbook[sheet_name]

    worksheet = workbook.create_sheet(sheet_name)
    headers = [
        "action",
        "combo",
        "station",
        "etabs_local_axis_component",
        "etabs_raw_signed_value",
        "design_demand_magnitude",
        "sign_convention",
    ]
    worksheet.append(headers)

    for row in rows:
        worksheet.append([row.get(header) for header in headers])

    workbook.save(xlsx_path)
def _capacity_design_check_statuses(check_types: Sequence[str]) -> dict[str, str]:
    check_set = set(check_types)
    return {
        "beam_shear_capacity_design_ve_le_vr": "executed" if "beam_shear_capacity_design_ve_le_vr" in check_set else "missing",
        "beam_shear_capacity_design_ve_le_085_vmax": "executed" if "beam_shear_capacity_design_ve_le_085_vmax" in check_set else "missing",
    }


def _render_markdown_summary(summary: Mapping[str, object]) -> str:
    lines = [
        "# Live ETABS Story Beam Batch FrameForce Summary",
        "",
        "BeamCore checks executed",
        "BeamCoreResult produced",
        "ETABS FrameForce bridge observed",
        "Selected story batch smoke observed",
        "",
        f"- selected story: {summary['selected_story']}",
        f"- selected combos: {', '.join(summary['selected_combos'])}",
        "- units: kN / kNm / mm",
        f"- min beam requirement: {summary['min_beam_requirement']}",
        f"- beam count discovered: {summary['beam_count_discovered']}",
        f"- beam count processed: {summary['beam_count_processed']}",
        f"- beam count failed: {summary['beam_count_failed']}",
        "- ACTIONS_SOURCE = ETABS_RESULTS",
        f"- envelope rules: {summary['envelope_rules_document']}",
        "",
        "| Object | Label | Story | Section | Vd_left_kN | Ve_left_kN | Md_left_neg_kNm | Md_mid_pos_kNm | Md_right_neg_kNm | axial_kN | BeamCore status |",
        "|---|---|---|---|---:|---:|---:|---:|---:|---:|---|",
    ]
    for beam in summary["beams"]:
        actions = beam["actions"]
        lines.append(
            f"| {beam['object_name']} | {beam['label']} | {beam['story']} | {beam['section']} | "
            f"{actions['Vd_left_kN']} | {actions['Ve_left_kN']} | {actions['Md_left_neg_kNm']} | "
            f"{actions['Md_mid_pos_kNm']} | {actions['Md_right_neg_kNm']} | {actions['axial_kN']} | {beam['beam_core_status']} |"
        )
    lines.extend(["", "## Failures", ""])
    if summary["failures"]:
        for failure in summary["failures"]:
            lines.append(f"- {failure['object_name']}: {failure['stage']} — {failure['error']}")
    else:
        lines.append("- none")
    lines.extend(["", "## Forbidden claims", ""])
    for claim in summary["forbidden_claims"]:
        lines.append(f"- {claim}")
    lines.append("")
    return "\n".join(lines)


def _require_env(name: str, expected: str, stage: str) -> None:
    if os.environ.get(name) != expected:
        raise StoryBeamBatchError(stage, f"{name}={expected} required")


def _optional_int_env(name: str) -> int | None:
    value = os.environ.get(name)
    if value in (None, ""):
        return None
    return int(value)


def _safe_name(name: str) -> str:
    return "".join(ch if ch.isalnum() or ch in ("-", "_") else "_" for ch in name)


def _git(args: list[str]) -> str | None:
    try:
        return subprocess.check_output(["git", *args], text=True).strip()
    except Exception:
        return None


__all__ = [
    "StoryBeamBatchError",
    "discover_story_beams",
    "run_live_etabs_story_beam_batch",
    "run_live_etabs_story_beam_batch_from_env",
]
